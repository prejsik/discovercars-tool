import json
import sys
import tempfile
import zipfile
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from xml.etree import ElementTree
from zoneinfo import ZoneInfo

import openpyxl
from openpyxl.styles import PatternFill

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from tools.update_excel_rates import (  # noqa: E402
    add_calendar_months,
    apply_updates,
    build_change_statistics,
    build_targets,
    build_validation_rows,
    ensure_fixed_rate_group_rows,
    find_fixed_rate_group_issues,
    get_duration_columns,
    get_delta_fill,
    get_import_row_limit,
    get_minimum_rate,
    load_baseline_confirmation,
    load_config,
    merge_config,
    parse_date_value,
    parse_number,
    target_matches_recommendation_types,
)


def assert_equal(actual, expected, message):
    if actual != expected:
        raise AssertionError(f"{message}: expected {expected!r}, got {actual!r}")


def assert_not_equal(actual, expected, message):
    if actual == expected:
        raise AssertionError(f"{message}: expected value different than {expected!r}")


def rgb(cell):
    return str(cell.fill.fgColor.rgb)[-6:]


def header_rows_snapshot(ws, rows=4):
    return {
        "row_heights": [ws.row_dimensions[row].height for row in range(1, rows + 1)],
        "merged_ranges": sorted(str(item) for item in ws.merged_cells.ranges),
        "cells": [
            [
                (
                    ws.cell(row, col).value,
                    str(ws.cell(row, col)._style),
                    ws.cell(row, col).number_format,
                )
                for col in range(1, ws.max_column + 1)
            ]
            for row in range(1, rows + 1)
        ],
    }


def workbook_zones(path):
    namespace = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    zones = set()
    with zipfile.ZipFile(path) as archive:
        shared_strings = []
        if "xl/sharedStrings.xml" in archive.namelist():
            shared_root = ElementTree.fromstring(archive.read("xl/sharedStrings.xml"))
            for item in shared_root.findall("x:si", namespace):
                shared_strings.append("".join(text.text or "" for text in item.findall(".//x:t", namespace)))

        sheet_root = ElementTree.fromstring(archive.read("xl/worksheets/sheet1.xml"))
        for row in sheet_root.findall(".//x:sheetData/x:row", namespace):
            if int(row.attrib.get("r", "0")) < 5:
                continue
            for cell in row.findall("x:c", namespace):
                reference = cell.attrib.get("r", "")
                column = "".join(char for char in reference if char.isalpha())
                if column != "D":
                    continue
                value_node = cell.find("x:v", namespace)
                if value_node is None or value_node.text is None:
                    continue
                value = value_node.text
                if cell.attrib.get("t") == "s":
                    value = shared_strings[int(value)]
                if str(value).strip():
                    zones.add(str(value).strip().upper())
    return zones


def build_workbook(path):
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "Sheet1"
    ws.append(["Rental rates for packages: INCLUSIVE FP"])
    ws.append(["Min days", None, None, None, "Date format:", None, None, None, 1, 2, 3, 5, 8, 21])
    ws.append(["Max days", None, None, None, "dd-MM-yy", None, None, None, 1, 2, 4, 7, 20, 35])
    ws.append([
        "Group",
        "Description",
        "Rate code",
        "Zone",
        "Booking start date",
        "Booking end date",
        "Pickup start date",
        "Pickup end date",
        "Per day",
        "Per day",
        "Per day",
        "Per day",
        "Per day",
        "Per day",
    ])
    rows = [
        ["CDMV", None, None, "WA1", "09-07-26", "10-07-26", "10-07-26", "11-07-26", 160, 70, 80, 90, 100, 120],
        ["CGAV", None, None, "WA1", "09-07-26", "10-07-26", "10-07-26", "11-07-26", 160, 70, 80, 90, 100, 120],
        ["CWAV", None, None, "WA1", "09-07-26", "10-07-26", "10-07-26", "11-07-26", 160, 70, 80, 90, 100, 120],
        ["EDMV", None, None, "WA1", "09-07-26", "10-07-26", "10-07-26", "11-07-26", 160, 70, 80, 90, 100, 120],
        ["FVMD", None, None, "WA1", "09-07-26", "10-07-26", "10-07-26", "11-07-26", 160, 70, 80, 90, 100, 120],
        ["SWAV", None, None, "WA1", "09-07-26", "10-07-26", "10-07-26", "11-07-26", 160, 70, 80, 90, 100, 120],
        ["CDMV", None, None, "WA1", "09-07-26", "10-07-26", "11-07-26", "12-07-26", 160, 90, 80, 90, 100, 120],
        ["EDMV", None, None, "WA1", "09-07-26", "10-07-26", "11-07-26", "12-07-26", 160, 90, 80, 90, 100, 120],
        ["CDMV", None, None, "WA1", "09-07-26", "10-07-26", "25-07-26", "26-07-26", 160, 90, 80, 90, 130, 120],
        ["EDMV", None, None, "WA1", "09-07-26", "10-07-26", "25-07-26", "26-07-26", 160, 90, 80, 90, 130, 120],
    ]
    for row in rows:
        ws.append(row)
    ws["A4"].fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    workbook.save(path)


def build_minimal_workbook(path, rows):
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "Sheet1"
    ws.append(["Rental rates for packages: INCLUSIVE FP"])
    ws.append(["Min days", None, None, None, "Date format:", None, None, None, 1, 2, 3, 5, 8, 21])
    ws.append(["Max days", None, None, None, "dd-MM-yy", None, None, None, 1, 2, 4, 7, 20, 35])
    ws.append([
        "Group",
        "Description",
        "Rate code",
        "Zone",
        "Booking start date",
        "Booking end date",
        "Pickup start date",
        "Pickup end date",
        "Per day",
        "Per day",
        "Per day",
        "Per day",
        "Per day",
        "Per day",
    ])
    for row in rows:
        ws.append(row)
    workbook.save(path)


def main():
    assert_equal(add_calendar_months(date(2026, 8, 27), 4), date(2026, 12, 27), "four-month pickup horizon")
    assert_equal(add_calendar_months(date(2026, 10, 31), 4), date(2027, 2, 28), "month-end pickup horizon")
    assert_equal(
        build_change_statistics([
            {"delta": 10},
            {"delta": 20},
            {"delta": -5},
            {"delta": -15},
            {"delta": 0},
            {"delta": None},
        ]),
        {
            "increase_count": 2,
            "decrease_count": 2,
            "average_increase_pln_day": 15.0,
            "average_decrease_pln_day": -10.0,
        },
        "change statistics use every non-zero applied delta",
    )

    warning_fill = get_delta_fill(
        {"new_rate": 59, "delta": -10},
        merge_config({"changed_rate_warning": {"below_pln_day": 60, "color": "FFF2CC"}}),
    )
    assert_equal(str(warning_fill.fgColor.rgb)[-6:], "FFF2CC", "changed rate below warning threshold fill")

    example_config = load_config(ROOT / "excel-rate-update.config.example.json")
    assert_equal(
        example_config["excluded_groups"],
        ["CGAV", "FVMD", "SWAV", "CFAV", "PDAH"],
        "excluded current and fixed-rate groups",
    )
    assert_equal(
        example_config["fixed_rate_groups"],
        {
            "CFAV": {
                "template_group": "CDMV",
                "rates_by_duration_band": {
                    "1": 300,
                    "2": 200,
                    "3-4": 180,
                    "5-7": 170,
                    "8-20": 160,
                    "21-35": 150,
                },
            },
            "PDAH": {
                "template_group": "CDMV",
                "rates_by_duration_band": {
                    "1": 400,
                    "2": 350,
                    "3-4": 300,
                    "5-7": 290,
                    "8-20": 260,
                    "21-35": 250,
                },
            },
        },
        "fixed-rate group configuration",
    )
    assert_equal(
        example_config["mirrored_rate_groups"],
        {"EDAV": {"template_group": "EDMV"}},
        "EDAV mirrors EDMV",
    )
    assert_equal(example_config["max_import_rows"], 28000, "broker import row limit")
    assert_equal(get_import_row_limit({"max_import_rows": 30000}), 28000, "broker row limit cannot be raised")
    assert_equal(
        example_config["excluded_group_highlights"],
        {"CGAV": 130, "SWAV": 150},
        "current highlight-only groups",
    )
    assert_equal(
        example_config["group_price_parity"]["base_groups"],
        ["CDMV", "CWAV", "CWMR"],
        "current base parity groups",
    )
    assert_equal(
        example_config["group_price_parity"]["premium_adjustments_pln_day"],
        {"EDMV": 1, "EDAV": 1},
        "current premium groups",
    )
    assert_equal(
        example_config["protected_rate_periods"],
        [
            {"start_date": "2026-10-31", "end_date": "2026-11-02"},
            {"start_date": "2026-12-15", "end_date": "2027-01-10"},
        ],
        "holiday rate protection periods",
    )
    assert_equal(example_config["city_top1_airport_cap"]["max_multiplier"], 1.3, "city-airport cap")
    assert_equal(example_config["max_recommendation_duration_days"], 7, "maximum recommendation duration")
    september_floor, _ = get_minimum_rate(
        {"target_date": date(2026, 9, 1), "duration_min_days": 1, "duration_max_days": 35},
        example_config,
    )
    assert_equal(september_floor, 50, "floor from September 2026")
    august_gap_floor, _ = get_minimum_rate(
        {"target_date": date(2026, 8, 31), "duration_min_days": 1, "duration_max_days": 35},
        example_config,
    )
    assert_equal(august_gap_floor, 0, "no period floor on 31 August 2026")
    baseline_manifest = json.loads((ROOT / "input" / "baseline-manifest.json").read_text(encoding="utf-8"))
    baseline_confirmation = load_baseline_confirmation(example_config, baseline_manifest["workbook_sha256"])
    assert_equal(baseline_confirmation["status"], "confirmed_imported", "confirmed baseline status")
    assert_equal(baseline_confirmation["calibration_eligible"], True, "confirmed baseline calibration eligibility")
    city_cap_types = {"top1_gap", "force_top1_maintain"}
    assert target_matches_recommendation_types({"recommendation_type": "top1_gap"}, city_cap_types)
    assert target_matches_recommendation_types({"recommendation_type": "force_top1_maintain"}, city_cap_types)
    assert not target_matches_recommendation_types({"recommendation_type": "top1_undercut"}, city_cap_types)

    duration_targets, duration_skipped = build_targets(
        [{
            "action": "decrease",
            "recommendation_type": "top1_undercut",
            "location": "Warsaw Train Station",
            "start_date": "2026-09-20",
            "rental_days": 8,
            "suggested_rate_pln_day": 80,
        }],
        {8: (13, "8-20", 8, 20)},
        example_config,
    )
    assert_equal(bool(duration_targets), False, "duration 8 target is blocked")
    assert_equal(len(duration_skipped), 1, "duration 8 skip count")
    assert "Maximum recommendation duration is 7 days" in duration_skipped[0]["skip_reason"]

    with tempfile.TemporaryDirectory() as temporary_dir:
        temporary_path = Path(temporary_dir)

        fixed_workbook_path = temporary_path / "fixed-rate-groups.xlsx"
        fixed_recommendations_path = temporary_path / "fixed-rate-recommendations.json"
        fixed_output_path = temporary_path / "fixed-rate-groups-output.xlsx"
        build_minimal_workbook(
            fixed_workbook_path,
            [
                ["CDMV", None, None, "WA1", "09-06-26", "20-06-26", "20-06-26", "20-06-26", 160, 70, 80, 90, 100, 120],
                ["CDMV", None, None, "WA2", "09-06-26", "20-06-26", "20-06-26", "20-06-26", 160, 70, 80, 90, 100, 120],
                ["EDMV", None, None, "WA1", "09-06-26", "20-06-26", "20-06-26", "20-06-26", 161, 71, 81, 91, 101, 121],
                ["EDMV", None, None, "WA2", "09-06-26", "20-06-26", "20-06-26", "20-06-26", 162, 72, 82, 92, 102, 122],
                ["CFAV", None, None, "WA1", "09-06-26", "20-06-26", "20-06-26", "20-06-26", 1, 2, 3, 4, 5, 6],
                ["CFAV", None, None, "WA3", "09-06-26", "20-06-26", "20-06-26", "20-06-26", 1, 2, 3, 4, 5, 6],
            ],
        )
        fixed_recommendations_path.write_text(
            json.dumps({
                "recommendations": [{
                    "action": "increase",
                    "recommendation_type": "top1_undercut",
                    "location": "Warsaw Test",
                    "start_date": "2026-06-20",
                    "rental_days": 2,
                    "suggested_rate_pln_day": 180,
                    "benchmark_rate_pln_day": 181,
                }]
            }),
            encoding="utf-8",
        )
        fixed_config = merge_config({
            "excluded_groups": ["CGAV", "FVMD", "SWAV", "CFAV", "PDAH"],
            "fixed_rate_groups": {
                "CFAV": {
                    "template_group": "CDMV",
                    "rates_by_duration_band": {"1": 300, "2": 200, "3-4": 180, "5-7": 170, "8-20": 160, "21-35": 150},
                },
                "PDAH": {
                    "template_group": "CDMV",
                    "rates_by_duration_band": {"1": 400, "2": 350, "3-4": 300, "5-7": 290, "8-20": 260, "21-35": 250},
                },
            },
            "mirrored_rate_groups": {"EDAV": {"template_group": "EDMV"}},
            "protected_rate_periods": example_config["protected_rate_periods"],
            "location_zones": {"Warsaw Test": ["WA1"]},
            "pickup_date_expansion": {"enabled": False},
        })
        fixed_dry_summary = apply_updates(
            workbook_path=fixed_workbook_path,
            recommendations_path=fixed_recommendations_path,
            output_path=None,
            config=fixed_config,
            cli_groups=None,
            dry_run=True,
        )
        fixed_dry_validation = {
            item["check"]: item["status"]
            for item in fixed_dry_summary["validation"]
        }
        assert_equal(
            fixed_dry_validation["Kompletne klasy ze stawkami stalymi"],
            "OK",
            "fixed-rate dry-run validation",
        )
        fixed_summary = apply_updates(
            workbook_path=fixed_workbook_path,
            recommendations_path=fixed_recommendations_path,
            output_path=fixed_output_path,
            config=fixed_config,
            cli_groups=None,
            dry_run=False,
        )
        assert_equal(fixed_summary["fixed_rate_groups"]["added_row_count"], 3, "fixed rows added")
        assert_equal(fixed_summary["fixed_rate_groups"]["updated_row_count"], 2, "fixed rows updated")
        assert_equal(fixed_summary["mirrored_rate_groups"]["added_row_count"], 2, "EDAV rows added")
        assert not ({"CFAV", "PDAH"} & {str(change["group"]) for change in fixed_summary["changes"]})
        fixed_book = openpyxl.load_workbook(fixed_output_path)
        fixed_ws = fixed_book["Sheet1"]
        fixed_legend = " ".join(
            str(fixed_book["Changed Positions"].cell(row, 2).value or "")
            for row in range(1, fixed_book["Changed Positions"].max_row + 1)
        )
        assert "CFAV: 1=300 PLN" in fixed_legend
        assert "PDAH: 1=400 PLN" in fixed_legend
        assert "EDAV=EDMV" in fixed_legend
        assert "2026-10-31-2026-11-02" in fixed_legend
        assert "2026-12-15-2027-01-10" in fixed_legend
        assert_equal(fixed_ws.max_row, 15, "fixed and mirrored classes cover template locations")
        fixed_rows = {
            (fixed_ws.cell(row, 1).value, fixed_ws.cell(row, 4).value): row
            for row in range(5, fixed_ws.max_row + 1)
        }
        expected_fixed_rates = {
            "CFAV": [300, 200, 180, 170, 160, 150],
            "PDAH": [400, 350, 300, 290, 260, 250],
        }
        for group, expected_rates in expected_fixed_rates.items():
            for zone in ("WA1", "WA2"):
                row = fixed_rows[(group, zone)]
                assert_equal(
                    [fixed_ws.cell(row, col).value for col in range(9, 15)],
                    expected_rates,
                    f"{group}/{zone} fixed rates",
                )
        assert_equal(
            [fixed_ws.cell(fixed_rows[("CFAV", "WA3")], col).value for col in range(9, 15)],
            expected_fixed_rates["CFAV"],
            "extra CFAV row fixed rates",
        )
        for zone in ("WA1", "WA2"):
            assert_equal(
                [fixed_ws.cell(fixed_rows[("EDAV", zone)], col).value for col in range(9, 15)],
                [fixed_ws.cell(fixed_rows[("EDMV", zone)], col).value for col in range(9, 15)],
                f"EDAV mirrors EDMV for {zone}",
            )
        assert_equal(
            find_fixed_rate_group_issues(fixed_ws, fixed_config, get_duration_columns(fixed_ws, fixed_config)),
            [],
            "fixed-rate group validation",
        )
        fixed_validation = fixed_book["Validation"]
        fixed_validation_rows = {
            fixed_validation.cell(row, 1).value: fixed_validation.cell(row, 2).value
            for row in range(2, fixed_validation.max_row + 1)
        }
        assert_equal(
            fixed_validation_rows["Kompletne klasy ze stawkami stalymi"],
            "OK",
            "fixed-rate group workbook validation",
        )
        second_pass = ensure_fixed_rate_group_rows(
            fixed_ws,
            fixed_config,
            get_duration_columns(fixed_ws, fixed_config),
            dry_run=False,
        )
        assert_equal(second_pass["added_row_count"], 0, "fixed rows are idempotent")
        assert_equal(second_pass["updated_row_count"], 0, "fixed rates are idempotent")

        holiday_workbook_path = temporary_path / "holiday-protection.xlsx"
        holiday_recommendations_path = temporary_path / "holiday-protection-recommendations.json"
        holiday_output_path = temporary_path / "holiday-protection-output.xlsx"
        holiday_dates = [
            "30-10-26",
            "31-10-26",
            "02-11-26",
            "03-11-26",
            "14-12-26",
            "15-12-26",
            "10-01-27",
            "11-01-27",
        ]
        holiday_group_rates = {
            "CDMV": [100, 101, 102, 103, 104, 105],
            "CWAV": [110, 111, 112, 113, 114, 115],
            "CWMR": [120, 121, 122, 123, 124, 125],
            "EDMV": [130, 131, 132, 133, 134, 135],
            "EDAV": [140, 141, 142, 143, 144, 145],
            "CFAV": [150, 151, 152, 153, 154, 155],
            "PDAH": [160, 161, 162, 163, 164, 165],
        }
        build_minimal_workbook(
            holiday_workbook_path,
            [
                [group, None, None, "WA1", "09-06-26", pickup_date, pickup_date, pickup_date, *rates]
                for pickup_date in holiday_dates
                for group, rates in holiday_group_rates.items()
            ],
        )
        holiday_recommendations = [
            {
                "action": "increase",
                "recommendation_type": "top1_gap",
                "location": "Warsaw Test",
                "start_date": parse_date_value(pickup_date).isoformat(),
                "rental_days": 2,
                "suggested_rate_pln_day": 222,
                "benchmark_rate_pln_day": 223,
            }
            for pickup_date in holiday_dates
        ]
        holiday_recommendations_path.write_text(
            json.dumps({"recommendations": holiday_recommendations}),
            encoding="utf-8",
        )
        holiday_config = merge_config({
            "excluded_groups": ["CGAV", "FVMD", "SWAV", "CFAV", "PDAH"],
            "fixed_rate_groups": example_config["fixed_rate_groups"],
            "mirrored_rate_groups": example_config["mirrored_rate_groups"],
            "group_rate_adjustments_pln_day": example_config["group_rate_adjustments_pln_day"],
            "group_price_parity": example_config["group_price_parity"],
            "protected_rate_periods": example_config["protected_rate_periods"],
            "location_zones": {"Warsaw Test": ["WA1"]},
            "pickup_date_expansion": {"enabled": False},
        })
        holiday_summary = apply_updates(
            workbook_path=holiday_workbook_path,
            recommendations_path=holiday_recommendations_path,
            output_path=holiday_output_path,
            config=holiday_config,
            cli_groups=None,
            dry_run=False,
        )
        assert_equal(holiday_summary["skipped_target_count"], 4, "four holiday recommendations skipped")
        holiday_book = openpyxl.load_workbook(holiday_output_path)
        holiday_ws = holiday_book["Sheet1"]
        holiday_rows = {
            (
                str(holiday_ws.cell(row, 1).value or "").strip().upper(),
                parse_date_value(holiday_ws.cell(row, 7).value),
            ): row
            for row in range(5, holiday_ws.max_row + 1)
        }
        protected_dates = {
            date(2026, 10, 31),
            date(2026, 11, 2),
            date(2026, 12, 15),
            date(2027, 1, 10),
        }
        for pickup_date in protected_dates:
            for group, original_rates in holiday_group_rates.items():
                row = holiday_rows[(group, pickup_date)]
                assert_equal(
                    [holiday_ws.cell(row, col).value for col in range(9, 15)],
                    original_rates,
                    f"protected holiday rates for {group}/{pickup_date.isoformat()}",
                )
        unprotected_dates = {
            date(2026, 10, 30),
            date(2026, 11, 3),
            date(2026, 12, 14),
            date(2027, 1, 11),
        }
        for pickup_date in unprotected_dates:
            for group, expected_rate in (("CDMV", 222), ("CWAV", 222), ("CWMR", 222), ("EDMV", 223), ("EDAV", 223)):
                row = holiday_rows[(group, pickup_date)]
                assert_equal(holiday_ws.cell(row, 10).value, expected_rate, f"unprotected rate for {group}/{pickup_date.isoformat()}")
            for group, expected_rates in expected_fixed_rates.items():
                row = holiday_rows[(group, pickup_date)]
                assert_equal(
                    [holiday_ws.cell(row, col).value for col in range(9, 15)],
                    expected_rates,
                    f"unprotected fixed rates for {group}/{pickup_date.isoformat()}",
                )
        holiday_book.close()

        pending_manifest = {
            "schema_version": 1,
            "status": "prepared",
            "workbook_sha256": baseline_manifest["workbook_sha256"],
        }
        (temporary_path / "baseline.json").write_text(json.dumps(pending_manifest), encoding="utf-8")
        pending_config = merge_config({"baseline_manifest_file": "baseline.json"})
        pending_config["_config_dir"] = str(temporary_path)
        try:
            load_baseline_confirmation(pending_config, baseline_manifest["workbook_sha256"])
        except ValueError as error:
            assert "not confirmed as imported" in str(error)
        else:
            raise AssertionError("prepared baseline should not be accepted")

    location_zones = {
        str(location): {str(zone).upper() for zone in zones}
        for location, zones in example_config["location_zones"].items()
    }
    expected_location_zones = {
        "Bydgoszcz Airport (BZG)": {"BYLO"},
        "Gdansk Downtown": {"GD1"},
        "Gdansk Airport (GDN)": {"GDLO"},
        "Katowice Downtown": {"KA1"},
        "Katowice Airport (KTW)": {"KALO"},
        "Krakow Train Station": {"KRDW"},
        "Galeria Krakowska Shopping Mall": {"KRGA"},
        "Krakow Airport (KRK)": {"KRLO", "KRTI"},
        "Lodz Downtown": {"LO1"},
        "Lodz Lublinek Airport (LCJ)": {"LOLO"},
        "Lubin Downtown": {"LU1"},
        "Olsztyn Downtown": {"OL1"},
        "Opole Downtown": {"OP1"},
        "Poznan Downtown": {"PO1"},
        "Poznan Airport (POZ)": {"POLO"},
        "Torun Downtown": {"TO1"},
        "Warsaw West Train Station": {"WA1"},
        "Warsaw Train Station": {"WA2"},
        "Warsaw Chopin Airport (WAW)": {"WALO"},
        "Wroclaw Downtown": {"WR1"},
        "Wroclaw Airport (WRO)": {"WRLO"},
    }
    for location, zones in expected_location_zones.items():
        assert_equal(location_zones.get(location), zones, f"location zone mapping for {location}")
    assert_equal(example_config["city_zone_airport_zones"]["WA1"], ["WALO"], "Warsaw West city-airport mapping")
    assert_equal(example_config["city_zone_airport_zones"]["WA2"], ["WALO"], "Warsaw city-airport mapping")
    assert_equal(example_config["city_zone_airport_zones"]["KRDW"], ["KRLO"], "Krakow station-airport mapping")
    assert_equal(example_config["city_zone_airport_zones"]["KRGA"], ["KRLO"], "Krakow gallery-airport mapping")
    assert "KRTI" not in example_config["city_zone_airport_zones"]

    covered_zones = set().union(*location_zones.values())
    real_zones = workbook_zones(ROOT / "input" / "mm-cars-rental-rates-inclusive-fp.xlsx")
    assert_equal(sorted(real_zones - covered_zones), [], "all real workbook zones are covered by location_zones")

    with tempfile.TemporaryDirectory() as tmp:
        tmpdir = Path(tmp)
        workbook_path = tmpdir / "rates.xlsx"
        recommendations_path = tmpdir / "pricing-recommendations.json"
        output_path = tmpdir / "rates-updated.xlsx"
        import_output_path = tmpdir / "rates-import-ready.xlsx"
        build_workbook(workbook_path)

        recommendations_path.write_text(
            json.dumps(
                {
                    "recommendations": [
                        {
                            "action": "increase",
                            "recommendation_type": "top1_gap",
                            "reason": "MM Cars Rental jest top1, a top2 jest drozszy o co najmniej 10 PLN/dzien; cel to 1 PLN ponizej top2.",
                            "location": "Warsaw",
                            "start_date": "2026-07-10",
                            "rental_days": 2,
                            "suggested_rate_pln_day": 81,
                            "mm_rate_pln_day": 70,
                            "benchmark_provider": "Flex To Go",
                            "benchmark_rate_pln_day": 82,
                            "scenario_id": "2026-07-10-2",
                        },
                        {
                            "action": "decrease",
                            "recommendation_type": "top1_undercut",
                            "reason": "MM Cars Rental jest top2 i brakuje mniej niz 10 PLN/dzien, zeby zostac top1; cel to 1 PLN ponizej top1.",
                            "location": "Warsaw",
                            "start_date": "2026-07-10",
                            "rental_days": 21,
                            "suggested_rate_pln_day": 80,
                            "mm_rate_pln_day": 120,
                            "benchmark_provider": "Flex To Go",
                            "benchmark_rate_pln_day": 101,
                            "scenario_id": "2026-07-10-21",
                        },
                        {
                            "action": "decrease",
                            "recommendation_type": "top1_undercut",
                            "reason": "MM Cars Rental jest top2 i brakuje mniej niz 10 PLN/dzien, zeby zostac top1; cel to 1 PLN ponizej top1.",
                            "location": "Warsaw",
                            "start_date": "2026-07-11",
                            "rental_days": 2,
                            "suggested_rate_pln_day": 60,
                            "mm_rate_pln_day": 90,
                            "benchmark_provider": "Car24",
                            "benchmark_rate_pln_day": 61,
                            "scenario_id": "2026-07-11-2",
                        },
                        {
                            "action": "decrease",
                            "recommendation_type": "top3_small_decrease",
                            "reason": "Cel top3 wymaga roznicy mniejszej niz 10 PLN/dzien; cel to 1 PLN ponizej top3.",
                            "location": "Warsaw",
                            "start_date": "2026-07-25",
                            "rental_days": 8,
                            "suggested_rate_pln_day": 90,
                            "mm_rate_pln_day": 120,
                            "benchmark_provider": "Kaizen Rent",
                            "benchmark_rate_pln_day": 91,
                            "scenario_id": "2026-07-25-8",
                        }
                    ]
                }
            ),
            encoding="utf-8",
        )

        config = merge_config(
            {
                "location_zones": {"Warsaw": ["WA1"]},
                "max_recommendation_duration_days": 35,
            }
        )

        summary = apply_updates(
            workbook_path=workbook_path,
            recommendations_path=recommendations_path,
            output_path=output_path,
            config=config,
            cli_groups=None,
            dry_run=False,
            import_output_path=import_output_path,
        )

        limited_output_path = tmpdir / "rates-over-limit.xlsx"
        limited_import_path = tmpdir / "rates-import-over-limit.xlsx"
        try:
            apply_updates(
                workbook_path=workbook_path,
                recommendations_path=recommendations_path,
                output_path=limited_output_path,
                config=merge_config({
                    "location_zones": {"Warsaw": ["WA1"]},
                    "max_recommendation_duration_days": 35,
                    "max_import_rows": 13,
                }),
                cli_groups=None,
                dry_run=False,
                import_output_path=limited_import_path,
            )
            raise AssertionError("workbook above the broker row limit should fail")
        except ValueError as error:
            assert "row limit exceeded" in str(error)
        assert_equal(limited_output_path.exists(), False, "over-limit recommendation workbook is not saved")
        assert_equal(limited_import_path.exists(), False, "over-limit import workbook is not saved")

        assert_equal(summary["change_count"], 10, "change_count")
        assert_equal(summary["group_price_parity_change_count"], 0, "group_price_parity_change_count")
        assert_equal(summary["group_price_parity_scope_count"], 4, "group_price_parity_scope_count")
        assert_equal(summary["import_output"], str(import_output_path), "import output path")
        assert_equal(summary["max_import_rows"], 28000, "summary import row limit")
        assert_equal(summary["import_row_count"], 14, "summary import row count")
        assert_equal(
            summary["recommendation_date_scope"],
            {"start_date": "2026-07-10", "end_date": "2026-07-25", "date_count": 3},
            "summary recommendation date scope",
        )
        assert_equal(summary["normalized_pickup_end_count"], 10, "normalized_pickup_end_count")
        assert_equal(summary["synced_booking_end_count"], 4, "synced_booking_end_count")
        updated = openpyxl.load_workbook(output_path)
        ws = updated["Sheet1"]
        changed_ws = updated["Changed Positions"]
        assert_equal(
            updated.sheetnames,
            ["Sheet1", "Changed Positions", "Recommendations Review", "Validation"],
            "workbook sheets",
        )
        assert str(ws["A4"].fill.fgColor.rgb).endswith("1F4E78")
        review_ws = updated["Recommendations Review"]
        validation_ws = updated["Validation"]
        import_ready = openpyxl.load_workbook(import_output_path)
        assert_equal(import_ready.sheetnames, ["Sheet1"], "import-ready workbook sheets")
        import_ready_ws = import_ready["Sheet1"]
        assert_equal(import_ready_ws["J5"].value, 81, "import-ready updated rate")
        assert_equal(import_ready_ws["N5"].value, 100, "import-ready long duration minimum")
        assert_equal(ws.max_row, 14, "main import sheet row count")
        assert_equal(ws["J5"].value, 81, "updated rate")
        assert_equal(ws["J6"].value, 70, "excluded CGAV rate")
        assert_equal(ws["J7"].value, 81, "CWAV parity rate")
        assert_equal(ws["J8"].value, 82, "EDMV adjusted rate")
        assert_equal(ws["J9"].value, 70, "excluded FVMD rate")
        assert_equal(ws["J10"].value, 70, "excluded SWAV rate")
        assert_equal(ws["I7"].value, 160, "CWAV duration 1 rate outside recommendations is unchanged")
        assert_equal(ws["I8"].value, 160, "EDMV duration 1 rate outside recommendations is unchanged")
        assert_equal(ws["I9"].value, 160, "excluded FVMD duration 1 rate")
        assert_equal(ws["K7"].value, 80, "CWAV duration 3-4 rate outside recommendations is unchanged")
        assert_equal(ws["K8"].value, 80, "EDMV duration 3-4 rate outside recommendations is unchanged")
        assert_equal(ws["N5"].value, 100, "long duration minimum")
        assert_equal(ws["N7"].value, 100, "long duration minimum for CWAV")
        assert_equal(ws["N8"].value, 101, "long duration minimum with EDMV adjustment")
        assert_equal(ws["N9"].value, 120, "excluded FVMD long duration rate")
        assert_equal(ws["J11"].value, 70, "global minimum")
        assert_equal(ws["J12"].value, 71, "global minimum with EDMV adjustment")
        assert_equal(ws["M13"].value, 115, "seasonal duration minimum")
        assert_equal(ws["M14"].value, 116, "seasonal duration minimum with EDMV adjustment")
        assert_equal(ws["H5"].value, ws["G5"].value, "pickup end normalized for CDMV")
        assert_equal(ws["H6"].value, ws["G6"].value, "pickup end normalized for excluded CGAV")
        for row in range(5, 15):
            assert_equal(ws.cell(row, 6).value, ws.cell(row, 8).value, f"booking end equals pickup end in row {row}")
        assert_not_equal(rgb(ws["J5"]), "C6EFCE", "increase color uses dynamic scale")
        assert_equal(rgb(ws["J5"]), "A9D18E", "increase color uses a stepped green scale")
        assert_not_equal(rgb(ws["J11"]), rgb(ws["M13"]), "larger decrease uses a stronger red")
        assert ws["J5"].comment is not None
        assert_equal(
            ws["J5"].comment.text,
            "Poprzednia stawka: 70 PLN\nNowa stawka: 81 PLN\nZmiana: +11 PLN\nCel na stronie: 81 PLN\nPrognoza na stronie: 81 PLN",
            "short Sheet1 comment",
        )
        assert ws["J8"].comment is not None
        assert_equal(
            ws["J8"].comment.text,
            "Poprzednia stawka: 70 PLN\nNowa stawka: 82 PLN\nZmiana: +12 PLN\nCel na stronie: 81 PLN\nPrognoza na stronie: 82 PLN\nCel rankingowy: wymaga kontroli",
            "short adjusted Sheet1 comment",
        )
        assert "brutto/dzien" not in ws["N5"].comment.text
        assert ws["J6"].comment is None
        assert_equal(changed_ws["A1"].value, "Legenda", "changed sheet legend title")
        assert_equal(changed_ws["A2"].value, "Top1 gap", "top1 legend label")
        assert "co najmniej 10 PLN" in changed_ws["B2"].value
        assert_equal(rgb(changed_ws["A2"]), "9DC3E6", "top1 legend color")
        assert_equal(changed_ws["A3"].value, "Male obnizenie top3", "top3 legend label")
        assert_equal(rgb(changed_ws["A3"]), "FFC7CE", "top3 legend color")
        assert_equal(changed_ws["A4"].value, "Przebicie top1", "top1 undercut legend label")
        assert_equal(rgb(changed_ws["A4"]), "F4B183", "top1 undercut legend color")
        assert_equal(changed_ws["A5"].value, "Scalanie duration", "duration aggregation legend label")
        assert_equal(changed_ws["A6"].value, "Kontrola celu", "target verification legend label")
        assert_equal(changed_ws["A9"].value, "Floor cenowy", "floor legend label")
        assert "Floor cenowy" in changed_ws["B9"].value
        assert_equal(changed_ws["O15"].value, "Komentarz zmiany", "changed sheet comment header")
        assert_equal(changed_ws.max_row, 20, "changed sheet row count")
        assert_equal(changed_ws["A16"].value, "CDMV, CWAV", "first changed group set")
        assert "Powod rekomendacji: MM Cars Rental jest na 1 miejscu" in changed_ws["O16"].value
        assert "co najmniej 10 PLN" in changed_ws["O16"].value
        assert "Co pozwoli osiagnac: utrzymanie top1" in changed_ws["O16"].value
        assert "Poprzednia stawka: 70 PLN" in changed_ws["O16"].value
        assert "Nowa stawka: 81 PLN" in changed_ws["O16"].value
        assert "Zmiana: +11 PLN" in changed_ws["O16"].value
        assert "EDMV: 82 PLN" not in changed_ws["O16"].value
        assert "brutto/dzien" not in changed_ws["O16"].value
        assert "Lokalizacja" not in changed_ws["O16"].value
        assert "Data odbioru" not in changed_ws["O16"].value
        assert "Duration" not in changed_ws["O16"].value
        assert "Korekta grupy" not in changed_ws["O16"].value
        assert "Komorka" not in changed_ws["O16"].value
        assert "Scenario" not in changed_ws["O16"].value
        assert "Zastosowane minimum" not in changed_ws["O16"].value
        assert_equal(changed_ws["J16"].value, 81, "grouped base rate cell")
        assert "Co pozwoli osiagnac: cel rankingowy nie jest gwarantowany" in changed_ws["O20"].value
        assert "Zastosowane minimum" not in changed_ws["O20"].value
        assert "Minimum sezonowe" not in changed_ws["O20"].value
        assert_equal(rgb(changed_ws["O16"]), "9DC3E6", "top1 gap row is blue")
        assert_equal(rgb(changed_ws["O17"]), "F4B183", "top1 undercut row is orange")
        assert_equal(rgb(changed_ws["O20"]), "FFC7CE", "top3 small decrease row is red")
        assert changed_ws["O16"].comment is not None
        for row in range(1, changed_ws.max_row + 1):
            for col in range(1, 15):
                assert changed_ws.cell(row, col).comment is None
        changed_groups = {changed_ws.cell(row, 1).value for row in range(16, changed_ws.max_row + 1)}
        assert "CGAV" not in ",".join(changed_groups)
        assert "FVMD" not in ",".join(changed_groups)
        assert "SWAV" not in ",".join(changed_groups)
        assert_equal(review_ws["A1"].value, "Akceptacja?", "review header")
        assert_equal(review_ws["B1"].value, "Status", "review status header")
        assert_equal(review_ws.max_row, 6, "review row count")
        assert_equal(review_ws["D2"].value, "Warsaw", "review location")
        assert_equal(review_ws["F2"].value, "CDMV, CWAV", "review grouped groups")
        assert review_ws["B2"].value in {"Gotowe", "Gotowe z uwaga", "Sprawdz"}
        assert review_ws["C2"].value == "OK"
        validation_rows = {
            validation_ws.cell(row, 1).value: validation_ws.cell(row, 2).value
            for row in range(2, validation_ws.max_row + 1)
        }
        assert_equal(validation_rows["Booking end date = Pickup end date"], "OK", "booking date validation")
        assert_equal(validation_rows["Pickup end date = Pickup start date"], "OK", "pickup date validation")
        assert_equal(validation_rows["Puste stawki w kolumnach duration"], "OK", "blank rate validation")
        assert_equal(validation_rows["Zmienione stawki ponizej floor cenowego"], "OK", "floor validation")
        parity_validation = build_validation_rows(
            ws,
            config,
            get_duration_columns(ws, config),
            [{"recommendation_type": "group_parity", "cell": "J5"}],
            [],
            {},
        )
        parity_validation_by_name = {row[0]: row for row in parity_validation}
        assert_equal(
            parity_validation_by_name["Zmienione rekomendacje bez ceny benchmarku"][2],
            0,
            "group parity does not require a competitor benchmark",
        )

        exact_location_workbook_path = tmpdir / "exact-location-rates.xlsx"
        exact_location_recommendations_path = tmpdir / "exact-location-recommendations.json"
        exact_location_output_path = tmpdir / "exact-location-rates-updated.xlsx"
        build_minimal_workbook(
            exact_location_workbook_path,
            [
                ["CDMV", None, None, "WA1", "09-06-26", "20-06-26", "20-06-26", "20-06-26", 160, 70, 80, 90, 100, 120],
                ["CDMV", None, None, "WA2", "09-06-26", "20-06-26", "20-06-26", "20-06-26", 160, 70, 80, 90, 100, 120],
            ],
        )
        exact_location_recommendations_path.write_text(
            json.dumps(
                {
                    "recommendations": [
                        {
                            "action": "increase",
                            "recommendation_type": "top1_gap",
                            "location": "Warsaw West Train Station",
                            "start_date": "2026-06-20",
                            "rental_days": 2,
                            "suggested_rate_pln_day": 90,
                            "mm_rate_pln_day": 70,
                            "benchmark_provider": "GO Rental Cars",
                            "benchmark_rate_pln_day": 91,
                            "scenario_id": "exact-location-2026-06-20-2",
                        }
                    ]
                }
            ),
            encoding="utf-8",
        )
        exact_location_summary = apply_updates(
            workbook_path=exact_location_workbook_path,
            recommendations_path=exact_location_recommendations_path,
            output_path=exact_location_output_path,
            config=merge_config(
                {
                    "location_zones": example_config["location_zones"],
                    "pickup_date_expansion": {"enabled": False},
                }
            ),
            cli_groups=None,
            dry_run=False,
        )
        exact_location_updated = openpyxl.load_workbook(exact_location_output_path)
        exact_location_ws = exact_location_updated["Sheet1"]
        assert_equal(exact_location_summary["change_count"], 1, "exact location updates only one zone")
        assert_equal(exact_location_ws["J5"].value, 90, "WA1 exact location update")
        assert_equal(exact_location_ws["J6"].value, 70, "WA2 is not changed by WA1 exact location")

        dedup_workbook_path = tmpdir / "dedup-rates.xlsx"
        dedup_recommendations_path = tmpdir / "dedup-recommendations.json"
        dedup_output_path = tmpdir / "dedup-rates-updated.xlsx"
        build_minimal_workbook(
            dedup_workbook_path,
            [
                ["CDMV", None, None, "WA1", "09-06-26", "10-06-26", "10-06-26", "11-06-26", 160, 70, 80, 90, 100, 120],
                ["MDMR", None, None, "WA1", "09-06-26", "10-06-26", "10-06-26", "11-06-26", 160, 70, 80, 90, 100, 120],
            ],
        )
        dedup_recommendations_path.write_text(
            json.dumps(
                {
                    "recommendations": [
                        {
                            "action": "increase",
                            "recommendation_type": "top1_gap",
                            "reason": "MM Cars Rental jest top1, a top2 jest drozszy o co najmniej 10 PLN/dzien; cel to 1 PLN ponizej top2.",
                            "location": "Warsaw",
                            "start_date": "2026-06-10",
                            "rental_days": 2,
                            "suggested_rate_pln_day": 81,
                            "mm_rate_pln_day": 70,
                            "benchmark_provider": "Flex To Go",
                            "benchmark_rate_pln_day": 82,
                            "scenario_id": "dedup-2026-06-10-2",
                        }
                    ]
                }
            ),
            encoding="utf-8",
        )
        dedup_summary = apply_updates(
            workbook_path=dedup_workbook_path,
            recommendations_path=dedup_recommendations_path,
            output_path=dedup_output_path,
            config=merge_config({"location_zones": {"Warsaw": ["WA1"]}}),
            cli_groups=None,
            dry_run=False,
        )
        assert_equal(dedup_summary["change_count"], 2, "deduplicated change count")
        dedup_updated = openpyxl.load_workbook(dedup_output_path)
        dedup_changed_ws = dedup_updated["Changed Positions"]
        assert_equal(dedup_changed_ws.max_row, 16, "deduplicated changed sheet row count")
        assert_equal(dedup_changed_ws["A16"].value, "CDMV, MDMR", "deduplicated group list")
        assert_equal(dedup_changed_ws["J16"].value, 81, "deduplicated changed rate cell")
        assert "70 PLN; 70 PLN" not in dedup_changed_ws["O16"].value
        assert "81 PLN; 81 PLN" not in dedup_changed_ws["O16"].value
        assert "+11 PLN; +11 PLN" not in dedup_changed_ws["O16"].value

        accepted_only_workbook_path = tmpdir / "accepted-only-rates.xlsx"
        accepted_only_recommendations_path = tmpdir / "accepted-only-recommendations.json"
        accepted_only_output_path = tmpdir / "accepted-only-rates-updated.xlsx"
        build_minimal_workbook(
            accepted_only_workbook_path,
            [
                ["CDMV", None, None, "WA1", "09-06-26", "10-06-26", "10-06-26", "11-06-26", 160, 70, 80, 90, 100, 120],
            ],
        )
        accepted_only_recommendations_path.write_text(
            json.dumps(
                {
                    "recommendations": [
                        {
                            "action": "increase",
                            "accepted": True,
                            "location": "Warsaw",
                            "start_date": "2026-06-10",
                            "rental_days": 2,
                            "suggested_rate_pln_day": 85,
                        },
                        {
                            "action": "increase",
                            "accepted": False,
                            "location": "Warsaw",
                            "start_date": "2026-06-10",
                            "rental_days": 3,
                            "suggested_rate_pln_day": 95,
                        },
                    ]
                }
            ),
            encoding="utf-8",
        )
        accepted_only_summary = apply_updates(
            workbook_path=accepted_only_workbook_path,
            recommendations_path=accepted_only_recommendations_path,
            output_path=accepted_only_output_path,
            config=merge_config({"location_zones": {"Warsaw": ["WA1"]}}),
            cli_groups=None,
            dry_run=False,
            accepted_only=True,
        )
        assert_equal(accepted_only_summary["accepted_target_count"], 1, "accepted-only target count")
        assert_equal(accepted_only_summary["filtered_unaccepted_target_count"], 1, "filtered target count")
        accepted_only_updated = openpyxl.load_workbook(accepted_only_output_path)
        accepted_only_ws = accepted_only_updated["Sheet1"]
        assert_equal(accepted_only_ws["J5"].value, 85, "accepted recommendation applied")
        assert_equal(accepted_only_ws["K5"].value, 80, "unaccepted recommendation skipped")

        acceptance_review_path = tmpdir / "acceptance-review.xlsx"
        acceptance_review = openpyxl.Workbook()
        acceptance_review_ws = acceptance_review.active
        acceptance_review_ws.title = "Recommendations Review"
        acceptance_review_ws.append(["Akceptacja?", "Lokalizacja", "Data odbioru", "Przedzial duration", "ID scenariusza"])
        acceptance_review_ws.append(["YES", "Warsaw", "2026-06-10", "2", ""])
        acceptance_review.save(acceptance_review_path)

        acceptance_workbook_path = tmpdir / "acceptance-rates.xlsx"
        acceptance_recommendations_path = tmpdir / "acceptance-recommendations.json"
        acceptance_output_path = tmpdir / "acceptance-rates-updated.xlsx"
        build_minimal_workbook(
            acceptance_workbook_path,
            [
                ["CDMV", None, None, "WA1", "09-06-26", "10-06-26", "10-06-26", "11-06-26", 160, 70, 80, 90, 100, 120],
            ],
        )
        acceptance_recommendations_path.write_text(
            json.dumps(
                {
                    "recommendations": [
                        {
                            "action": "increase",
                            "location": "Warsaw",
                            "start_date": "2026-06-10",
                            "rental_days": 2,
                            "suggested_rate_pln_day": 86,
                        },
                        {
                            "action": "increase",
                            "location": "Warsaw",
                            "start_date": "2026-06-10",
                            "rental_days": 3,
                            "suggested_rate_pln_day": 96,
                        },
                    ]
                }
            ),
            encoding="utf-8",
        )
        acceptance_summary = apply_updates(
            workbook_path=acceptance_workbook_path,
            recommendations_path=acceptance_recommendations_path,
            output_path=acceptance_output_path,
            config=merge_config({"location_zones": {"Warsaw": ["WA1"]}}),
            cli_groups=None,
            dry_run=False,
            accepted_only=True,
            acceptance_workbook_path=acceptance_review_path,
        )
        assert_equal(acceptance_summary["accepted_target_count"], 1, "acceptance workbook target count")
        acceptance_updated = openpyxl.load_workbook(acceptance_output_path)
        acceptance_ws = acceptance_updated["Sheet1"]
        assert_equal(acceptance_ws["J5"].value, 86, "acceptance workbook recommendation applied")
        assert_equal(acceptance_ws["K5"].value, 80, "non-accepted workbook recommendation skipped")

        expansion_workbook_path = tmpdir / "expansion-rates.xlsx"
        expansion_recommendations_path = tmpdir / "expansion-recommendations.json"
        expansion_output_path = tmpdir / "expansion-rates-updated.xlsx"
        build_minimal_workbook(
            expansion_workbook_path,
            [
                ["CDMV", None, None, "WA1", "09-06-26", "10-06-26", "10-06-26", "10-06-26", 155, 65, 75, 85, 95, 115],
                ["CDMV", None, None, "WA1", "09-06-26", "10-06-26", "10-06-26", "12-06-26", 160, 70, 80, 90, 100, 120],
                ["CDMV", None, None, "WA1", "09-06-26", "13-06-26", "13-06-26", "13-06-26", 165, 75, 85, 95, 105, 125],
            ],
        )
        expansion_before = openpyxl.load_workbook(expansion_workbook_path)
        expansion_before_snapshot = header_rows_snapshot(expansion_before["Sheet1"])
        expansion_recommendations_path.write_text(
            json.dumps(
                {
                    "recommendations": [
                        {
                            "action": "decrease",
                            "recommendation_type": "top1_undercut",
                            "reason": "MM Cars Rental jest top2 i brakuje mniej niz 10 PLN/dzien, zeby zostac top1; cel to 1 PLN ponizej top1.",
                            "location": "Warsaw",
                            "start_date": "2026-06-11",
                            "rental_days": 1,
                            "suggested_rate_pln_day": 75,
                            "mm_rate_pln_day": 160,
                            "benchmark_provider": "Car24",
                            "benchmark_rate_pln_day": 76,
                            "scenario_id": "expansion-2026-06-11-1",
                        },
                        {
                            "action": "increase",
                            "recommendation_type": "top1_gap",
                            "reason": "MM Cars Rental jest top1, a top2 jest drozszy o co najmniej 10 PLN/dzien; cel to 1 PLN ponizej top2.",
                            "location": "Warsaw",
                            "start_date": "2026-06-11",
                            "rental_days": 2,
                            "suggested_rate_pln_day": 81,
                            "mm_rate_pln_day": 70,
                            "benchmark_provider": "Flex To Go",
                            "benchmark_rate_pln_day": 82,
                            "scenario_id": "expansion-2026-06-11-2",
                        },
                    ]
                }
            ),
            encoding="utf-8",
        )
        expansion_summary = apply_updates(
            workbook_path=expansion_workbook_path,
            recommendations_path=expansion_recommendations_path,
            output_path=expansion_output_path,
            config=merge_config(
                {
                    "location_zones": {"Warsaw": ["WA1"]},
                    "normalize_pickup_end_to_start": False,
                    "pickup_date_expansion": {
                        "enabled": True,
                        "start_date": "2026-06-11",
                        "end_date": "2026-06-12",
                        "drop_rows_before_start_date": True,
                        "drop_rows_after_end_date": True,
                        "time_zone": "Europe/Warsaw",
                    },
                }
            ),
            cli_groups=None,
            dry_run=False,
        )
        assert_equal(expansion_summary["pickup_date_expansion"]["source_row_count"], 1, "expanded source row count")
        assert_equal(expansion_summary["pickup_date_expansion"]["expanded_row_count"], 2, "expanded row count")
        assert_equal(expansion_summary["pickup_date_expansion"]["preserved_out_of_range_row_count"], 0, "preserved source row count")
        assert_equal(expansion_summary["pickup_date_expansion"]["dropped_before_start_row_count"], 2, "past row count")
        assert_equal(expansion_summary["pickup_date_expansion"]["dropped_after_end_row_count"], 1, "future row count")
        assert_equal(expansion_summary["pickup_date_expansion"]["output_row_count"], 2, "full output row count")
        assert_equal(expansion_summary["normalized_pickup_end_count"], 0, "expansion disables pickup end normalization")
        assert_equal(expansion_summary["synced_booking_end_count"], 2, "expanded booking end sync count")
        assert_equal(expansion_summary["change_count"], 2, "expanded duration-specific change count")
        expansion_updated = openpyxl.load_workbook(expansion_output_path)
        expansion_ws = expansion_updated["Sheet1"]
        expansion_after_snapshot = header_rows_snapshot(expansion_ws)
        assert_equal(expansion_after_snapshot, expansion_before_snapshot, "expanded Sheet1 rows 1-4 values and formatting")
        assert_equal(expansion_ws.max_row, 6, "expanded Sheet1 row count")
        assert_equal(expansion_ws["G5"].value, "11-06-26", "first expanded pickup date")
        assert_equal(expansion_ws["H5"].value, "11-06-26", "first expanded pickup end")
        assert_equal(expansion_ws["F5"].value, expansion_ws["H5"].value, "first expanded booking end")
        assert_equal(expansion_ws["G6"].value, "12-06-26", "second expanded pickup date")
        assert_equal(expansion_ws["H6"].value, "12-06-26", "second expanded pickup end")
        assert_equal(expansion_ws["F6"].value, expansion_ws["H6"].value, "unchanged date booking end")
        assert_equal(expansion_ws["I5"].value, 75, "duration 1 rate update")
        assert_equal(expansion_ws["J5"].value, 81, "duration 2 rate update on the same pickup date row")
        assert_equal(expansion_ws["I6"].value, 165, "duration 1 rate does not update a different pickup date")
        assert_equal(expansion_ws["J6"].value, 75, "duration 2 rate does not update a different pickup date")

        partial_range_workbook_path = tmpdir / "partial-range-rates.xlsx"
        partial_range_recommendations_path = tmpdir / "partial-range-recommendations.json"
        partial_range_output_path = tmpdir / "partial-range-updated.xlsx"
        partial_range_recommendations_path.write_text(json.dumps({"recommendations": []}), encoding="utf-8")
        build_minimal_workbook(
            partial_range_workbook_path,
            [["CDMV", None, None, "WA1", "09-06-26", "11-06-26", "11-06-26", "11-06-26", 155, 65, 75, 85, 95, 115]],
        )
        partial_range_summary = apply_updates(
            workbook_path=partial_range_workbook_path,
            recommendations_path=partial_range_recommendations_path,
            output_path=partial_range_output_path,
            config=merge_config(
                {
                    "pickup_date_expansion": {
                        "enabled": True,
                        "start_date": "2026-06-11",
                        "end_date": "2026-06-12",
                        "drop_rows_before_start_date": True,
                        "drop_rows_after_end_date": True,
                        "time_zone": "Europe/Warsaw",
                    },
                }
            ),
            cli_groups=None,
            dry_run=False,
        )
        assert_equal(
            partial_range_summary["pickup_date_expansion"]["appended_horizon_row_count"],
            1,
            "partial source range is extended through configured end date",
        )
        partial_range_ws = openpyxl.load_workbook(partial_range_output_path)["Sheet1"]
        assert_equal(partial_range_ws.max_row, 6, "partial source range output row count")
        assert_equal(partial_range_ws["G6"].value, "12-06-26", "partial source range final pickup date")
        assert_equal(partial_range_ws["J6"].value, 65, "partial source range carries the latest rate forward")

        internal_gap_workbook_path = tmpdir / "internal-gap-rates.xlsx"
        internal_gap_output_path = tmpdir / "internal-gap-updated.xlsx"
        build_minimal_workbook(
            internal_gap_workbook_path,
            [
                ["CDMV", None, None, "WA1", "09-06-26", "11-06-26", "11-06-26", "11-06-26", 155, 65, 75, 85, 95, 115],
                ["CDMV", None, None, "WA1", "09-06-26", "13-06-26", "13-06-26", "13-06-26", 165, 75, 85, 95, 105, 125],
            ],
        )
        apply_updates(
            workbook_path=internal_gap_workbook_path,
            recommendations_path=partial_range_recommendations_path,
            output_path=internal_gap_output_path,
            config=merge_config(
                {
                    "pickup_date_expansion": {
                        "enabled": True,
                        "start_date": "2026-06-11",
                        "end_date": "2026-06-14",
                        "drop_rows_before_start_date": True,
                        "drop_rows_after_end_date": True,
                        "time_zone": "Europe/Warsaw",
                    },
                }
            ),
            cli_groups=None,
            dry_run=False,
        )
        internal_gap_ws = openpyxl.load_workbook(internal_gap_output_path)["Sheet1"]
        assert_equal(
            [internal_gap_ws.cell(row, 7).value for row in range(5, internal_gap_ws.max_row + 1)],
            ["11-06-26", "13-06-26", "14-06-26"],
            "internal source gaps are preserved while the horizon is extended",
        )

        aggregate_workbook_path = tmpdir / "aggregate-duration-rates.xlsx"
        aggregate_recommendations_path = tmpdir / "aggregate-duration-recommendations.json"
        aggregate_output_path = tmpdir / "aggregate-duration-updated.xlsx"
        build_minimal_workbook(
            aggregate_workbook_path,
            [["CDMV", None, None, "WA1", "14-07-26", "15-07-26", "15-07-26", "15-07-26", 160, 100, 100, 100, 100, 120]],
        )
        aggregate_recommendations_path.write_text(
            json.dumps(
                {
                    "decisions": [
                        {
                            "action": "increase",
                            "recommendation_type": "top1_gap",
                            "location": "Warsaw",
                            "start_date": "2026-07-15",
                            "rental_days": 3,
                            "suggested_rate_pln_day": 120,
                            "maximum_import_rate_pln_day": 120,
                            "site_cap_rate_pln_day": 120,
                            "broker_markup_multiplier": 1,
                            "data_quality_status": "ok",
                        },
                        {
                            "action": "decrease",
                            "recommendation_type": "top1_undercut",
                            "location": "Warsaw",
                            "start_date": "2026-07-15",
                            "rental_days": 4,
                            "suggested_rate_pln_day": 80,
                            "maximum_import_rate_pln_day": 80,
                            "site_cap_rate_pln_day": 80,
                            "broker_markup_multiplier": 1,
                            "data_quality_status": "ok",
                        },
                    ]
                }
            ),
            encoding="utf-8",
        )
        aggregate_summary = apply_updates(
            workbook_path=aggregate_workbook_path,
            recommendations_path=aggregate_recommendations_path,
            output_path=aggregate_output_path,
            config=merge_config({"location_zones": {"Warsaw": ["WA1"]}}),
            cli_groups=None,
            dry_run=False,
        )
        aggregate_ws = openpyxl.load_workbook(aggregate_output_path)["Sheet1"]
        assert_equal(aggregate_ws["K5"].value, 80, "duration band uses the most restrictive recommendation")
        assert_equal(aggregate_summary["change_count"], 1, "duration band writes one cell once")
        assert_equal(aggregate_summary["changes"][0]["source_decision_count"], 2, "duration band source count")
        assert_equal(aggregate_summary["changes"][0]["aggregation_conflict"], True, "duration band conflict flag")

        partial_workbook_path = tmpdir / "partial-duration-rates.xlsx"
        partial_recommendations_path = tmpdir / "partial-duration-recommendations.json"
        partial_output_path = tmpdir / "partial-duration-updated.xlsx"
        build_minimal_workbook(
            partial_workbook_path,
            [["CDMV", None, None, "WA1", "14-09-26", "15-09-26", "15-09-26", "15-09-26", 160, 100, 100, 100, 100, 120]],
        )
        partial_recommendations_path.write_text(
            json.dumps(
                {
                    "decisions": [{
                        "action": "increase",
                        "recommendation_type": "top1_gap",
                        "location": "Warsaw",
                        "start_date": "2026-09-15",
                        "rental_days": 8,
                        "suggested_rate_pln_day": 150,
                        "maximum_import_rate_pln_day": 150,
                        "site_cap_rate_pln_day": 150,
                        "broker_markup_multiplier": 1,
                        "data_quality_status": "ok",
                    }]
                }
            ),
            encoding="utf-8",
        )
        partial_summary = apply_updates(
            workbook_path=partial_workbook_path,
            recommendations_path=partial_recommendations_path,
            output_path=partial_output_path,
            config=merge_config({"location_zones": {"Warsaw": ["WA1"]}}),
            cli_groups=None,
            dry_run=False,
        )
        partial_ws = openpyxl.load_workbook(partial_output_path)["Sheet1"]
        assert_equal(partial_ws["M5"].value, 100, "incomplete duration band does not raise the current rate")
        assert_equal(partial_summary["change_count"], 0, "incomplete duration band increase is skipped")
        assert_equal(partial_summary["skipped_target_count"], 1, "incomplete duration band is reported")

        force_top1_workbook_path = tmpdir / "force-top1-rates.xlsx"
        force_top1_recommendations_path = tmpdir / "force-top1-recommendations.json"
        force_top1_output_path = tmpdir / "force-top1-updated.xlsx"
        build_minimal_workbook(
            force_top1_workbook_path,
            [
                ["CDMV", None, None, "WA1", "14-09-26", "15-09-26", "15-09-26", "15-09-26", 160, 100, 100, 100, 100, 120],
                ["EDMV", None, None, "WA1", "14-09-26", "15-09-26", "15-09-26", "15-09-26", 160, 101, 100, 100, 100, 120],
            ],
        )
        force_top1_recommendations_path.write_text(
            json.dumps(
                {
                    "decisions": [{
                        "action": "decrease",
                        "recommendation_type": "force_top1_undercut",
                        "location": "Warsaw",
                        "start_date": "2026-09-15",
                        "rental_days": 2,
                        "suggested_rate_pln_day": 80,
                        "site_cap_rate_pln_day": 80,
                        "broker_markup_multiplier": 1,
                        "data_quality_status": "ok",
                    }]
                }
            ),
            encoding="utf-8",
        )
        force_top1_summary = apply_updates(
            workbook_path=force_top1_workbook_path,
            recommendations_path=force_top1_recommendations_path,
            output_path=force_top1_output_path,
            config=merge_config({"location_zones": {"Warsaw": ["WA1"]}}),
            cli_groups=None,
            dry_run=False,
        )
        force_top1_ws = openpyxl.load_workbook(force_top1_output_path)["Sheet1"]
        assert_equal(force_top1_ws["J5"].value, 79, "force top1 reserves premium adjustment in base rate")
        assert_equal(force_top1_ws["J6"].value, 80, "force top1 keeps EDMV below the site cap")
        assert all(
            change["target_achievable"]
            for change in force_top1_summary["changes"]
            if change["recommendation_type"] != "group_parity"
        )

        city_cap_workbook_path = tmpdir / "city-cap-rates.xlsx"
        city_cap_recommendations_path = tmpdir / "city-cap-recommendations.json"
        city_cap_output_path = tmpdir / "city-cap-updated.xlsx"
        parity_groups = ["CDMV", "CWAV", "CWMR", "EDMV"]
        city_cap_rows = []
        for zone, base_rate in (("WA1", 90), ("WALO", 100)):
            for group in parity_groups:
                adjustment = 1 if group == "EDMV" else 0
                city_cap_rows.append([
                    group,
                    None,
                    None,
                    zone,
                    "14-09-26",
                    "15-09-26",
                    "15-09-26",
                    "15-09-26",
                    160,
                    base_rate + adjustment,
                    100,
                    100,
                    100,
                    120,
                ])
        build_minimal_workbook(city_cap_workbook_path, city_cap_rows)
        city_cap_recommendations_path.write_text(
            json.dumps({
                "decisions": [{
                    "action": "increase",
                    "recommendation_type": "top1_gap",
                    "location": "Warsaw Train Station",
                    "start_date": "2026-09-15",
                    "rental_days": 2,
                    "suggested_rate_pln_day": 150,
                    "maximum_import_rate_pln_day": 150,
                    "site_cap_rate_pln_day": 150,
                    "broker_markup_multiplier": 1,
                    "benchmark_provider": "Car24",
                    "benchmark_rate_pln_day": 151,
                    "mm_rate_pln_day": 90,
                    "data_quality_status": "ok",
                }]
            }),
            encoding="utf-8",
        )
        city_cap_summary = apply_updates(
            workbook_path=city_cap_workbook_path,
            recommendations_path=city_cap_recommendations_path,
            output_path=city_cap_output_path,
            config=merge_config({
                "location_zones": {"Warsaw Train Station": ["WA1"]},
                "city_zone_airport_zones": {"WA1": ["WALO"]},
                "zone_location_labels": {"WALO": "Warsaw Chopin Airport (WAW)"},
            }),
            cli_groups=None,
            dry_run=False,
        )
        city_cap_workbook = openpyxl.load_workbook(city_cap_output_path)
        city_cap_ws = city_cap_workbook["Sheet1"]
        for offset, group in enumerate(parity_groups, start=5):
            expected_rate = 131 if group == "EDMV" else 130
            assert_equal(city_cap_ws.cell(offset, 10).value, expected_rate, f"city cap rate for {group}")
        for offset, group in enumerate(parity_groups, start=5 + len(parity_groups)):
            expected_rate = 101 if group == "EDMV" else 100
            assert_equal(city_cap_ws.cell(offset, 10).value, expected_rate, f"unchanged airport rate for {group}")
        assert_equal(city_cap_summary["city_top1_airport_cap_scope_count"], 1, "city cap scope count")
        assert_equal(city_cap_summary["city_top1_airport_cap_applied_count"], 4, "city cap applied count")
        assert_equal(city_cap_summary["city_top1_airport_cap_violation_count"], 0, "city cap violations")
        assert "max 130% ceny lotniskowej" in city_cap_ws["J5"].comment.text
        assert_equal(city_cap_workbook["Changed Positions"]["A10"].value, "Limit miasto vs lotnisko", "city cap legend")
        assert "maksymalnie 130%" in city_cap_workbook["Changed Positions"]["B10"].value

        city_cap_out_of_scope_output_path = tmpdir / "city-cap-out-of-scope.xlsx"
        city_cap_out_of_scope_summary = apply_updates(
            workbook_path=city_cap_workbook_path,
            recommendations_path=city_cap_recommendations_path,
            output_path=city_cap_out_of_scope_output_path,
            config=merge_config({
                "location_zones": {"Warsaw Train Station": ["WA1"]},
                "city_zone_airport_zones": {"WA1": ["WALO"]},
                "pickup_date_expansion": {
                    "enabled": True,
                    "start_date": "2026-09-16",
                    "end_date": "2026-09-16",
                    "drop_rows_before_start_date": True,
                    "drop_rows_after_end_date": True,
                    "time_zone": "Europe/Warsaw",
                },
            }),
            cli_groups=None,
            dry_run=False,
        )
        assert_equal(
            city_cap_out_of_scope_summary["recommendation_out_of_pickup_range_count"],
            1,
            "out-of-range recommendation count",
        )
        assert_equal(city_cap_out_of_scope_summary["city_top1_airport_cap_scope_count"], 0, "out-of-range city cap scope")
        city_cap_out_of_scope_ws = openpyxl.load_workbook(city_cap_out_of_scope_output_path)["Sheet1"]
        assert_equal(
            {parse_date_value(city_cap_out_of_scope_ws.cell(row, 7).value) for row in range(5, city_cap_out_of_scope_ws.max_row + 1)},
            {date(2026, 9, 16)},
            "out-of-range recommendations do not restore removed dates",
        )

        try:
            apply_updates(
                workbook_path=city_cap_workbook_path,
                recommendations_path=city_cap_recommendations_path,
                output_path=tmpdir / "city-cap-floor-conflict.xlsx",
                config=merge_config({
                    "location_zones": {"Warsaw Train Station": ["WA1"]},
                    "city_zone_airport_zones": {"WA1": ["WALO"]},
                    "minimum_rates": {"global_min_pln_day": 131},
                }),
                cli_groups=None,
                dry_run=False,
            )
            raise AssertionError("city cap and floor conflict should block the workbook")
        except ValueError as error:
            assert "conflicts with the configured price floor" in str(error)

        city_only_workbook_path = tmpdir / "city-cap-missing-airport.xlsx"
        build_minimal_workbook(city_only_workbook_path, city_cap_rows[:len(parity_groups)])
        try:
            apply_updates(
                workbook_path=city_only_workbook_path,
                recommendations_path=city_cap_recommendations_path,
                output_path=tmpdir / "city-cap-missing-airport-output.xlsx",
                config=merge_config({
                    "location_zones": {"Warsaw Train Station": ["WA1"]},
                    "city_zone_airport_zones": {"WA1": ["WALO"]},
                }),
                cli_groups=None,
                dry_run=False,
            )
            raise AssertionError("missing airport reference should block the workbook")
        except ValueError as error:
            assert "Missing airport rate required" in str(error)

        expired_config = merge_config(
            {
                "location_zones": {"Warsaw": ["WA1"]},
                "pickup_date_expansion": {
                    "enabled": True,
                    "start_date": "2027-02-01",
                    "end_date": "2027-01-31",
                    "time_zone": "Europe/Warsaw",
                },
            }
        )
        try:
            apply_updates(
                workbook_path=aggregate_workbook_path,
                recommendations_path=aggregate_recommendations_path,
                output_path=tmpdir / "expired-output.xlsx",
                config=expired_config,
                cli_groups=None,
                dry_run=False,
            )
            raise AssertionError("expired pickup range should fail before modifying Sheet1")
        except ValueError as error:
            assert "Sheet1 was not modified" in str(error)

        real_workbook_path = ROOT / "input" / "mm-cars-rental-rates-inclusive-fp.xlsx"
        real_recommendations_path = tmpdir / "real-recommendations.json"
        real_output_path = tmpdir / "real-rates-updated.xlsx"
        real_before = openpyxl.load_workbook(real_workbook_path)
        real_ws = real_before["Sheet1"]
        before_snapshot = header_rows_snapshot(real_ws)
        real_target = None
        for row in range(5, real_ws.max_row + 1):
            group = str(real_ws.cell(row, 1).value or "").strip().upper()
            zone = str(real_ws.cell(row, 4).value or "").strip().upper()
            pickup_date = parse_date_value(real_ws.cell(row, 7).value)
            old_rate = parse_number(real_ws.cell(row, 10).value)
            if group == "CDMV" and zone in {"KRDW", "KRGA", "KRLO", "KRTI"} and pickup_date and old_rate is not None:
                real_target = (pickup_date, old_rate)
                break
        if real_target is None:
            raise AssertionError("Real workbook smoke test needs a CDMV Krakow row with a duration 2 rate.")
        real_pickup_date, real_old_rate = real_target
        real_recommendations_path.write_text(
            json.dumps(
                {
                    "recommendations": [
                        {
                            "action": "increase",
                            "recommendation_type": "top1_undercut",
                            "reason": "MM Cars Rental jest top2 i brakuje mniej niz 10 PLN/dzien, zeby zostac top1; cel to 1 PLN ponizej top1.",
                            "location": "Krakow",
                            "start_date": real_pickup_date.isoformat(),
                            "rental_days": 2,
                            "suggested_rate_pln_day": real_old_rate + 10,
                            "mm_rate_pln_day": real_old_rate,
                            "benchmark_provider": "Car24",
                            "benchmark_rate_pln_day": real_old_rate + 11,
                            "scenario_id": f"real-template-{real_pickup_date.isoformat()}-2",
                        }
                    ]
                }
            ),
            encoding="utf-8",
        )

        real_summary = apply_updates(
            workbook_path=real_workbook_path,
            recommendations_path=real_recommendations_path,
            output_path=real_output_path,
            config=merge_config({"location_zones": {"Krakow": ["KRDW", "KRGA", "KRLO", "KRTI"]}}),
            cli_groups=None,
            dry_run=False,
        )
        assert real_summary["change_count"] > 0
        real_after = openpyxl.load_workbook(real_output_path)
        after_snapshot = header_rows_snapshot(real_after["Sheet1"])
        assert_equal(after_snapshot, before_snapshot, "Sheet1 rows 1-4 values and formatting")

        fixed_real_recommendations_path = tmpdir / "fixed-real-recommendations.json"
        fixed_real_output_path = tmpdir / "fixed-real-recommendations.xlsx"
        fixed_real_import_path = tmpdir / "fixed-real-import.xlsx"
        fixed_real_recommendations_path.write_text(json.dumps({"recommendations": []}), encoding="utf-8")
        fixed_real_summary = apply_updates(
            workbook_path=real_workbook_path,
            recommendations_path=fixed_real_recommendations_path,
            output_path=fixed_real_output_path,
            config=load_config(ROOT / "excel-rate-update.config.example.json"),
            cli_groups=None,
            dry_run=False,
            import_output_path=fixed_real_import_path,
        )
        assert fixed_real_summary["fixed_rate_groups"]["enabled"] is True
        expected_pickup_start = datetime.now(ZoneInfo("Europe/Warsaw")).date()
        expected_pickup_end = add_calendar_months(expected_pickup_start, 4)
        for output_file in (fixed_real_output_path, fixed_real_import_path):
            fixed_real_workbook = openpyxl.load_workbook(output_file, read_only=True)
            fixed_real_ws = fixed_real_workbook["Sheet1"]
            group_counts = Counter()
            pickup_dates = []
            for values in fixed_real_ws.iter_rows(min_row=5, min_col=1, max_col=14, values_only=True):
                group = str(values[0] or "").strip().upper()
                group_counts[group] += 1
                pickup_date = parse_date_value(values[6])
                if pickup_date is not None:
                    pickup_dates.append(pickup_date)
                if group in expected_fixed_rates:
                    assert_equal(list(values[8:14]), expected_fixed_rates[group], f"real {group} fixed rates")
            assert_equal(fixed_real_ws.max_row <= 28000, True, "real workbook stays within broker row limit")
            assert_equal(min(pickup_dates), expected_pickup_start, "real workbook pickup start")
            assert_equal(max(pickup_dates), expected_pickup_end, "real workbook pickup end")
            assert_equal(group_counts["CFAV"], group_counts["CDMV"], "real CFAV row coverage")
            assert_equal(group_counts["PDAH"], group_counts["CDMV"], "real PDAH row coverage")
            assert_equal(group_counts["EDAV"], group_counts["EDMV"], "real EDAV row coverage")
            fixed_real_workbook.close()

    print("All Excel rate updater tests passed.")


if __name__ == "__main__":
    main()
