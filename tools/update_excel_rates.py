#!/usr/bin/env python3
"""Update MM Cars Rental rate workbook from scraper pricing recommendations."""

from __future__ import annotations

import argparse
import calendar
import hashlib
import json
import math
from collections import defaultdict
from copy import copy
from datetime import date, datetime, timedelta
from pathlib import Path
from statistics import median
from typing import Any, Iterator
from zoneinfo import ZoneInfo

try:
    import openpyxl
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover - runtime environment guard
    raise SystemExit("Missing dependency: openpyxl. Install it with: pip install openpyxl") from exc


BROKER_IMPORT_ROW_LIMIT = 28000


DEFAULT_CONFIG = {
    "worksheet": "Sheet1",
    "header_row": 4,
    "data_start_row": 5,
    "duration_min_row": 2,
    "duration_max_row": 3,
    "columns": {
        "group": 1,
        "zone": 4,
        "booking_end_date": 6,
        "pickup_start_date": 7,
        "pickup_end_date": 8,
        "rate_start": 9,
    },
    "location_zones": {},
    "apply_groups": "all",
    "max_import_rows": BROKER_IMPORT_ROW_LIMIT,
    "max_recommendation_duration_days": 7,
    "excluded_groups": ["FVMD", "SWAV", "CFAV", "EDAV", "PDAH"],
    "fixed_rate_groups": {},
    "mirrored_rate_groups": {},
    "protected_rate_periods": [],
    "excluded_group_highlights": {
        "SWAV": 150,
    },
    "group_rate_adjustments_pln_day": {
        "EDMV": 1,
    },
    "group_price_parity": {
        "enabled": True,
        "base_groups": ["CDMV", "CGAV", "CWAV", "CWMR"],
        "premium_adjustments_pln_day": {
            "EDMV": 1,
        },
    },
    "city_top1_airport_cap": {
        "enabled": True,
        "max_multiplier": 1.3,
        "recommendation_types": ["top1_gap", "force_top1_maintain", "force_top1_undercut"],
    },
    "normalize_pickup_end_to_start": True,
    "pickup_date_expansion": {
        "enabled": False,
        "start_date": "today",
        "end_date": "2027-01-31",
        "time_zone": "Europe/Warsaw",
    },
    "changed_positions_sheet": "Changed Positions",
    "recommendations_review_sheet": "Recommendations Review",
    "competitor_evidence_sheet": "",
    "validation_sheet": "Validation",
    "pricing_rules_file": "pricing-rules.config.example.json",
    "minimum_rates": {
        "global_min_pln_day": 0,
        "bands": [
            {"start_date": "2026-07-01", "end_date": "2026-08-30", "min_days": 1, "max_days": 7, "min_pln_day": 70},
            {"start_date": "2026-07-01", "end_date": "2026-08-30", "min_days": 8, "max_days": 20, "min_pln_day": 115},
            {"start_date": "2026-07-01", "end_date": "2026-08-30", "min_days": 21, "max_days": 35, "min_pln_day": 100},
        ],
    },
    "delta_color_scale": {
        "max_delta_pln_day": 30,
        "increase_light": "E2F0D9",
        "increase_dark": "00B050",
        "decrease_light": "F4CCCC",
        "decrease_dark": "C00000",
    },
    "delta_color_steps": {
        "thresholds_pln_day": [5, 10, 15, 20],
        "increase": ["E2F0D9", "C6E0B4", "A9D18E", "70AD47", "00B050"],
        "decrease": ["F4CCCC", "F8B4B4", "EA9999", "E06666", "C00000"],
    },
    "changed_rate_warning": {
        "below_pln_day": None,
        "color": "FFF2CC",
    },
    "recommendation_colors": {
        "top1_gap": "9DC3E6",
        "top3_small_decrease": "FFC7CE",
        "top1_undercut": "F4B183",
        "force_top1_maintain": "9DC3E6",
        "force_top1_undercut": "F4B183",
    },
    "min_excel_change_pln_day": 0.01,
    "colors": {
        "increase": "C6EFCE",
        "decrease": "FFC7CE",
        "hold": "D9EAF7",
        "limited": "FCE4D6",
    },
}

CONFIRMED_BASELINE_STATUSES = {"confirmed_imported", "verified_live"}


def load_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def get_pricing_rules(config: dict[str, Any]) -> dict[str, Any]:
    path = Path(str(config.get("pricing_rules_file") or "pricing-rules.config.example.json"))
    if not path.exists():
        return {
            "top1GapThresholdPlnDay": 10,
            "top1UndercutThresholdPlnDay": 10,
            "top3SmallDecreaseThresholdPlnDay": 10,
            "undercutBufferPlnDay": 1,
        }
    payload = load_json(path)
    return payload.get("pricing") or payload


def merge_config(raw: dict[str, Any] | None) -> dict[str, Any]:
    raw = raw or {}
    merged = json.loads(json.dumps(DEFAULT_CONFIG))
    for key, value in raw.items():
        if isinstance(value, dict) and isinstance(merged.get(key), dict):
            merged[key].update(value)
        else:
            merged[key] = value
    return merged


def load_config(config_path: Path) -> dict[str, Any]:
    resolved_config_path = config_path.resolve()
    config = merge_config(load_json(resolved_config_path))
    config["_config_dir"] = str(resolved_config_path.parent)

    registry_file = str(config.get("location_registry_file") or "").strip()
    if not registry_file:
        return config

    registry_path = (resolved_config_path.parent / registry_file).resolve()
    registry = load_json(registry_path)
    if registry.get("schema_version") != 1 or not isinstance(registry.get("locations"), list):
        raise ValueError("Location registry must use schema_version 1 and contain a locations array.")

    location_zones: dict[str, list[str]] = {}
    locations_by_id: dict[str, str] = {}
    locations_by_city: dict[str, dict[str, list[dict[str, Any]]]] = defaultdict(lambda: defaultdict(list))
    zone_location_labels: dict[str, str] = {}
    for location in registry["locations"]:
        location_id = str(location.get("id") or "").strip()
        label = str(location.get("scraper_label") or "").strip()
        city = normalize_key(location.get("city"))
        kind = normalize_key(location.get("kind"))
        zones = sorted({normalize_code(zone) for zone in location.get("zones", []) if normalize_code(zone)})
        if not location_id or not label or not city or kind not in {"airport", "city"} or not zones:
            raise ValueError(
                "Every registry location requires id, scraper_label, city, kind (airport/city), and at least one zone."
            )
        if location_id in locations_by_id or label in location_zones:
            raise ValueError(f"Duplicate location registry entry: {location_id or label}.")
        locations_by_id[location_id] = label
        location_zones[label] = zones
        if city and kind in {"airport", "city"}:
            locations_by_city[city][kind].append({"label": label, "zones": zones})
        for zone in zones:
            if zone in zone_location_labels and zone_location_labels[zone] != label:
                raise ValueError(f"Zone {zone} belongs to more than one registry location.")
            zone_location_labels[zone] = label

    for alias, zones in (registry.get("aliases") or {}).items():
        normalized_zones = sorted({normalize_code(zone) for zone in zones if normalize_code(zone)})
        if not str(alias).strip() or not normalized_zones:
            raise ValueError("Every location alias requires a name and at least one zone.")
        location_zones[str(alias)] = normalized_zones

    daily_locations = []
    for location_id in (registry.get("profiles") or {}).get("daily", []):
        if location_id not in locations_by_id:
            raise ValueError(f"Daily location profile references unknown location id: {location_id}.")
        daily_locations.append(locations_by_id[location_id])

    zone_mirrors = {
        normalize_code(source): normalize_code(target)
        for source, target in (registry.get("zone_mirrors") or {}).items()
        if normalize_code(source) and normalize_code(target)
    }

    def canonical_zone(zone: str) -> str:
        current = normalize_code(zone)
        seen: set[str] = set()
        while current in zone_mirrors:
            if current in seen:
                raise ValueError(f"Cycle detected in zone_mirrors at {current}.")
            seen.add(current)
            current = zone_mirrors[current]
        return current

    city_zone_airport_zones: dict[str, list[str]] = {}
    for locations in locations_by_city.values():
        airport_zones = sorted({
            canonical_zone(zone)
            for location in locations.get("airport", [])
            for zone in location["zones"]
        })
        if not airport_zones:
            continue
        for location in locations.get("city", []):
            for zone in location["zones"]:
                city_zone_airport_zones[zone] = airport_zones

    config["location_zones"] = location_zones
    config["daily_locations"] = daily_locations
    config["zone_mirrors"] = zone_mirrors
    config["city_zone_airport_zones"] = city_zone_airport_zones
    config["zone_location_labels"] = zone_location_labels
    config["location_registry_path"] = str(registry_path)
    return config


def load_baseline_confirmation(config: dict[str, Any], input_workbook_sha256: str) -> dict[str, Any]:
    manifest_file = str(config.get("baseline_manifest_file") or "").strip()
    legacy_expected_hash = str(config.get("baseline_workbook_sha256") or "").strip().lower()

    if manifest_file:
        config_dir = Path(str(config.get("_config_dir") or "."))
        manifest_path = (config_dir / manifest_file).resolve()
        if not manifest_path.exists():
            raise ValueError(f"Baseline manifest not found: {manifest_path}")
        manifest = load_json(manifest_path)
        status = str(manifest.get("status") or "").strip().lower()
        expected_hash = str(manifest.get("workbook_sha256") or "").strip().lower()
        if not expected_hash:
            raise ValueError("Baseline manifest does not contain workbook_sha256.")
        if input_workbook_sha256.lower() != expected_hash:
            raise ValueError(
                "Input workbook does not match the baseline manifest; no workbook changes were made. "
                f"Expected {expected_hash}, got {input_workbook_sha256}."
            )
        if status not in CONFIRMED_BASELINE_STATUSES:
            raise ValueError(
                "Baseline workbook is not confirmed as imported. "
                f"Current status: {status or 'missing'}."
            )
        return {
            "status": status,
            "confirmed": True,
            "calibration_eligible": True,
            "workbook_sha256": expected_hash,
            "confirmed_at": manifest.get("confirmed_at"),
            "confirmed_by": manifest.get("confirmed_by"),
            "manifest_path": str(manifest_path),
        }

    if legacy_expected_hash and input_workbook_sha256.lower() != legacy_expected_hash:
        raise ValueError(
            "Input workbook does not match the confirmed baseline hash; no workbook changes were made. "
            f"Expected {legacy_expected_hash}, got {input_workbook_sha256}."
        )

    return {
        "status": "legacy_confirmed" if legacy_expected_hash else "legacy_unmanaged",
        "confirmed": bool(legacy_expected_hash),
        "calibration_eligible": True,
        "workbook_sha256": legacy_expected_hash or input_workbook_sha256,
        "confirmed_at": None,
        "confirmed_by": None,
        "manifest_path": None,
    }


def parse_date_value(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text:
        return None
    if "T" in text:
        text = text.split("T", 1)[0]

    for fmt in ("%Y-%m-%d", "%d-%m-%y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def resolve_config_date(value: Any, time_zone: str) -> date:
    text = str(value or "today").strip().lower()
    if text == "today":
        return datetime.now(ZoneInfo(time_zone)).date()

    parsed = parse_date_value(value)
    if parsed is None:
        raise ValueError(f"Invalid pickup date expansion date: {value!r}")
    return parsed


def add_calendar_months(value: date, months: int) -> date:
    if months < 0:
        raise ValueError("pickup_date_expansion.months_ahead must not be negative.")
    month_index = value.month - 1 + months
    year = value.year + month_index // 12
    month = month_index % 12 + 1
    day = min(value.day, calendar.monthrange(year, month)[1])
    return date(year, month, day)


def iter_dates_inclusive(start_date: date, end_date: date) -> Iterator[date]:
    current = start_date
    while current <= end_date:
        yield current
        current += timedelta(days=1)


def format_pickup_date_like_template(value: date, template_value: Any) -> Any:
    if isinstance(template_value, datetime):
        return datetime.combine(value, template_value.time())
    if isinstance(template_value, date):
        return value
    return value.strftime("%d-%m-%y")


def parse_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and math.isfinite(float(value)):
        return float(value)
    text = str(value).strip().replace(" ", "").replace(",", ".")
    if not text:
        return None
    try:
        parsed = float(text)
    except ValueError:
        return None
    return parsed if math.isfinite(parsed) else None


def build_change_statistics(changes: list[dict[str, Any]]) -> dict[str, Any]:
    deltas = [parse_number(change.get("delta")) for change in changes]
    increases = [delta for delta in deltas if delta is not None and delta > 0]
    decreases = [delta for delta in deltas if delta is not None and delta < 0]
    return {
        "increase_count": len(increases),
        "decrease_count": len(decreases),
        "average_increase_pln_day": round(sum(increases) / len(increases), 2) if increases else None,
        "average_decrease_pln_day": round(sum(decreases) / len(decreases), 2) if decreases else None,
    }


def normalize_key(value: Any) -> str:
    return str(value or "").strip().lower()


def normalize_code(value: Any) -> str:
    return str(value or "").strip().upper()


def get_duration_columns(ws: Any, config: dict[str, Any]) -> dict[int, tuple[int, str, int, int]]:
    min_row = int(config["duration_min_row"])
    max_row = int(config["duration_max_row"])
    rate_start = int(config["columns"]["rate_start"])
    duration_columns: dict[int, tuple[int, str]] = {}

    for col in range(rate_start, ws.max_column + 1):
        min_days = parse_number(ws.cell(min_row, col).value)
        max_days = parse_number(ws.cell(max_row, col).value)
        if min_days is None or max_days is None:
            continue

        left = int(min_days)
        right = int(max_days)
        label = str(left) if left == right else f"{left}-{right}"
        for duration in range(left, right + 1):
            duration_columns[duration] = (col, label, left, right)

    return duration_columns


def load_recommendation_items(path: Path) -> list[dict[str, Any]]:
    payload = load_json(path)
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict) and isinstance(payload.get("decisions"), list):
        return payload["decisions"]
    if isinstance(payload, dict) and isinstance(payload.get("recommendations"), list):
        return payload["recommendations"]
    raise ValueError("Recommendations file must be a list or an object with a 'decisions' or 'recommendations' list.")


def get_recommendation_scope_dates(recommendations: list[dict[str, Any]]) -> set[date]:
    dates: set[date] = set()
    for recommendation in recommendations:
        pickup_date = parse_date_value(
            recommendation.get("start_date") or recommendation.get("pickup_date")
        )
        if pickup_date is not None:
            dates.add(pickup_date)
    return dates


def filter_recommendations_to_pickup_date_range(
    recommendations: list[dict[str, Any]],
    config: dict[str, Any],
    expansion_summary: dict[str, Any],
) -> tuple[list[dict[str, Any]], int]:
    if not expansion_summary.get("enabled"):
        return recommendations, 0

    settings = config.get("pickup_date_expansion") or {}
    start_date = (
        parse_date_value(expansion_summary.get("start_date"))
        if settings.get("drop_rows_before_start_date")
        else None
    )
    end_date = (
        parse_date_value(expansion_summary.get("end_date"))
        if settings.get("drop_rows_after_end_date")
        else None
    )
    filtered: list[dict[str, Any]] = []
    removed_count = 0
    for recommendation in recommendations:
        pickup_date = parse_date_value(
            recommendation.get("start_date") or recommendation.get("pickup_date")
        )
        if pickup_date is not None and (
            (start_date is not None and pickup_date < start_date)
            or (end_date is not None and pickup_date > end_date)
        ):
            removed_count += 1
            continue
        filtered.append(recommendation)
    return filtered, removed_count


def is_accepted_value(value: Any) -> bool:
    text = str(value or "").strip().lower()
    return text in {"1", "yes", "y", "true", "tak", "t", "x", "accepted", "accept"}


def get_acceptance_key(
    scenario_id: Any,
    location: Any,
    pickup_date: Any,
    duration_band: Any,
) -> tuple[str, str, str, str]:
    parsed_date = parse_date_value(pickup_date)
    return (
        normalize_key(scenario_id),
        normalize_key(location),
        parsed_date.isoformat() if parsed_date else normalize_key(pickup_date),
        normalize_key(duration_band),
    )


def load_acceptance_keys(path: Path, sheet_name: str) -> set[tuple[str, str, str, str]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Acceptance workbook does not contain sheet '{sheet_name}'.")

    ws = workbook[sheet_name]
    headers = {
        normalize_key(cell.value): index
        for index, cell in enumerate(ws[1], start=1)
        if cell.value not in (None, "")
    }

    def header_col(*names: str) -> int | None:
        for name in names:
            col = headers.get(normalize_key(name))
            if col:
                return col
        return None

    accept_col = header_col("Accept?", "Akceptacja?")
    if not accept_col:
        raise ValueError("Acceptance sheet must contain an 'Accept?' or 'Akceptacja?' column.")

    scenario_col = header_col("Scenario ID", "ID scenariusza")
    location_col = header_col("Location", "Lokalizacja")
    pickup_col = header_col("Pickup date", "Data odbioru")
    duration_col = header_col("Duration band", "Przedzial duration")

    accepted: set[tuple[str, str, str, str]] = set()
    for row in range(2, ws.max_row + 1):
        if not is_accepted_value(ws.cell(row, accept_col).value):
            continue
        accepted.add(
            get_acceptance_key(
                ws.cell(row, scenario_col).value if scenario_col else "",
                ws.cell(row, location_col).value if location_col else "",
                ws.cell(row, pickup_col).value if pickup_col else "",
                ws.cell(row, duration_col).value if duration_col else "",
            )
        )
    return accepted


def resolve_apply_groups(config: dict[str, Any], cli_groups: str | None) -> set[str] | str:
    raw_groups: Any = cli_groups if cli_groups is not None else config.get("apply_groups")
    if isinstance(raw_groups, str):
        if raw_groups.strip().lower() == "all":
            return "all"
        values = [item.strip() for item in raw_groups.split(",")]
    else:
        values = [str(item).strip() for item in raw_groups or []]

    groups = {normalize_code(item) for item in values if str(item).strip()}
    if not groups:
        raise ValueError("Set apply_groups to an explicit list of car groups, or pass --groups=all intentionally.")
    return groups


def group_is_allowed(group: Any, allowed_groups: set[str] | str) -> bool:
    return allowed_groups == "all" or normalize_code(group) in allowed_groups


def group_is_excluded(group: Any, config: dict[str, Any]) -> bool:
    excluded = {normalize_code(item) for item in config.get("excluded_groups", [])}
    return normalize_code(group) in excluded


def get_fixed_rate_groups(config: dict[str, Any]) -> dict[str, dict[str, Any]]:
    fixed_groups: dict[str, dict[str, Any]] = {}
    for raw_group, raw_settings in (config.get("fixed_rate_groups") or {}).items():
        group = normalize_code(raw_group)
        settings = raw_settings if isinstance(raw_settings, dict) else {"rate_pln_day": raw_settings}
        rate = parse_number(settings.get("rate_pln_day"))
        raw_band_rates = settings.get("rates_by_duration_band") or {}
        band_rates = {
            str(label).strip(): parse_number(value)
            for label, value in raw_band_rates.items()
            if str(label).strip()
        } if isinstance(raw_band_rates, dict) else {}
        template_group = normalize_code(settings.get("template_group") or "CDMV")
        invalid_band_rate = any(value is None or value < 0 for value in band_rates.values())
        if (
            not group
            or (rate is None and not band_rates)
            or (rate is not None and rate < 0)
            or invalid_band_rate
            or not template_group
            or group == template_group
        ):
            raise ValueError(f"Invalid fixed_rate_groups configuration for '{raw_group}'.")
        fixed_groups[group] = {
            "rate_pln_day": rate,
            "rates_by_duration_band": band_rates,
            "template_group": template_group,
        }
    return fixed_groups


def get_fixed_rate_for_duration_band(settings: dict[str, Any], duration_band: str) -> float:
    flat_rate = parse_number(settings.get("rate_pln_day"))
    if flat_rate is not None:
        return float(flat_rate)
    band_rate = parse_number((settings.get("rates_by_duration_band") or {}).get(duration_band))
    if band_rate is None:
        raise ValueError(f"Missing fixed rate for duration band '{duration_band}'.")
    return float(band_rate)


def get_mirrored_rate_groups(config: dict[str, Any]) -> dict[str, dict[str, str]]:
    mirrored_groups: dict[str, dict[str, str]] = {}
    for raw_group, raw_settings in (config.get("mirrored_rate_groups") or {}).items():
        group = normalize_code(raw_group)
        settings = raw_settings if isinstance(raw_settings, dict) else {"template_group": raw_settings}
        template_group = normalize_code(settings.get("template_group"))
        if not group or not template_group or group == template_group:
            raise ValueError(f"Invalid mirrored_rate_groups configuration for '{raw_group}'.")
        mirrored_groups[group] = {"template_group": template_group}
    return mirrored_groups


def get_protected_rate_periods(config: dict[str, Any]) -> list[tuple[date, date]]:
    periods: list[tuple[date, date]] = []
    for item in config.get("protected_rate_periods") or []:
        settings = item if isinstance(item, dict) else {}
        start_date = parse_date_value(settings.get("start_date"))
        end_date = parse_date_value(settings.get("end_date"))
        if start_date is None or end_date is None or start_date > end_date:
            raise ValueError("Invalid protected_rate_periods configuration.")
        periods.append((start_date, end_date))
    return periods


def date_is_rate_protected(value: Any, periods: list[tuple[date, date]]) -> bool:
    parsed = value if isinstance(value, date) else parse_date_value(value)
    return parsed is not None and any(start_date <= parsed <= end_date for start_date, end_date in periods)


def get_excluded_group_highlight_threshold(group: Any, config: dict[str, Any]) -> float | None:
    highlights = {
        normalize_code(key): parse_number(value)
        for key, value in (config.get("excluded_group_highlights") or {}).items()
    }
    return highlights.get(normalize_code(group))


def get_group_rate_adjustment(group: Any, config: dict[str, Any]) -> float:
    adjustments = {
        normalize_code(key): parse_number(value) or 0
        for key, value in (config.get("group_rate_adjustments_pln_day") or {}).items()
    }
    return float(adjustments.get(normalize_code(group), 0))


def get_group_price_parity(config: dict[str, Any]) -> tuple[list[str], dict[str, float]] | None:
    parity = config.get("group_price_parity") or {}
    if parity.get("enabled") is False:
        return None

    base_groups = [normalize_code(item) for item in parity.get("base_groups", []) if normalize_code(item)]
    premium_adjustments = {
        normalize_code(key): float(parse_number(value) or 0)
        for key, value in (parity.get("premium_adjustments_pln_day") or {}).items()
        if normalize_code(key)
    }
    if not base_groups or not premium_adjustments:
        return None
    return base_groups, premium_adjustments


def get_city_top1_airport_cap(config: dict[str, Any]) -> tuple[float, set[str]] | None:
    settings = config.get("city_top1_airport_cap") or {}
    if settings.get("enabled") is False:
        return None

    multiplier = parse_number(settings.get("max_multiplier")) or 1.3
    if multiplier < 1:
        raise ValueError("city_top1_airport_cap.max_multiplier must be at least 1.")
    recommendation_types = {
        str(item).strip()
        for item in settings.get("recommendation_types", [])
        if str(item).strip()
    }
    if not recommendation_types:
        return None
    return float(multiplier), recommendation_types


def target_matches_recommendation_types(target: dict[str, Any], recommendation_types: set[str]) -> bool:
    source_types = {
        str(item).strip()
        for item in target.get("source_recommendation_types", [])
        if str(item).strip()
    }
    direct_type = str(target.get("recommendation_type") or "").strip()
    if direct_type:
        source_types.add(direct_type)
    return bool(source_types & recommendation_types)


def get_force_top1_base_offset(target: dict[str, Any], config: dict[str, Any]) -> float:
    if target.get("recommendation_type") not in {"force_top1_maintain", "force_top1_undercut"}:
        return 0
    parity = get_group_price_parity(config)
    if parity is None:
        return 0
    _base_groups, premium_adjustments = parity
    return max([0, *premium_adjustments.values()])


def calculate_target_base_rate(
    target: dict[str, Any],
    current_base_equivalent: float | None,
    config: dict[str, Any],
) -> tuple[float, float, float, str]:
    minimum_rate, minimum_reason = get_minimum_rate(target, config)
    suggested_rate = float(target["suggested_rate_pln_day"]) - get_force_top1_base_offset(target, config)
    base_rate = max(suggested_rate, minimum_rate)
    if (
        target.get("duration_band_coverage_complete") is False
        and current_base_equivalent is not None
        and base_rate > current_base_equivalent
        and minimum_rate <= current_base_equivalent
    ):
        base_rate = current_base_equivalent
    return base_rate, suggested_rate, minimum_rate, minimum_reason


def classify_actual_action(old_rate: float | None, new_rate: float, fallback_action: str) -> str:
    if old_rate is None:
        return fallback_action
    if new_rate > old_rate:
        return "increase"
    if new_rate < old_rate:
        return "decrease"
    return "hold"


def evaluate_target_constraints(target: dict[str, Any], new_rate: float) -> dict[str, Any]:
    constraints = target.get("constraint_items") or [target]
    evaluated: list[dict[str, Any]] = []
    for item in constraints:
        multiplier = parse_number(item.get("broker_markup_multiplier")) or 1
        site_cap = parse_number(item.get("site_cap_rate_pln_day"))
        if site_cap is None:
            site_cap = parse_number(item.get("site_target_rate_pln_day"))
        predicted = round(new_rate * multiplier, 2)
        satisfied = site_cap is not None and predicted <= site_cap + 0.01
        evaluated.append({
            "duration_days": item.get("rental_days"),
            "action": item.get("action"),
            "site_cap_rate": site_cap,
            "predicted_site_rate": predicted,
            "satisfied": satisfied,
        })

    unmet = [item for item in evaluated if not item["satisfied"]]
    predicted_values = [float(item["predicted_site_rate"]) for item in evaluated]
    return {
        "target_achievable": not unmet,
        "unmet_constraint_count": len(unmet),
        "unmet_constraints": unmet,
        "predicted_site_rate_min": min(predicted_values) if predicted_values else None,
        "predicted_site_rate_max": max(predicted_values) if predicted_values else None,
    }


def format_rate_for_comment(value: float | None) -> str:
    if value is None:
        return "puste"
    if float(value).is_integer():
        return str(int(value))
    return f"{value:.2f}"


def format_delta_for_comment(value: float | None) -> str:
    formatted = format_rate_for_comment(value)
    if value is not None and value > 0:
        return f"+{formatted}"
    return formatted


def format_percent_for_comment(value: float | None) -> str:
    if value is None:
        return ""
    if float(value).is_integer():
        return f"{int(value)}%"
    return f"{value:.2f}%"


def clamp(value: float, minimum: float, maximum: float) -> float:
    return max(minimum, min(maximum, value))


def parse_hex_color(value: Any) -> tuple[int, int, int]:
    text = str(value or "").strip().replace("#", "")
    if len(text) == 8:
        text = text[-6:]
    if len(text) != 6:
        text = "FFFFFF"
    return int(text[0:2], 16), int(text[2:4], 16), int(text[4:6], 16)


def interpolate_color(light: Any, dark: Any, ratio: float) -> str:
    start = parse_hex_color(light)
    end = parse_hex_color(dark)
    ratio = clamp(ratio, 0, 1)
    return "".join(f"{round(start[index] + (end[index] - start[index]) * ratio):02X}" for index in range(3))


def get_delta_fill(change: dict[str, Any], config: dict[str, Any]) -> PatternFill:
    warning = config.get("changed_rate_warning") or {}
    warning_threshold = parse_number(warning.get("below_pln_day"))
    new_rate = parse_number(change.get("new_rate"))
    if warning_threshold is not None and new_rate is not None and new_rate < warning_threshold:
        color = str(warning.get("color") or "FFF2CC").replace("#", "")
        return PatternFill(fill_type="solid", fgColor=color)

    delta = parse_number(change.get("delta"))
    if delta is None or delta == 0:
        color = (config.get("colors") or {}).get("hold", "D9EAF7")
        return PatternFill(fill_type="solid", fgColor=str(color).replace("#", ""))

    steps = config.get("delta_color_steps") or {}
    palette = steps.get("increase" if delta > 0 else "decrease") or []
    thresholds = [parse_number(item) for item in (steps.get("thresholds_pln_day") or [])]
    thresholds = [item for item in thresholds if item is not None]
    if palette:
        color_index = 0
        for threshold in thresholds:
            if abs(delta) > threshold:
                color_index += 1
        color = palette[min(color_index, len(palette) - 1)]
        return PatternFill(fill_type="solid", fgColor=str(color).replace("#", ""))

    scale = config.get("delta_color_scale") or {}
    max_delta = parse_number(scale.get("max_delta_pln_day")) or 30
    ratio = abs(delta) / max_delta
    if delta > 0:
        color = interpolate_color(scale.get("increase_light", "E2F0D9"), scale.get("increase_dark", "00B050"), ratio)
    else:
        color = interpolate_color(scale.get("decrease_light", "F4CCCC"), scale.get("decrease_dark", "C00000"), ratio)
    return PatternFill(fill_type="solid", fgColor=color)


def get_recommendation_fill(change: dict[str, Any], config: dict[str, Any]) -> PatternFill | None:
    recommendation_type = change.get("recommendation_type")
    color = (config.get("recommendation_colors") or {}).get(recommendation_type)
    if not color:
        return None
    return PatternFill(fill_type="solid", fgColor=str(color).replace("#", ""))


def append_city_top1_airport_cap_reason(change: dict[str, Any], reason: str) -> str:
    if not change.get("city_top1_airport_cap_active"):
        return reason
    multiplier = parse_number(change.get("city_top1_airport_max_multiplier")) or 1.3
    airport_location = change.get("airport_reference_location") or change.get("airport_reference_zone") or "lotnisko"
    airport_rate = format_rate_for_comment(parse_number(change.get("airport_reference_rate")))
    maximum_rate = format_rate_for_comment(parse_number(change.get("city_top1_airport_max_rate")))
    cap_text = (
        f"Dla oddzialu miejskiego obowiazuje dodatkowo limit {format_percent_for_comment((multiplier - 1) * 100)} "
        f"powyzej stawki {airport_location}: {airport_rate} PLN, czyli maksymalnie {maximum_rate} PLN."
    )
    return f"{reason} {cap_text}".strip()


def get_recommendation_reason_pl(change: dict[str, Any]) -> str:
    recommendation_type = change.get("recommendation_type")
    benchmark_provider = change.get("benchmark_provider") or "konkurent"
    benchmark_rate = format_rate_for_comment(parse_number(change.get("benchmark_rate")))
    if recommendation_type == "top1_gap":
        reason = (
            "MM Cars Rental jest na 1 miejscu, a druga oferta jest drozsza o co najmniej "
            "10 PLN/dzien. Cel jest ustawiony 1 PLN ponizej top2: "
            f"{benchmark_provider} ({benchmark_rate} PLN)."
        )
    elif recommendation_type == "top3_small_decrease":
        target_rank = int(parse_number(change.get("target_rank")) or 3)
        reason = (
            "Male obnizenie ceny, ponizej 10 PLN/dzien, pozwala przeskoczyc "
            f"rywala z top{target_rank}: {benchmark_provider} ({benchmark_rate} PLN)."
        )
    elif recommendation_type == "top1_undercut":
        reason = (
            "MM Cars Rental jest na 2 miejscu i brakuje mniej niz 10 PLN/dzien, "
            "zeby zostac top1. Cel jest ustawiony 1 PLN ponizej "
            f"benchmarku {benchmark_provider} ({benchmark_rate} PLN)."
        )
    elif recommendation_type == "force_top1_maintain":
        reason = (
            "MM Cars Rental jest top1. Cel utrzymuje top1 przy cenie 1 PLN ponizej "
            f"top2: {benchmark_provider} ({benchmark_rate} PLN)."
        )
    elif recommendation_type == "force_top1_undercut":
        reason = (
            "MM Cars Rental nie jest top1. Cel jest ustawiony 1 PLN ponizej "
            f"aktualnego top1: {benchmark_provider} ({benchmark_rate} PLN)."
        )
    else:
        reason = str(change.get("reason") or "")
    return append_city_top1_airport_cap_reason(change, reason)


def get_recommendation_outcome_pl(change: dict[str, Any]) -> str:
    if change.get("target_achievable") is False:
        outcome = "cel rankingowy nie jest gwarantowany przy finalnej stawce; pozycja wymaga kontroli."
        return outcome
    recommendation_type = change.get("recommendation_type")
    if recommendation_type == "group_parity":
        outcome = "spojnosc stawek grup bazowych oraz korekta skonfigurowanych grup premium."
    elif recommendation_type == "top1_gap":
        outcome = "utrzymanie top1 przy cenie 1 PLN ponizej top2."
    elif recommendation_type == "top3_small_decrease":
        target_rank = int(parse_number(change.get("target_rank")) or 3)
        outcome = f"top{target_rank} przy cenie 1 PLN ponizej rywala z top{target_rank}."
    elif recommendation_type == "top1_undercut":
        outcome = "top1 przy cenie 1 PLN ponizej obecnego top1."
    elif recommendation_type in {"force_top1_maintain", "force_top1_undercut"}:
        outcome = "top1 przy cenie 1 PLN ponizej najblizszego benchmarku."
    else:
        outcome = ""
    if change.get("city_top1_airport_cap_active"):
        multiplier = parse_number(change.get("city_top1_airport_max_multiplier")) or 1.3
        suffix = (
            " Cena oddzialu miejskiego pozostaje nie wyzsza niz "
            f"{format_percent_for_comment(multiplier * 100)} odpowiadajacej stawki lotniskowej."
        )
        outcome = f"{outcome}{suffix}".strip()
    return outcome


def get_minimum_rate(target: dict[str, Any], config: dict[str, Any]) -> tuple[float, str]:
    rules = config.get("minimum_rates") or {}
    minimum = parse_number(rules.get("global_min_pln_day")) or 0
    reason = f"Minimum globalne: {format_rate_for_comment(minimum)} PLN brutto/dzien." if minimum else ""

    target_date = target.get("target_date")
    duration_min = int(parse_number(target.get("duration_min_days")) or parse_number(target.get("rental_days")) or 0)
    duration_max = int(parse_number(target.get("duration_max_days")) or duration_min)
    matching_bands: list[tuple[float, str]] = []
    for band in rules.get("bands") or []:
        band_start = parse_date_value(band.get("start_date"))
        band_end = parse_date_value(band.get("end_date"))
        band_min_days = int(parse_number(band.get("min_days")) or 0)
        band_max_days = int(parse_number(band.get("max_days")) or band_min_days)
        band_rate = parse_number(band.get("min_pln_day"))
        if (
            isinstance(target_date, date)
            and band_start is not None
            and band_end is not None
            and band_start <= target_date <= band_end
            and duration_min <= band_max_days
            and duration_max >= band_min_days
            and band_rate is not None
        ):
            matching_bands.append((band_rate, f"{band_start.isoformat()} - {band_end.isoformat()}, duration {band_min_days}-{band_max_days} dni"))

    if matching_bands:
        band_rate, band_label = max(matching_bands, key=lambda item: item[0])
        if band_rate > minimum:
            minimum = band_rate
            reason = f"Minimum dla {band_label}: {format_rate_for_comment(minimum)} PLN brutto/dzien."
        return minimum, reason

    duration = int(parse_number(target.get("rental_days")) or 0)
    duration_min = int(parse_number(target.get("duration_min_days")) or duration)
    long_duration_min_days = int(parse_number(rules.get("long_duration_min_days")) or 21)
    long_duration_min_rate = parse_number(rules.get("long_duration_min_pln_day"))
    if long_duration_min_rate is not None and duration >= long_duration_min_days and long_duration_min_rate > minimum:
        minimum = long_duration_min_rate
        reason = f"Minimum dla duration od {long_duration_min_days} dni: {format_rate_for_comment(minimum)} PLN brutto/dzien."

    target_date = target.get("target_date")
    season_start = parse_date_value(rules.get("season_start"))
    season_end = parse_date_value(rules.get("season_end"))
    season_column_min_days = int(parse_number(rules.get("season_duration_column_min_days")) or 8)
    season_min_rate = parse_number(rules.get("season_min_pln_day"))
    if (
        isinstance(target_date, date)
        and season_start is not None
        and season_end is not None
        and season_start <= target_date <= season_end
        and duration_min >= season_column_min_days
        and season_min_rate is not None
        and season_min_rate > minimum
    ):
        minimum = season_min_rate
        reason = (
            f"Minimum sezonowe {season_start.isoformat()} - {season_end.isoformat()} "
            f"dla kolumn od {season_column_min_days} dni: {format_rate_for_comment(minimum)} PLN brutto/dzien."
        )

    return minimum, reason


def format_for_changed_positions(value: str) -> str:
    return (
        str(value or "")
        .replace(" PLN brutto/dzien", " PLN")
        .replace(" brutto/dzien", "")
        .replace("/dzien", "")
    )


def build_rate_comment(change: dict[str, Any]) -> Comment:
    lines = [
        f"Poprzednia stawka: {format_rate_for_comment(change.get('old_rate'))} PLN",
        f"Nowa stawka: {format_rate_for_comment(change.get('new_rate'))} PLN",
        f"Zmiana: {format_delta_for_comment(change.get('delta'))} PLN",
    ]
    site_target = parse_number(change.get("site_target_rate"))
    predicted_site_rate = parse_number(change.get("predicted_site_rate"))
    broker_markup_percent = parse_number(change.get("broker_markup_percent"))
    if site_target is not None:
        lines.append(f"Cel na stronie: {format_rate_for_comment(site_target)} PLN")
    if predicted_site_rate is not None:
        lines.append(f"Prognoza na stronie: {format_rate_for_comment(predicted_site_rate)} PLN")
    if broker_markup_percent is not None:
        lines.append(f"Szac. narzut brokera: {format_percent_for_comment(broker_markup_percent)}")
    if change.get("city_top1_airport_cap_active"):
        multiplier = parse_number(change.get("city_top1_airport_max_multiplier")) or 1.3
        lines.append(
            "Limit oddzialu miejskiego: "
            f"max {format_percent_for_comment(multiplier * 100)} ceny lotniskowej "
            f"({format_rate_for_comment(parse_number(change.get('city_top1_airport_max_rate')))} PLN)"
        )
    if change.get("target_achievable") is False:
        lines.append("Cel rankingowy: wymaga kontroli")
    return Comment("\n".join(lines), "Codex")


def unique_display_values(changes: list[dict[str, Any]], field: str) -> list[tuple[Any, str]]:
    output: list[tuple[Any, str]] = []
    seen: set[str] = set()
    for change in get_display_changes_for_changed_positions(changes):
        raw_value = change.get(field)
        display_value = f"{format_rate_for_comment(raw_value)} PLN"
        if display_value in seen:
            continue
        seen.add(display_value)
        output.append((raw_value, display_value))
    return output


def unique_delta_values(changes: list[dict[str, Any]]) -> list[str]:
    output: list[str] = []
    seen: set[str] = set()
    for change in get_display_changes_for_changed_positions(changes):
        display_value = f"{format_delta_for_comment(change.get('delta'))} PLN"
        if display_value in seen:
            continue
        seen.add(display_value)
        output.append(display_value)
    return output


def format_grouped_rates(changes: list[dict[str, Any]], field: str) -> str:
    return "; ".join(display_value for _, display_value in unique_display_values(changes, field))


def format_grouped_deltas(changes: list[dict[str, Any]]) -> str:
    return "; ".join(unique_delta_values(changes))


def get_display_changes_for_changed_positions(changes: list[dict[str, Any]]) -> list[dict[str, Any]]:
    base_changes = [
        change
        for change in changes
        if not parse_number(change.get("group_adjustment_pln_day"))
    ]
    return base_changes or changes


def build_change_explanation(changes: list[dict[str, Any]]) -> str:
    change = changes[0]
    lines = [
        f"Poprzednia stawka: {format_grouped_rates(changes, 'old_rate')}",
        f"Nowa stawka: {format_grouped_rates(changes, 'new_rate')}",
        f"Zmiana: {format_grouped_deltas(changes)}",
        f"Powod rekomendacji: {format_for_changed_positions(get_recommendation_reason_pl(change))}",
    ]
    outcome = get_recommendation_outcome_pl(change)
    if outcome:
        lines.append(f"Co pozwoli osiagnac: {outcome}")
    benchmark_provider = change.get("benchmark_provider")
    benchmark_rate = parse_number(change.get("benchmark_rate"))
    if benchmark_provider or benchmark_rate is not None:
        lines.append(
            f"Benchmark: {benchmark_provider or 'n/a'} "
            f"{format_rate_for_comment(benchmark_rate)} PLN"
        )
    site_target = parse_number(change.get("site_target_rate"))
    predicted_site_rate = parse_number(change.get("predicted_site_rate"))
    broker_markup_percent = parse_number(change.get("broker_markup_percent"))
    if site_target is not None:
        lines.append(f"Cel na stronie DiscoverCars: {format_rate_for_comment(site_target)} PLN")
    if predicted_site_rate is not None or broker_markup_percent is not None:
        lines.append(
            "Kalibracja brokera: "
            f"narzut {format_percent_for_comment(broker_markup_percent) or 'n/a'}, "
            f"prognoza na stronie {format_rate_for_comment(predicted_site_rate)} PLN."
        )
    return "\n".join(lines)


def get_changed_positions_group_key(change: dict[str, Any]) -> tuple[Any, ...]:
    return (
        change.get("recommendation_type"),
        change.get("recommendation_action"),
        change.get("location"),
        change.get("zone"),
        change.get("pickup_date"),
        change.get("duration_band"),
        change.get("benchmark_provider"),
        change.get("benchmark_rate"),
        change.get("minimum_reason"),
        change.get("target_achievable"),
        change.get("airport_reference_zone"),
        change.get("city_top1_airport_max_rate"),
    )


def group_changes_for_changed_positions(changes: list[dict[str, Any]]) -> list[list[dict[str, Any]]]:
    grouped: dict[tuple[Any, ...], list[dict[str, Any]]] = {}
    for change in changes:
        grouped.setdefault(get_changed_positions_group_key(change), []).append(change)
    return list(grouped.values())


def get_strongest_delta_change(changes: list[dict[str, Any]]) -> dict[str, Any]:
    display_changes = get_display_changes_for_changed_positions(changes)
    return max(display_changes, key=lambda change: abs(parse_number(change.get("delta")) or 0))


def get_grouped_rate_summary(changes: list[dict[str, Any]]) -> Any:
    values = unique_display_values(changes, "new_rate")
    if len(values) == 1:
        new_rate = values[0][0]
        return int(new_rate) if isinstance(new_rate, float) and float(new_rate).is_integer() else new_rate
    return "\n".join(display_value for _, display_value in values)


def get_grouped_groups(changes: list[dict[str, Any]]) -> str:
    groups: list[str] = []
    seen: set[str] = set()
    for change in changes:
        group = normalize_code(change.get("group"))
        if not group or group in seen:
            continue
        seen.add(group)
        groups.append(group)
    return ", ".join(groups)


def get_floor_legend_text(config: dict[str, Any]) -> str:
    rules = config.get("minimum_rates") or {}
    parts: list[str] = []
    global_min = parse_number(rules.get("global_min_pln_day"))
    if global_min is not None:
        if global_min > 0:
            parts.append(f"globalnie {format_rate_for_comment(global_min)} PLN")

    bands = rules.get("bands") or []
    for band in bands:
        start_date = band.get("start_date")
        end_date = band.get("end_date")
        min_days = int(parse_number(band.get("min_days")) or 0)
        max_days = int(parse_number(band.get("max_days")) or min_days)
        min_rate = parse_number(band.get("min_pln_day"))
        if start_date and end_date and min_days and max_days and min_rate is not None:
            parts.append(
                f"{start_date}-{end_date}, duration {min_days}-{max_days}: {format_rate_for_comment(min_rate)} PLN"
            )

    if bands:
        return "Floor cenowy chroni przed rekomendacja i zmiana ponizej: " + "; ".join(parts) + "."

    long_days = int(parse_number(rules.get("long_duration_min_days")) or 0)
    long_rate = parse_number(rules.get("long_duration_min_pln_day"))
    if long_days and long_rate is not None:
        parts.append(f"od {long_days} dni {format_rate_for_comment(long_rate)} PLN")

    season_start = rules.get("season_start")
    season_end = rules.get("season_end")
    season_days = int(parse_number(rules.get("season_duration_column_min_days")) or 0)
    season_rate = parse_number(rules.get("season_min_pln_day"))
    if season_start and season_end and season_days and season_rate is not None:
        parts.append(
            f"{season_start}-{season_end} od {season_days} dni {format_rate_for_comment(season_rate)} PLN"
        )

    if not parts:
        return "Brak aktywnych regul floor cenowego."
    return "Floor cenowy chroni przed rekomendacja i zmiana ponizej: " + "; ".join(parts) + "."


def format_group_list(groups: list[str]) -> str:
    return ", ".join(groups) if groups else "brak"


def get_group_rules_legend_text(config: dict[str, Any]) -> str:
    parity = config.get("group_price_parity") or {}
    base_groups = [normalize_code(item) for item in parity.get("base_groups", []) if normalize_code(item)]
    premium_adjustments = {
        normalize_code(group): parse_number(adjustment) or 0
        for group, adjustment in (parity.get("premium_adjustments_pln_day") or {}).items()
        if normalize_code(group)
    }
    premium_text = ", ".join(
        f"{group}=baza+{format_rate_for_comment(adjustment)} PLN"
        for group, adjustment in premium_adjustments.items()
    )
    excluded_groups = [normalize_code(item) for item in config.get("excluded_groups", []) if normalize_code(item)]
    fixed_groups = get_fixed_rate_groups(config)
    fixed_rate_parts: list[str] = []
    for group, settings in fixed_groups.items():
        flat_rate = parse_number(settings.get("rate_pln_day"))
        if flat_rate is not None:
            fixed_rate_parts.append(f"{group}={format_rate_for_comment(flat_rate)} PLN")
            continue
        band_text = ", ".join(
            f"{duration_band}={format_rate_for_comment(rate)} PLN"
            for duration_band, rate in (settings.get("rates_by_duration_band") or {}).items()
        )
        fixed_rate_parts.append(f"{group}: {band_text}")
    fixed_rate_text = "; ".join(fixed_rate_parts)
    fixed_rate_rule = (
        f" Stawki dobowe niezalezne od rekomendacji konkurencyjnych: {fixed_rate_text}."
        if fixed_rate_text
        else ""
    )
    mirror_text = ", ".join(
        f"{group}={settings['template_group']}"
        for group, settings in get_mirrored_rate_groups(config).items()
    )
    mirror_rule = f" Klasy lustrzane: {mirror_text}." if mirror_text else ""
    protected_periods = get_protected_rate_periods(config)
    protected_rule = (
        " Istniejace stawki pozostaja bez zmian dla pickup start date: "
        + "; ".join(f"{start_date.isoformat()}-{end_date.isoformat()}" for start_date, end_date in protected_periods)
        + " (daty wlacznie)."
        if protected_periods
        else ""
    )
    max_duration = int(parse_number(config.get("max_recommendation_duration_days")) or 0)
    duration_rule = (
        f" Rekomendacje i zmiany stawek tylko dla duration 1-{max_duration} dni; "
        f"od {max_duration + 1} dni bez rekomendacji i bez zmian."
        if max_duration > 0
        else ""
    )
    return (
        f"Zmiana stawek: {format_group_list(base_groups)} maja taka sama cene bazowa; "
        f"{premium_text or 'brak grup premium'}. "
        f"Bez zmian z rekomendacji konkurencyjnych: {format_group_list(excluded_groups)}."
        f"{fixed_rate_rule}"
        f"{mirror_rule}"
        f"{duration_rule}"
        f"{protected_rule}"
    )


def get_excluded_group_highlight_legend_text(config: dict[str, Any]) -> str:
    highlights = config.get("excluded_group_highlights") or {}
    parts = [
        f"{normalize_code(group)} < {format_rate_for_comment(parse_number(threshold))} PLN/dzien"
        for group, threshold in sorted(highlights.items())
        if parse_number(threshold) is not None
    ]
    if not parts:
        return "Wykluczone klasy nie sa zmieniane; brak dodatkowych progow podswietlenia."
    return (
        "Wykluczone klasy nie sa zmieniane; moga byc tylko podswietlone kontrolnie: "
        + "; ".join(parts)
        + "."
    )


def get_city_top1_airport_cap_legend_text(config: dict[str, Any]) -> str:
    cap_rule = get_city_top1_airport_cap(config)
    if cap_rule is None:
        return "Limit ceny oddzialu miejskiego wzgledem lotniska jest wylaczony."
    multiplier, _ = cap_rule
    return (
        "Przy rekomendacji utrzymania top1 stawka oddzialu miejskiego dla tej samej daty, "
        "grupy i przedzialu duration moze wynosic maksymalnie "
        f"{format_percent_for_comment(multiplier * 100)} stawki odpowiadajacego lotniska. "
        "Lotniska nie sa ograniczane ta regula."
    )


def build_review_notes(changes: list[dict[str, Any]]) -> str:
    notes: list[str] = []
    strongest = get_strongest_delta_change(changes)
    strongest_delta = abs(parse_number(strongest.get("delta")) or 0)
    if strongest_delta >= 40:
        notes.append("duza zmiana >= 40 PLN")

    if any(change.get("minimum_reason") for change in changes):
        notes.append("ochrona floor cenowego")
    if any(change.get("city_top1_airport_cap_applied") for change in changes):
        multiplier = parse_number(changes[0].get("city_top1_airport_max_multiplier")) or 1.3
        notes.append(f"limit oddzialu miejskiego do {format_percent_for_comment(multiplier * 100)} stawki lotniskowej")
    if any(parse_number(change.get("group_adjustment_pln_day")) for change in changes):
        notes.append("korekta grupy EDMV")
    if any(parse_number(change.get("broker_markup_multiplier")) not in (None, 1) for change in changes):
        notes.append("uwzgledniono szacowany narzut brokera")
    if any(change.get("action") != change.get("recommendation_action") for change in changes):
        notes.append("kierunek po floor rozny od rekomendacji")
    if any(parse_number(change.get("benchmark_rate")) is None for change in changes):
        notes.append("brak ceny benchmarku")
    if any(parse_number(change.get("mm_rate")) is None for change in changes):
        notes.append("brak ceny MM")
    if any(change.get("target_achievable") is False for change in changes):
        notes.append("finalna stawka nie gwarantuje celu rankingowego")
    if any(change.get("aggregation_conflict") for change in changes):
        notes.append("sprzeczne kierunki w jednym przedziale duration")
    max_decisions = max((int(change.get("source_decision_count") or 0) for change in changes), default=0)
    if max_decisions > 1:
        notes.append(f"scalono {max_decisions} scenariuszy duration")
    if any(change.get("duration_band_coverage_complete") is False for change in changes):
        missing = sorted({
            int(duration)
            for change in changes
            for duration in change.get("missing_duration_days", [])
        })
        notes.append("brak danych dla duration: " + ",".join(str(item) for item in missing))

    return "; ".join(notes) if notes else "OK"


def get_review_status(changes: list[dict[str, Any]]) -> str:
    strongest = get_strongest_delta_change(changes)
    strongest_delta = abs(parse_number(strongest.get("delta")) or 0)
    critical = (
        strongest_delta >= 40
        or any(change.get("action") != change.get("recommendation_action") for change in changes)
        or any(parse_number(change.get("benchmark_rate")) is None for change in changes)
        or any(parse_number(change.get("mm_rate")) is None for change in changes)
        or any(change.get("target_achievable") is False for change in changes)
        or any(change.get("aggregation_conflict") for change in changes)
    )
    if critical:
        return "Sprawdz"
    if any(change.get("duration_band_coverage_complete") is False for change in changes):
        return "Gotowe z uwaga"
    if any(change.get("minimum_reason") for change in changes) or any(
        parse_number(change.get("group_adjustment_pln_day")) for change in changes
    ):
        return "Gotowe z uwaga"
    return "Gotowe"


def get_recommendation_label_pl(change: dict[str, Any]) -> str:
    recommendation_type = change.get("recommendation_type")
    if recommendation_type == "group_parity":
        return "Ujednolicenie grup"
    if recommendation_type == "top1_gap":
        return "Top1 gap"
    if recommendation_type == "top3_small_decrease":
        target_rank = int(parse_number(change.get("target_rank")) or 3)
        return f"Male obnizenie do top{target_rank}"
    if recommendation_type == "top1_undercut":
        return "Przebicie top1"
    if recommendation_type == "force_top1_maintain":
        return "Utrzymanie top1"
    if recommendation_type == "force_top1_undercut":
        return "Przebicie top1"
    return str(recommendation_type or "")


def copy_cell(source_cell: Any, target_cell: Any) -> None:
    target_cell.value = source_cell.value
    if source_cell.has_style:
        target_cell._style = copy(source_cell._style)
    if source_cell.hyperlink:
        target_cell._hyperlink = copy(source_cell.hyperlink)


def write_changed_positions_sheet(
    workbook: Any,
    source_ws: Any,
    config: dict[str, Any],
    changes: list[dict[str, Any]],
) -> None:
    sheet_name = str(config.get("changed_positions_sheet") or "").strip()
    if not sheet_name:
        return
    if sheet_name == source_ws.title:
        raise ValueError("changed_positions_sheet must be different from the import worksheet name.")

    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]

    header_row = int(config["header_row"])
    pricing_rules = get_pricing_rules(config)
    top1_gap = format_rate_for_comment(parse_number(pricing_rules.get("top1GapThresholdPlnDay")) or 10)
    undercut_limit = format_rate_for_comment(parse_number(pricing_rules.get("top1UndercutThresholdPlnDay")) or 10)
    top3_limit = format_rate_for_comment(parse_number(pricing_rules.get("top3SmallDecreaseThresholdPlnDay")) or 10)
    undercut_buffer = format_rate_for_comment(parse_number(pricing_rules.get("undercutBufferPlnDay")) or 1)
    if pricing_rules.get("forceTop1"):
        recommendation_legend_items = [
            ("9DC3E6", "Utrzymanie top1", f"Gdy MM Cars Rental jest top1, stawka jest ustawiana {undercut_buffer} PLN/dzien ponizej top2."),
            ("F4B183", "Przebicie top1", f"Gdy MM Cars Rental nie jest top1, stawka jest ustawiana {undercut_buffer} PLN/dzien ponizej aktualnego top1, niezaleznie od obecnej pozycji."),
        ]
    else:
        recommendation_legend_items = [
            ("9DC3E6", "Top1 gap", f"MM Cars Rental jest top1, a jego cena jest co najmniej {top1_gap} PLN/dzien nizsza niz top2; rekomendacja podnosi cene do {undercut_buffer} PLN ponizej top2."),
            ("FFC7CE", "Male obnizenie top3", f"Obnizka ponizej {top3_limit} PLN/dzien pozwala przeskoczyc wyzej ustawionego rywala z top3 ofert; cel to {undercut_buffer} PLN ponizej tej oferty."),
            ("F4B183", "Przebicie top1", f"MM Cars Rental jest top2 i brakuje mniej niz {undercut_limit} PLN/dzien, zeby zostac top1; rekomendacja ustawia cene {undercut_buffer} PLN ponizej obecnego top1."),
        ]
    legend_items = [
        *recommendation_legend_items,
        ("D9EAD3", "Scalanie duration", "Jedna komorka Sheet1 obsluguje caly przedzial duration. Stawka jest wyliczana raz z wszystkich scenariuszy w przedziale i respektuje najbardziej restrykcyjny limit."),
        ("FFF2CC", "Kontrola celu", "Po zastosowaniu finalnej stawki narzedzie ponownie przelicza prognoze na stronie. Cel nieosiagalny jest oznaczany jako wymagajacy kontroli i nie jest opisywany jako gwarantowane top1/top2/top3."),
        ("D9EAF7", "Grupy zmieniane", get_group_rules_legend_text(config)),
        ("FCE4D6", "Grupy tylko kontrolne", get_excluded_group_highlight_legend_text(config)),
        ("FCE4D6", "Floor cenowy", get_floor_legend_text(config)),
    ]
    if config.get("city_zone_airport_zones"):
        legend_items.append(("DDEBF7", "Limit miasto vs lotnisko", get_city_top1_airport_cap_legend_text(config)))
    legend_items.append(("FFFFFF", "Kolory w Sheet1", "Zielony oznacza podwyzke, czerwony obnizke; im mocniejszy kolor, tym wieksza zmiana PLN/dzien. Komentarze sa tylko w kolumnie O tego arkusza."))
    warning = config.get("changed_rate_warning") or {}
    warning_threshold = parse_number(warning.get("below_pln_day"))
    if warning_threshold is not None:
        legend_items.append(
            (
                str(warning.get("color") or "FFF2CC").replace("#", ""),
                f"Stawka ponizej {format_rate_for_comment(warning_threshold)} PLN",
                f"Zmieniona stawka ponizej {format_rate_for_comment(warning_threshold)} PLN/dzien jest oznaczona tym kolorem.",
            )
        )
    legend_rows = len(legend_items) + 2
    header_start_row = legend_rows + 1
    changed_groups = group_changes_for_changed_positions(changes)
    comment_col = 15
    source_col_count = comment_col - 1
    target_ws = workbook.create_sheet(sheet_name, workbook.index(source_ws) + 1)

    for col in range(1, comment_col + 1):
        letter = get_column_letter(col)
        source_width = source_ws.column_dimensions[letter].width
        if source_width:
            target_ws.column_dimensions[letter].width = source_width
    target_ws.column_dimensions["A"].width = 22
    target_ws.column_dimensions["B"].width = 70
    target_ws.column_dimensions[get_column_letter(comment_col)].width = 80

    target_ws["A1"] = "Legenda"
    target_ws["A1"].font = Font(bold=True)
    for row, (color, label, description) in enumerate(legend_items, start=2):
        target_ws.cell(row, 1).value = label
        target_ws.cell(row, 1).fill = PatternFill(fill_type="solid", fgColor=color)
        target_ws.cell(row, 1).font = Font(bold=True)
        target_ws.cell(row, 2).value = description
        target_ws.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")

    for row in range(1, header_row + 1):
        target_row = header_start_row + row - 1
        target_ws.row_dimensions[target_row].height = source_ws.row_dimensions[row].height
        for col in range(1, source_col_count + 1):
            copy_cell(source_ws.cell(row, col), target_ws.cell(target_row, col))

    target_header_row = header_start_row + header_row - 1
    data_start_row = target_header_row + 1
    target_ws.cell(target_header_row, comment_col).value = "Komentarz zmiany"
    target_ws.freeze_panes = f"A{data_start_row}"

    for index, grouped_changes in enumerate(changed_groups, start=data_start_row):
        change = grouped_changes[0]
        source_row = source_ws[change["cell"]].row
        target_ws.row_dimensions[index].height = source_ws.row_dimensions[source_row].height
        recommendation_fill = get_recommendation_fill(change, config)
        for col in range(1, source_col_count + 1):
            copy_cell(source_ws.cell(source_row, col), target_ws.cell(index, col))
            if recommendation_fill:
                target_ws.cell(index, col).fill = copy(recommendation_fill)

        target_ws.cell(index, 1).value = ", ".join(change["group"] for change in grouped_changes)
        target_ws.cell(index, 1).alignment = Alignment(wrap_text=True, vertical="top")
        changed_rate_col = source_ws[change["cell"]].column
        rate_cell = target_ws.cell(index, changed_rate_col)
        rate_cell.value = get_grouped_rate_summary(grouped_changes)
        rate_cell.fill = get_delta_fill(get_strongest_delta_change(grouped_changes), config)
        rate_cell.alignment = Alignment(wrap_text=True, vertical="top")

        explanation = build_change_explanation(grouped_changes)
        comment_cell = target_ws.cell(index, comment_col)
        comment_cell.value = explanation
        comment_cell.comment = Comment(explanation, "Codex")
        comment_cell.alignment = Alignment(wrap_text=True, vertical="top")
        if recommendation_fill:
            comment_cell.fill = copy(recommendation_fill)


def create_replaced_sheet(workbook: Any, sheet_name: str, index: int | None = None) -> Any:
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    if index is None:
        return workbook.create_sheet(sheet_name)
    return workbook.create_sheet(sheet_name, index)


def write_table_sheet(
    workbook: Any,
    sheet_name: str,
    index: int | None,
    headers: list[str],
    rows: list[list[Any]],
    widths: dict[str, int] | None = None,
) -> Any:
    ws = create_replaced_sheet(workbook, sheet_name, index)
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    widths = widths or {}

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = widths.get(header, min(max(len(header) + 4, 12), 36))

    for row_index, row_values in enumerate(rows, start=2):
        for col, value in enumerate(row_values, start=1):
            cell = ws.cell(row_index, col)
            cell.value = value
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    return ws


def write_recommendations_review_sheet(
    workbook: Any,
    source_ws: Any,
    config: dict[str, Any],
    changes: list[dict[str, Any]],
) -> None:
    sheet_name = str(config.get("recommendations_review_sheet") or "").strip()
    if not sheet_name:
        return

    headers = [
        "Akceptacja?",
        "Status",
        "Uwagi kontroli",
        "Lokalizacja",
        "Strefa",
        "Grupy",
        "Data odbioru",
        "Przedzial duration",
        "Poprzednia stawka",
        "Nowa stawka",
        "Delta",
        "Cel na stronie",
        "Prognoza na stronie",
        "Narzut brokera",
        "Typ rekomendacji",
        "Cel zmiany",
        "Benchmark",
        "Cena benchmarku",
        "Cena MM",
        "Top1",
        "Top2",
        "Top3",
        "Powod",
        "ID scenariusza",
        "Komorki Excel",
    ]
    rows: list[list[Any]] = []
    for grouped_changes in group_changes_for_changed_positions(changes):
        change = grouped_changes[0]
        notes = build_review_notes(grouped_changes)
        status = get_review_status(grouped_changes)
        rows.append([
            "",
            status,
            notes,
            change.get("location", ""),
            change.get("zone", ""),
            get_grouped_groups(grouped_changes),
            change.get("pickup_date", ""),
            change.get("duration_band", ""),
            format_grouped_rates(grouped_changes, "old_rate"),
            format_grouped_rates(grouped_changes, "new_rate"),
            format_grouped_deltas(grouped_changes),
            parse_number(change.get("site_target_rate")),
            parse_number(change.get("predicted_site_rate")),
            format_percent_for_comment(parse_number(change.get("broker_markup_percent"))),
            get_recommendation_label_pl(change),
            get_recommendation_outcome_pl(change),
            change.get("benchmark_provider", ""),
            parse_number(change.get("benchmark_rate")),
            parse_number(change.get("mm_rate")),
            format_provider_rate(change.get("top1_provider"), change.get("top1_rate")),
            format_provider_rate(change.get("top2_provider"), change.get("top2_rate")),
            format_provider_rate(change.get("top3_provider"), change.get("top3_rate")),
            format_for_changed_positions(get_recommendation_reason_pl(change)),
            change.get("scenario_id", ""),
            ", ".join(item.get("cell", "") for item in grouped_changes),
        ])

    widths = {
        "Akceptacja?": 12,
        "Uwagi kontroli": 34,
        "Grupy": 34,
        "Cel na stronie": 18,
        "Prognoza na stronie": 20,
        "Narzut brokera": 16,
        "Powod": 76,
        "Komorki Excel": 26,
        "Top1": 30,
        "Top2": 30,
        "Top3": 30,
    }
    ws = write_table_sheet(workbook, sheet_name, workbook.index(source_ws) + 2, headers, rows, widths)
    for row in range(2, ws.max_row + 1):
        status = ws.cell(row, 2).value
        if status == "Gotowe":
            ws.cell(row, 2).fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
        elif status == "Gotowe z uwaga":
            ws.cell(row, 2).fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
        elif status == "Sprawdz":
            ws.cell(row, 2).fill = PatternFill(fill_type="solid", fgColor="FFC7CE")


def format_provider_rate(provider: Any, rate: Any) -> str:
    provider_text = str(provider or "").strip()
    rate_number = parse_number(rate)
    if provider_text and rate_number is not None:
        return f"{provider_text} ({format_rate_for_comment(rate_number)} PLN)"
    if provider_text:
        return provider_text
    if rate_number is not None:
        return f"{format_rate_for_comment(rate_number)} PLN"
    return ""


def write_competitor_evidence_sheet(
    workbook: Any,
    source_ws: Any,
    config: dict[str, Any],
    changes: list[dict[str, Any]],
) -> None:
    sheet_name = str(config.get("competitor_evidence_sheet") or "").strip()
    if not sheet_name:
        return

    headers = [
        "ID scenariusza",
        "Wygenerowano z danych",
        "Lokalizacja",
        "Strefa",
        "Data odbioru",
        "Data zwrotu",
        "Dni najmu",
        "Przedzial duration",
        "Waluta",
        "Pozycja MM",
        "Dostawca MM",
        "Cena MM",
        "Dostawca top1",
        "Cena top1",
        "Dostawca top2",
        "Cena top2",
        "Dostawca top3",
        "Cena top3",
        "Benchmark",
        "Cena benchmarku",
        "Cel konkurencyjny przed floor",
        "Stawka zastosowana",
        "Zmiana",
        "Grupy",
    ]
    rows: list[list[Any]] = []
    for grouped_changes in group_changes_for_changed_positions(changes):
        change = grouped_changes[0]
        rows.append([
            change.get("scenario_id", ""),
            change.get("source_generated_at", ""),
            change.get("location", ""),
            change.get("zone", ""),
            change.get("pickup_date", ""),
            change.get("dropoff_date", ""),
            change.get("duration_days", ""),
            change.get("duration_band", ""),
            change.get("currency", ""),
            change.get("mm_rank", ""),
            change.get("mm_provider", ""),
            parse_number(change.get("mm_rate")),
            change.get("top1_provider", ""),
            parse_number(change.get("top1_rate")),
            change.get("top2_provider", ""),
            parse_number(change.get("top2_rate")),
            change.get("top3_provider", ""),
            parse_number(change.get("top3_rate")),
            change.get("benchmark_provider", ""),
            parse_number(change.get("benchmark_rate")),
            parse_number(change.get("suggested_rate_before_minimum")),
            format_grouped_rates(grouped_changes, "new_rate"),
            format_grouped_deltas(grouped_changes),
            get_grouped_groups(grouped_changes),
        ])

    widths = {
        "ID scenariusza": 28,
        "Wygenerowano z danych": 24,
        "Dostawca top1": 24,
        "Dostawca top2": 24,
        "Dostawca top3": 24,
        "Benchmark": 26,
        "Stawka zastosowana": 18,
        "Grupy": 34,
    }
    write_table_sheet(workbook, sheet_name, workbook.index(source_ws) + 3, headers, rows, widths)


def first_items(values: list[str], limit: int = 8) -> str:
    if not values:
        return ""
    text = "; ".join(values[:limit])
    if len(values) > limit:
        text += f"; +{len(values) - limit} more"
    return text


def get_validation_status(issue_count: int, warning: bool = False) -> str:
    if issue_count == 0:
        return "OK"
    return "WARNING" if warning else "FAIL"


def build_validation_rows(
    ws: Any,
    config: dict[str, Any],
    duration_columns: dict[int, tuple[int, str, int, int]],
    changes: list[dict[str, Any]],
    skipped_targets: list[dict[str, Any]],
    expansion_summary: dict[str, Any] | None = None,
    assume_fixed_rate_changes_applied: bool = False,
) -> list[list[Any]]:
    columns = config["columns"]
    data_start_row = int(config["data_start_row"])
    group_col = int(columns["group"])
    zone_col = int(columns["zone"])
    booking_end_col = int(columns.get("booking_end_date", 0) or 0)
    pickup_start_col = int(columns["pickup_start_date"])
    pickup_end_col = int(columns["pickup_end_date"])
    rate_cols = sorted({value[0] for value in duration_columns.values()})
    excluded_groups = {normalize_code(item) for item in config.get("excluded_groups", [])}

    data_rows = 0
    booking_mismatch: list[str] = []
    pickup_mismatch: list[str] = []
    missing_rates: list[str] = []
    duplicates: list[str] = []
    seen_keys: set[tuple[str, str, date]] = set()
    duplicate_keys: set[tuple[str, str, date]] = set()

    for row in range(data_start_row, ws.max_row + 1):
        group = normalize_code(ws.cell(row, group_col).value)
        zone = normalize_code(ws.cell(row, zone_col).value)
        pickup_start = parse_date_value(ws.cell(row, pickup_start_col).value)
        pickup_end = parse_date_value(ws.cell(row, pickup_end_col).value)
        if not group and not zone and pickup_start is None:
            continue
        data_rows += 1

        if booking_end_col:
            booking_end = parse_date_value(ws.cell(row, booking_end_col).value)
            if booking_end != pickup_end:
                booking_mismatch.append(f"row {row}: {group}/{zone}")

        if pickup_start != pickup_end:
            pickup_mismatch.append(f"row {row}: {group}/{zone}")

        if group and zone and pickup_start is not None:
            key = (group, zone, pickup_start)
            if key in seen_keys and key not in duplicate_keys:
                duplicate_keys.add(key)
                duplicates.append(f"{group}/{zone}/{pickup_start.isoformat()}")
            seen_keys.add(key)

        for col in rate_cols:
            value = parse_number(ws.cell(row, col).value)
            if value is None:
                missing_rates.append(f"row {row} {get_column_letter(col)}")

    excluded_changed = [
        f"{change.get('group')}/{change.get('zone')}/{change.get('pickup_date')}"
        for change in changes
        if normalize_code(change.get("group")) in excluded_groups
    ]
    missing_benchmark = sorted({
        str(change.get("scenario_id") or change.get("cell"))
        for change in changes
        if change.get("recommendation_type") != "group_parity"
        and parse_number(change.get("benchmark_rate")) is None
    })
    below_floor_changes: list[str] = []
    city_airport_cap_violations: list[str] = []
    for change in changes:
        minimum_rate = parse_number(change.get("minimum_rate_pln_day"))
        new_rate = parse_number(change.get("new_rate"))
        if minimum_rate is not None and new_rate is not None and new_rate < minimum_rate:
            below_floor_changes.append(
                f"{change.get('group')}/{change.get('zone')}/{change.get('pickup_date')} "
                f"{change.get('duration_band')}: {format_rate_for_comment(new_rate)} < "
                f"{format_rate_for_comment(minimum_rate)}"
            )
        maximum_city_rate = parse_number(change.get("city_top1_airport_max_rate"))
        if (
            change.get("city_top1_airport_cap_active")
            and maximum_city_rate is not None
            and new_rate is not None
            and new_rate > maximum_city_rate + 0.001
        ):
            city_airport_cap_violations.append(
                f"{change.get('group')}/{change.get('zone')}/{change.get('pickup_date')} "
                f"{change.get('duration_band')}: {format_rate_for_comment(new_rate)} > "
                f"{format_rate_for_comment(maximum_city_rate)}"
            )

    unachievable_targets = sorted({
        f"{change.get('group')}/{change.get('zone')}/{change.get('pickup_date')} {change.get('duration_band')}"
        for change in changes
        if change.get("target_achievable") is False
    })
    aggregation_conflicts = sorted({
        f"{change.get('zone')}/{change.get('pickup_date')} {change.get('duration_band')}"
        for change in changes
        if change.get("aggregation_conflict")
    })
    incomplete_duration_coverage = sorted({
        f"{change.get('zone')}/{change.get('pickup_date')} {change.get('duration_band')}: "
        + ",".join(str(item) for item in change.get("missing_duration_days", []))
        for change in changes
        if change.get("duration_band_coverage_complete") is False
    })

    expansion_summary = expansion_summary or {}
    missing_groups_after_expansion = [
        str(item)
        for item in expansion_summary.get("missing_source_groups_after_expansion", [])
        if str(item).strip()
    ]
    restored_group_zones = [
        str(item)
        for item in expansion_summary.get("restored_group_zones", [])
        if str(item).strip()
    ]
    missing_group_zones_after_expansion = [
        str(item)
        for item in expansion_summary.get("missing_source_group_zones_after_expansion", [])
        if str(item).strip()
    ]
    fixed_rate_group_issues = (
        []
        if assume_fixed_rate_changes_applied
        else find_fixed_rate_group_issues(ws, config, duration_columns)
    )
    mirrored_rate_group_issues = (
        []
        if assume_fixed_rate_changes_applied
        else find_mirrored_rate_group_issues(ws, config, duration_columns)
    )

    return [
        ["Wiersze danych w Sheet1", "INFO", data_rows, ""],
        ["Zmienione komorki stawek", "INFO", len(changes), ""],
        ["Brakujace grupy po ekspansji dat", get_validation_status(len(missing_groups_after_expansion)), len(missing_groups_after_expansion), first_items(missing_groups_after_expansion)],
        ["Brakujace Group + Zone po ekspansji dat", get_validation_status(len(missing_group_zones_after_expansion)), len(missing_group_zones_after_expansion), first_items(missing_group_zones_after_expansion)],
        ["Odtworzone Group + Zone po ekspansji dat", "INFO", len(restored_group_zones), first_items(restored_group_zones)],
        ["Pominiete rekomendacje", get_validation_status(len(skipped_targets), warning=True), len(skipped_targets), first_items([str(item.get("skip_reason", "")) for item in skipped_targets])],
        ["Booking end date = Pickup end date", get_validation_status(len(booking_mismatch)), len(booking_mismatch), first_items(booking_mismatch)],
        ["Pickup end date = Pickup start date", get_validation_status(len(pickup_mismatch)), len(pickup_mismatch), first_items(pickup_mismatch)],
        ["Duplikaty Group + Zone + Pickup date", get_validation_status(len(duplicates), warning=True), len(duplicates), first_items(duplicates)],
        ["Puste stawki w kolumnach duration", get_validation_status(len(missing_rates)), len(missing_rates), first_items(missing_rates)],
        ["Kompletne klasy ze stawkami stalymi", get_validation_status(len(fixed_rate_group_issues)), len(fixed_rate_group_issues), first_items(fixed_rate_group_issues)],
        ["Kompletne klasy lustrzane", get_validation_status(len(mirrored_rate_group_issues)), len(mirrored_rate_group_issues), first_items(mirrored_rate_group_issues)],
        ["Zmienione stawki ponizej floor cenowego", get_validation_status(len(below_floor_changes)), len(below_floor_changes), first_items(below_floor_changes)],
        ["Stawki miejskie powyzej 130% ceny lotniskowej", get_validation_status(len(city_airport_cap_violations)), len(city_airport_cap_violations), first_items(city_airport_cap_violations)],
        ["Cele rankingowe nieosiagalne po finalnej stawce", get_validation_status(len(unachievable_targets), warning=True), len(unachievable_targets), first_items(unachievable_targets)],
        ["Sprzeczne rekomendacje w przedziale duration", get_validation_status(len(aggregation_conflicts), warning=True), len(aggregation_conflicts), first_items(aggregation_conflicts)],
        ["Niepelne pokrycie przedzialu duration", get_validation_status(len(incomplete_duration_coverage), warning=True), len(incomplete_duration_coverage), first_items(incomplete_duration_coverage)],
        ["Zmienione grupy wykluczone", get_validation_status(len(excluded_changed)), len(excluded_changed), first_items(excluded_changed)],
        ["Zmienione rekomendacje bez ceny benchmarku", get_validation_status(len(missing_benchmark), warning=True), len(missing_benchmark), first_items(missing_benchmark)],
    ]


def write_validation_sheet(
    workbook: Any,
    source_ws: Any,
    config: dict[str, Any],
    duration_columns: dict[int, tuple[int, str, int, int]],
    changes: list[dict[str, Any]],
    skipped_targets: list[dict[str, Any]],
    expansion_summary: dict[str, Any] | None = None,
    validation_rows: list[list[Any]] | None = None,
) -> None:
    sheet_name = str(config.get("validation_sheet") or "").strip()
    if not sheet_name:
        return
    headers = ["Kontrola", "Status", "Liczba problemow", "Szczegoly"]
    rows = validation_rows if validation_rows is not None else build_validation_rows(
        source_ws,
        config,
        duration_columns,
        changes,
        skipped_targets,
        expansion_summary,
    )
    widths = {"Kontrola": 48, "Status": 14, "Liczba problemow": 18, "Szczegoly": 90}
    ws = write_table_sheet(workbook, sheet_name, workbook.index(source_ws) + 4, headers, rows, widths)
    for row in range(2, ws.max_row + 1):
        status_cell = ws.cell(row, 2)
        if status_cell.value == "OK":
            status_cell.fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
        elif status_cell.value == "WARNING":
            status_cell.fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
        elif status_cell.value == "FAIL":
            status_cell.fill = PatternFill(fill_type="solid", fgColor="FFC7CE")


def average(values: list[float]) -> float | None:
    if not values:
        return None
    return sum(values) / len(values)


def build_broker_markup_observations(
    changes: list[dict[str, Any]],
    config: dict[str, Any],
    baseline_confirmation: dict[str, Any] | None = None,
) -> dict[str, Any]:
    settings = config.get("broker_markup_learning") or {}
    if baseline_confirmation and not baseline_confirmation.get("calibration_eligible"):
        return {
            "enabled": False,
            "count": 0,
            "reason": "baseline_not_confirmed",
            "observations": [],
        }
    if settings.get("enabled") is False:
        return {
            "enabled": False,
            "count": 0,
            "observations": [],
        }

    min_multiplier = parse_number(settings.get("min_multiplier")) or 1
    max_multiplier = parse_number(settings.get("max_multiplier")) or 1.25
    observations_by_key: dict[tuple[Any, ...], dict[str, Any]] = {}
    ambiguous_observation_keys: set[tuple[Any, ...]] = set()

    for change in changes:
        old_rate = parse_number(change.get("old_rate"))
        mm_rate = parse_number(change.get("mm_rate"))
        if old_rate is None or old_rate <= 0 or mm_rate is None or mm_rate <= 0:
            continue
        if parse_number(change.get("group_adjustment_pln_day")):
            continue

        multiplier = mm_rate / old_rate
        if multiplier < min_multiplier or multiplier > max_multiplier:
            continue

        observation = {
            "location": change.get("location", ""),
            "zone": change.get("zone", ""),
            "group": change.get("group", ""),
            "pickup_date": change.get("pickup_date", ""),
            "duration_days": change.get("duration_days"),
            "duration_band": change.get("duration_band"),
            "old_import_rate_pln_day": old_rate,
            "live_mm_rate_pln_day": mm_rate,
            "observed_multiplier": round(multiplier, 6),
            "observed_markup_percent": round((multiplier - 1) * 100, 2),
        }
        dedupe_key = (
            normalize_key(observation["location"]),
            observation["pickup_date"],
            observation["duration_days"],
        )
        if dedupe_key in ambiguous_observation_keys:
            continue
        existing = observations_by_key.get(dedupe_key)
        if existing and abs(float(existing["observed_multiplier"]) - multiplier) > 0.02:
            observations_by_key.pop(dedupe_key, None)
            ambiguous_observation_keys.add(dedupe_key)
            continue
        observations_by_key.setdefault(dedupe_key, observation)

    observations = list(observations_by_key.values())

    by_location: dict[str, list[float]] = defaultdict(list)
    by_duration: dict[str, list[float]] = defaultdict(list)
    by_location_duration: dict[str, dict[str, list[float]]] = defaultdict(lambda: defaultdict(list))
    for observation in observations:
        multiplier = float(observation["observed_multiplier"])
        location = str(observation.get("location") or "").strip()
        duration = str(observation.get("duration_days") or "").strip()
        if location:
            by_location[location].append(multiplier)
        if duration:
            by_duration[duration].append(multiplier)
        if location and duration:
            by_location_duration[location][duration].append(multiplier)

    def summarize(values: list[float]) -> dict[str, Any]:
        result = average(values)
        robust_result = median(values) if values else None
        return {
            "count": len(values),
            "average_multiplier": round(result, 6) if result is not None else None,
            "average_markup_percent": round((result - 1) * 100, 2) if result is not None else None,
            "median_multiplier": round(robust_result, 6) if robust_result is not None else None,
            "median_markup_percent": round((robust_result - 1) * 100, 2) if robust_result is not None else None,
        }

    all_values = [float(item["observed_multiplier"]) for item in observations]
    return {
        "enabled": True,
        "count": len(observations),
        "ambiguous_observation_count": len(ambiguous_observation_keys),
        **summarize(all_values),
        "by_location": {
            location: summarize(values)
            for location, values in sorted(by_location.items())
        },
        "by_duration": {
            duration: summarize(values)
            for duration, values in sorted(by_duration.items(), key=lambda item: int(item[0]) if item[0].isdigit() else 0)
        },
        "by_location_duration": {
            location: {
                duration: summarize(values)
                for duration, values in sorted(durations.items(), key=lambda item: int(item[0]) if item[0].isdigit() else 0)
            }
            for location, durations in sorted(by_location_duration.items())
        },
        "observations": observations[:100],
    }


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def build_targets(
    recommendations: list[dict[str, Any]],
    duration_columns: dict[int, tuple[int, str, int, int]],
    config: dict[str, Any],
) -> tuple[dict[str, dict[date, list[dict[str, Any]]]], list[dict[str, Any]]]:
    location_zones = {
        normalize_key(location): [normalize_code(zone) for zone in zones]
        for location, zones in (config.get("location_zones") or {}).items()
    }
    targets: dict[str, dict[date, list[dict[str, Any]]]] = defaultdict(lambda: defaultdict(list))
    candidates_by_cell: dict[tuple[str, date, int], list[dict[str, Any]]] = defaultdict(list)
    skipped: list[dict[str, Any]] = []
    protected_periods = get_protected_rate_periods(config)

    for item in recommendations:
        action = item.get("action")
        if action not in {"increase", "decrease", "hold"}:
            continue

        suggested_rate = parse_number(item.get("suggested_rate_pln_day"))
        maximum_rate = parse_number(item.get("maximum_import_rate_pln_day"))
        if maximum_rate is None and action in {"increase", "decrease"}:
            maximum_rate = suggested_rate
        rental_days = parse_number(item.get("rental_days"))
        target_date = parse_date_value(item.get("start_date") or item.get("pickup_date"))
        location = normalize_key(item.get("location"))
        zones = location_zones.get(location, [])

        if rental_days is None or target_date is None or not zones or (action in {"increase", "decrease"} and suggested_rate is None):
            if action in {"increase", "decrease"}:
                skipped.append({**item, "skip_reason": "Missing suggested rate, rental days, start date, or location zone mapping."})
            continue

        if date_is_rate_protected(target_date, protected_periods):
            if action in {"increase", "decrease"}:
                skipped.append({
                    **item,
                    "skip_reason": f"Pickup start date {target_date.isoformat()} is protected; the existing rate was preserved.",
                })
            continue

        duration = int(rental_days)
        max_duration = int(parse_number(config.get("max_recommendation_duration_days")) or 0)
        if max_duration > 0 and duration > max_duration:
            if action in {"increase", "decrease"}:
                skipped.append({
                    **item,
                    "skip_reason": f"Maximum recommendation duration is {max_duration} days; {duration} days was ignored.",
                })
            continue
        duration_column = duration_columns.get(duration)
        if not duration_column:
            skipped.append({**item, "skip_reason": f"No Excel duration column for {duration} days."})
            continue

        col, duration_band, duration_min_days, duration_max_days = duration_column
        for zone in zones:
            broker_multiplier = parse_number(item.get("broker_markup_multiplier")) or 1
            site_cap_rate = parse_number(item.get("site_cap_rate_pln_day"))
            if site_cap_rate is None and maximum_rate is not None:
                site_cap_rate = round(maximum_rate * broker_multiplier, 2)
            candidates_by_cell[(zone, target_date, col)].append({
                **item,
                "zone": zone,
                "target_date": target_date,
                "rate_col": col,
                "duration_band": duration_band,
                "duration_min_days": duration_min_days,
                "duration_max_days": duration_max_days,
                "suggested_rate_pln_day": suggested_rate,
                "constraint_import_rate_pln_day": maximum_rate,
                "site_cap_rate_pln_day": site_cap_rate,
            })

    for (zone, target_date, _col), candidates in candidates_by_cell.items():
        active = [item for item in candidates if item.get("action") in {"increase", "decrease"}]
        if not active:
            continue

        invalid_constraints = [
            item for item in candidates
            if item.get("data_quality_status") not in {None, "", "ok"}
            or parse_number(item.get("constraint_import_rate_pln_day")) is None
        ]
        if invalid_constraints:
            skipped.append({
                **active[0],
                "skip_reason": (
                    "Duration-band update blocked because at least one covered scenario has missing or invalid constraint data: "
                    + ", ".join(
                        f"{item.get('rental_days')}d/{item.get('data_quality_status') or 'missing cap'}"
                        for item in invalid_constraints[:8]
                    )
                ),
            })
            continue

        controlling = min(candidates, key=lambda item: float(item["constraint_import_rate_pln_day"]))
        representative = min(active, key=lambda item: float(item["suggested_rate_pln_day"]))
        aggregate_rate = float(controlling["constraint_import_rate_pln_day"])
        source_actions = sorted({str(item.get("action")) for item in active})
        source_types = sorted({str(item.get("recommendation_type") or "") for item in active if item.get("recommendation_type")})
        source_locations = sorted({str(item.get("location") or "") for item in candidates if item.get("location")})
        covered_duration_days = sorted({int(parse_number(item.get("rental_days")) or 0) for item in candidates})
        expected_duration_days = list(range(int(representative["duration_min_days"]), int(representative["duration_max_days"]) + 1))
        missing_duration_days = [duration for duration in expected_duration_days if duration not in covered_duration_days]
        if missing_duration_days:
            skipped.append({
                **representative,
                "skip_reason": (
                    f"Duration band {representative['duration_band']} lacks scenarios for "
                    + ",".join(str(item) for item in missing_duration_days)
                    + "; increases are capped at the current workbook rate."
                ),
            })
        targets[zone][target_date].append({
            **representative,
            "zone": zone,
            "target_date": target_date,
            "rate_col": representative["rate_col"],
            "duration_band": representative["duration_band"],
            "duration_min_days": representative["duration_min_days"],
            "duration_max_days": representative["duration_max_days"],
            "suggested_rate_pln_day": aggregate_rate,
            "site_cap_rate_pln_day": controlling.get("site_cap_rate_pln_day"),
            "broker_markup_multiplier": controlling.get("broker_markup_multiplier"),
            "broker_markup_percent": controlling.get("broker_markup_percent"),
            "broker_markup_source": controlling.get("broker_markup_source", ""),
            "constraint_items": candidates,
            "source_decision_count": len(candidates),
            "source_active_count": len(active),
            "source_actions": source_actions,
            "source_recommendation_types": source_types,
            "source_locations": source_locations,
            "controlling_duration_days": controlling.get("rental_days"),
            "aggregation_conflict": len(source_actions) > 1,
            "covered_duration_days": covered_duration_days,
            "missing_duration_days": missing_duration_days,
            "duration_band_coverage_complete": not missing_duration_days,
        })

    return targets, skipped


def get_target_acceptance_key(target: dict[str, Any]) -> tuple[str, str, str, str]:
    return get_acceptance_key(
        target.get("scenario_id", ""),
        target.get("location", ""),
        target.get("target_date"),
        target.get("duration_band", ""),
    )


def target_has_inline_acceptance(target: dict[str, Any]) -> bool:
    return is_accepted_value(target.get("accepted"))


def filter_targets_by_acceptance(
    targets: dict[str, dict[date, list[dict[str, Any]]]],
    accepted_keys: set[tuple[str, str, str, str]],
) -> tuple[dict[str, dict[date, list[dict[str, Any]]]], int, int]:
    filtered: dict[str, dict[date, list[dict[str, Any]]]] = defaultdict(lambda: defaultdict(list))
    accepted_count = 0
    filtered_count = 0

    for zone, targets_by_date in targets.items():
        for target_date, row_targets in targets_by_date.items():
            for target in row_targets:
                if target_has_inline_acceptance(target) or get_target_acceptance_key(target) in accepted_keys:
                    filtered[zone][target_date].append(target)
                    accepted_count += 1
                else:
                    filtered_count += 1

    return filtered, accepted_count, filtered_count


def build_group_price_parity_scope(
    targets: dict[str, dict[date, list[dict[str, Any]]]],
) -> set[tuple[str, date, int]]:
    scope: set[tuple[str, date, int]] = set()
    for zone, targets_by_date in targets.items():
        normalized_zone = normalize_code(zone)
        for target_date, row_targets in targets_by_date.items():
            for target in row_targets:
                rate_col = target.get("rate_col")
                if isinstance(target_date, date) and isinstance(rate_col, int):
                    scope.add((normalized_zone, target_date, rate_col))
    return scope


def build_city_top1_airport_rate_caps(
    ws: Any,
    config: dict[str, Any],
    targets: dict[str, dict[date, list[dict[str, Any]]]],
) -> tuple[dict[tuple[str, date, int], dict[str, Any]], list[dict[str, Any]]]:
    cap_rule = get_city_top1_airport_cap(config)
    city_zone_airport_zones = {
        normalize_code(city_zone): [normalize_code(zone) for zone in airport_zones if normalize_code(zone)]
        for city_zone, airport_zones in (config.get("city_zone_airport_zones") or {}).items()
        if normalize_code(city_zone)
    }
    if cap_rule is None or not city_zone_airport_zones:
        return {}, []

    multiplier, recommendation_types = cap_rule
    columns = config["columns"]
    data_start_row = int(config["data_start_row"])
    group_col = int(columns["group"])
    zone_col = int(columns["zone"])
    pickup_col = int(columns["pickup_start_date"])
    rows_by_zone_date: dict[tuple[str, date], dict[str, int]] = defaultdict(dict)
    for row in range(data_start_row, ws.max_row + 1):
        group = normalize_code(ws.cell(row, group_col).value)
        zone = normalize_code(ws.cell(row, zone_col).value)
        pickup_date = parse_date_value(ws.cell(row, pickup_col).value)
        if group and zone and pickup_date is not None:
            rows_by_zone_date[(zone, pickup_date)][group] = row

    targets_by_cell: dict[tuple[str, date, int], dict[str, Any]] = {}
    for zone, targets_by_date in targets.items():
        normalized_zone = normalize_code(zone)
        for target_date, row_targets in targets_by_date.items():
            for target in row_targets:
                rate_col = target.get("rate_col")
                if isinstance(target_date, date) and isinstance(rate_col, int):
                    targets_by_cell[(normalized_zone, target_date, rate_col)] = target

    parity = get_group_price_parity(config)
    tracked_groups = set(parity[0]) | set(parity[1]) if parity else set()
    zone_labels = {
        normalize_code(zone): str(label)
        for zone, label in (config.get("zone_location_labels") or {}).items()
        if normalize_code(zone)
    }
    caps: dict[tuple[str, date, int], dict[str, Any]] = {}
    missing_references: list[dict[str, Any]] = []

    for city_zone, targets_by_date in targets.items():
        normalized_city_zone = normalize_code(city_zone)
        airport_zones = city_zone_airport_zones.get(normalized_city_zone, [])
        if not airport_zones:
            continue
        for target_date, row_targets in targets_by_date.items():
            for city_target in row_targets:
                rate_col = city_target.get("rate_col")
                if not isinstance(rate_col, int) or not target_matches_recommendation_types(city_target, recommendation_types):
                    continue

                airport_options: list[dict[str, Any]] = []
                for airport_zone in airport_zones:
                    airport_rows = rows_by_zone_date.get((airport_zone, target_date), {})
                    if not airport_rows:
                        continue
                    candidate_groups = tracked_groups or {
                        group for group in airport_rows if not group_is_excluded(group, config)
                    }
                    airport_target = targets_by_cell.get((airport_zone, target_date, rate_col))
                    projected_rates: dict[str, float] = {}

                    if airport_target is not None:
                        reference_group = next(
                            (
                                group
                                for group in (parity[0] if parity else sorted(candidate_groups))
                                if group in airport_rows
                            ),
                            None,
                        )
                        if reference_group is not None:
                            reference_adjustment = get_group_rate_adjustment(reference_group, config)
                            current_rate = parse_number(ws.cell(airport_rows[reference_group], rate_col).value)
                            current_base = None if current_rate is None else current_rate - reference_adjustment
                            projected_base, _, _, _ = calculate_target_base_rate(airport_target, current_base, config)
                            for group in candidate_groups:
                                if group in airport_rows:
                                    projected_rates[group] = projected_base + get_group_rate_adjustment(group, config)
                    else:
                        for group in candidate_groups:
                            row = airport_rows.get(group)
                            rate = parse_number(ws.cell(row, rate_col).value) if row is not None else None
                            if rate is not None:
                                projected_rates[group] = rate

                    base_cap_candidates = [
                        rate * multiplier - get_group_rate_adjustment(group, config)
                        for group, rate in projected_rates.items()
                    ]
                    if not base_cap_candidates:
                        continue
                    airport_options.append({
                        "airport_zone": airport_zone,
                        "airport_location": zone_labels.get(airport_zone, airport_zone),
                        "airport_rates_by_group": projected_rates,
                        "base_rate_cap_pln_day": math.floor((min(base_cap_candidates) + 1e-9) * 100) / 100,
                        "max_multiplier": multiplier,
                    })

                if airport_options:
                    caps[(normalized_city_zone, target_date, rate_col)] = min(
                        airport_options,
                        key=lambda item: float(item["base_rate_cap_pln_day"]),
                    )
                else:
                    missing_references.append({
                        "city_zone": normalized_city_zone,
                        "pickup_date": target_date.isoformat(),
                        "rate_col": rate_col,
                        "duration_band": city_target.get("duration_band", ""),
                        "airport_zones": airport_zones,
                    })

    return caps, missing_references


def find_city_top1_airport_cap_violations(
    ws: Any,
    config: dict[str, Any],
    caps: dict[tuple[str, date, int], dict[str, Any]],
) -> list[str]:
    if not caps:
        return []
    columns = config["columns"]
    data_start_row = int(config["data_start_row"])
    group_col = int(columns["group"])
    zone_col = int(columns["zone"])
    pickup_col = int(columns["pickup_start_date"])
    caps_by_zone_date: dict[tuple[str, date], list[tuple[int, dict[str, Any]]]] = defaultdict(list)
    for (zone, pickup_date, rate_col), cap_info in caps.items():
        caps_by_zone_date[(zone, pickup_date)].append((rate_col, cap_info))
    violations: list[str] = []
    for row in range(data_start_row, ws.max_row + 1):
        group = normalize_code(ws.cell(row, group_col).value)
        zone = normalize_code(ws.cell(row, zone_col).value)
        pickup_date = parse_date_value(ws.cell(row, pickup_col).value)
        if not group or not zone or pickup_date is None or group_is_excluded(group, config):
            continue
        for rate_col, cap_info in caps_by_zone_date.get((zone, pickup_date), []):
            rate = parse_number(ws.cell(row, rate_col).value)
            maximum_rate = float(cap_info["base_rate_cap_pln_day"]) + get_group_rate_adjustment(group, config)
            if rate is not None and rate > maximum_rate + 0.001:
                violations.append(
                    f"{group}/{zone}/{pickup_date.isoformat()} {get_column_letter(rate_col)}: "
                    f"{format_rate_for_comment(rate)} > {format_rate_for_comment(maximum_rate)}"
                )
    return violations


def snapshot_row(ws: Any, row: int, max_col: int) -> dict[str, Any]:
    row_dimension = ws.row_dimensions[row]
    return {
        "height": row_dimension.height,
        "hidden": row_dimension.hidden,
        "outline_level": row_dimension.outlineLevel,
        "cells": [
            {
                "value": ws.cell(row, col).value,
                "style": copy(ws.cell(row, col)._style) if ws.cell(row, col).has_style else None,
                "hyperlink": copy(ws.cell(row, col).hyperlink) if ws.cell(row, col).hyperlink else None,
            }
            for col in range(1, max_col + 1)
        ],
    }


def write_row_snapshot(ws: Any, row: int, snapshot: dict[str, Any]) -> None:
    ws.row_dimensions[row].height = snapshot.get("height")
    ws.row_dimensions[row].hidden = bool(snapshot.get("hidden"))
    ws.row_dimensions[row].outlineLevel = int(snapshot.get("outline_level") or 0)
    for col, cell_snapshot in enumerate(snapshot["cells"], start=1):
        cell = ws.cell(row, col)
        cell.value = cell_snapshot["value"]
        if cell_snapshot["style"] is not None:
            cell._style = copy(cell_snapshot["style"])
        if cell_snapshot["hyperlink"]:
            cell._hyperlink = copy(cell_snapshot["hyperlink"])


def ensure_fixed_rate_group_rows(
    ws: Any,
    config: dict[str, Any],
    duration_columns: dict[int, tuple[int, str, int, int]],
    dry_run: bool,
) -> dict[str, Any]:
    fixed_groups = get_fixed_rate_groups(config)
    if not fixed_groups:
        return {
            "enabled": False,
            "added_row_count": 0,
            "updated_row_count": 0,
            "updated_cell_count": 0,
            "groups": {},
        }

    columns = config["columns"]
    data_start_row = int(config["data_start_row"])
    group_col = int(columns["group"])
    zone_col = int(columns["zone"])
    pickup_col = int(columns["pickup_start_date"])
    rate_cols = sorted({value[0] for value in duration_columns.values()})
    duration_band_by_col = {
        value[0]: value[1]
        for value in duration_columns.values()
    }
    protected_periods = get_protected_rate_periods(config)
    max_col = ws.max_column
    rows_by_group_key: dict[tuple[str, str, date], list[int]] = defaultdict(list)

    for row in range(data_start_row, ws.max_row + 1):
        group = normalize_code(ws.cell(row, group_col).value)
        zone = normalize_code(ws.cell(row, zone_col).value)
        pickup_date = parse_date_value(ws.cell(row, pickup_col).value)
        if group and zone and pickup_date is not None:
            rows_by_group_key[(group, zone, pickup_date)].append(row)

    added_row_count = 0
    updated_row_count = 0
    updated_cell_count = 0
    group_summaries: dict[str, dict[str, Any]] = {}
    next_row = ws.max_row + 1

    for group, settings in fixed_groups.items():
        template_group = settings["template_group"]
        template_rows = {
            (zone, pickup_date): rows[0]
            for (row_group, zone, pickup_date), rows in rows_by_group_key.items()
            if row_group == template_group
        }
        if not template_rows:
            raise ValueError(
                f"Fixed-rate group {group} cannot be created because template group {template_group} is missing."
            )

        group_added = 0
        group_updated = 0
        existing_group_rows = [
            (pickup_date, row)
            for (row_group, _zone, pickup_date), rows in rows_by_group_key.items()
            if row_group == group
            for row in rows
        ]
        protected_row_count = 0
        for pickup_date, existing_row in existing_group_rows:
            if date_is_rate_protected(pickup_date, protected_periods):
                protected_row_count += 1
                continue
            row_changed = False
            for col in rate_cols:
                fixed_rate = get_fixed_rate_for_duration_band(settings, duration_band_by_col[col])
                current_rate = parse_number(ws.cell(existing_row, col).value)
                if current_rate is not None and abs(current_rate - fixed_rate) < 0.001:
                    continue
                row_changed = True
                updated_cell_count += 1
                if not dry_run:
                    ws.cell(existing_row, col).value = int(fixed_rate) if fixed_rate.is_integer() else fixed_rate
            if row_changed:
                group_updated += 1
                updated_row_count += 1

        for (zone, pickup_date), template_row in sorted(template_rows.items()):
            if rows_by_group_key.get((group, zone, pickup_date)):
                continue
            group_added += 1
            added_row_count += 1
            if dry_run:
                continue
            write_row_snapshot(ws, next_row, snapshot_row(ws, template_row, max_col))
            ws.cell(next_row, group_col).value = group
            for col in rate_cols:
                fixed_rate = get_fixed_rate_for_duration_band(settings, duration_band_by_col[col])
                ws.cell(next_row, col).value = int(fixed_rate) if fixed_rate.is_integer() else fixed_rate
            rows_by_group_key[(group, zone, pickup_date)].append(next_row)
            next_row += 1

        group_summaries[group] = {
            "rate_pln_day": settings.get("rate_pln_day"),
            "rates_by_duration_band": settings.get("rates_by_duration_band") or {},
            "template_group": template_group,
            "expected_row_count": len(template_rows),
            "added_row_count": group_added,
            "updated_row_count": group_updated,
            "protected_existing_row_count": protected_row_count,
        }

    return {
        "enabled": True,
        "added_row_count": added_row_count,
        "updated_row_count": updated_row_count,
        "updated_cell_count": updated_cell_count,
        "groups": group_summaries,
    }


def find_fixed_rate_group_issues(
    ws: Any,
    config: dict[str, Any],
    duration_columns: dict[int, tuple[int, str, int, int]],
) -> list[str]:
    fixed_groups = get_fixed_rate_groups(config)
    if not fixed_groups:
        return []

    columns = config["columns"]
    data_start_row = int(config["data_start_row"])
    group_col = int(columns["group"])
    zone_col = int(columns["zone"])
    pickup_col = int(columns["pickup_start_date"])
    rate_cols = sorted({value[0] for value in duration_columns.values()})
    duration_band_by_col = {
        value[0]: value[1]
        for value in duration_columns.values()
    }
    protected_periods = get_protected_rate_periods(config)
    rows_by_group_key: dict[tuple[str, str, date], list[int]] = defaultdict(list)

    for row in range(data_start_row, ws.max_row + 1):
        group = normalize_code(ws.cell(row, group_col).value)
        zone = normalize_code(ws.cell(row, zone_col).value)
        pickup_date = parse_date_value(ws.cell(row, pickup_col).value)
        if group and zone and pickup_date is not None:
            rows_by_group_key[(group, zone, pickup_date)].append(row)

    issues: list[str] = []
    for group, settings in fixed_groups.items():
        template_group = settings["template_group"]
        template_keys = {
            (zone, pickup_date)
            for row_group, zone, pickup_date in rows_by_group_key
            if row_group == template_group
        }
        for zone, pickup_date in sorted(template_keys):
            fixed_rows = rows_by_group_key.get((group, zone, pickup_date), [])
            if not fixed_rows:
                issues.append(f"brak {group}/{zone}/{pickup_date.isoformat()}")
        fixed_rows = [
            (zone, pickup_date, row)
            for (row_group, zone, pickup_date), rows in rows_by_group_key.items()
            if row_group == group
            for row in rows
        ]
        for zone, pickup_date, row in sorted(fixed_rows):
            if date_is_rate_protected(pickup_date, protected_periods):
                continue
            for col in rate_cols:
                expected_rate = get_fixed_rate_for_duration_band(settings, duration_band_by_col[col])
                rate = parse_number(ws.cell(row, col).value)
                if rate is None or abs(rate - expected_rate) >= 0.001:
                    issues.append(
                        f"{group}/{zone}/{pickup_date.isoformat()} {get_column_letter(col)}: "
                        f"{format_rate_for_comment(rate)} != {format_rate_for_comment(expected_rate)}"
                    )
    return issues


def ensure_mirrored_rate_group_rows(
    ws: Any,
    config: dict[str, Any],
    duration_columns: dict[int, tuple[int, str, int, int]],
    dry_run: bool,
) -> dict[str, Any]:
    mirrored_groups = get_mirrored_rate_groups(config)
    if not mirrored_groups:
        return {
            "enabled": False,
            "added_row_count": 0,
            "updated_row_count": 0,
            "updated_cell_count": 0,
            "groups": {},
        }

    columns = config["columns"]
    data_start_row = int(config["data_start_row"])
    group_col = int(columns["group"])
    zone_col = int(columns["zone"])
    pickup_col = int(columns["pickup_start_date"])
    rate_cols = sorted({value[0] for value in duration_columns.values()})
    protected_periods = get_protected_rate_periods(config)
    max_col = ws.max_column
    rows_by_group_key: dict[tuple[str, str, date], list[int]] = defaultdict(list)

    for row in range(data_start_row, ws.max_row + 1):
        group = normalize_code(ws.cell(row, group_col).value)
        zone = normalize_code(ws.cell(row, zone_col).value)
        pickup_date = parse_date_value(ws.cell(row, pickup_col).value)
        if group and zone and pickup_date is not None:
            rows_by_group_key[(group, zone, pickup_date)].append(row)

    added_row_count = 0
    updated_row_count = 0
    updated_cell_count = 0
    group_summaries: dict[str, dict[str, Any]] = {}
    next_row = ws.max_row + 1

    for group, settings in mirrored_groups.items():
        template_group = settings["template_group"]
        template_rows = {
            (zone, pickup_date): rows[0]
            for (row_group, zone, pickup_date), rows in rows_by_group_key.items()
            if row_group == template_group
        }
        if not template_rows:
            raise ValueError(
                f"Mirrored-rate group {group} cannot be created because template group {template_group} is missing."
            )

        group_added = 0
        group_updated = 0
        protected_row_count = 0
        existing_group_rows = [
            (zone, pickup_date, row)
            for (row_group, zone, pickup_date), rows in rows_by_group_key.items()
            if row_group == group
            for row in rows
        ]
        for zone, pickup_date, existing_row in existing_group_rows:
            if date_is_rate_protected(pickup_date, protected_periods):
                protected_row_count += 1
                continue
            template_row = template_rows.get((zone, pickup_date))
            if template_row is None:
                continue
            row_changed = False
            for col in rate_cols:
                template_rate = parse_number(ws.cell(template_row, col).value)
                current_rate = parse_number(ws.cell(existing_row, col).value)
                if template_rate is None or (current_rate is not None and abs(current_rate - template_rate) < 0.001):
                    continue
                row_changed = True
                updated_cell_count += 1
                if not dry_run:
                    ws.cell(existing_row, col).value = int(template_rate) if template_rate.is_integer() else template_rate
            if row_changed:
                group_updated += 1
                updated_row_count += 1

        for (zone, pickup_date), template_row in sorted(template_rows.items()):
            if rows_by_group_key.get((group, zone, pickup_date)):
                continue
            group_added += 1
            added_row_count += 1
            if dry_run:
                continue
            write_row_snapshot(ws, next_row, snapshot_row(ws, template_row, max_col))
            ws.cell(next_row, group_col).value = group
            rows_by_group_key[(group, zone, pickup_date)].append(next_row)
            next_row += 1

        group_summaries[group] = {
            "template_group": template_group,
            "expected_row_count": len(template_rows),
            "added_row_count": group_added,
            "updated_row_count": group_updated,
            "protected_existing_row_count": protected_row_count,
        }

    return {
        "enabled": True,
        "added_row_count": added_row_count,
        "updated_row_count": updated_row_count,
        "updated_cell_count": updated_cell_count,
        "groups": group_summaries,
    }


def find_mirrored_rate_group_issues(
    ws: Any,
    config: dict[str, Any],
    duration_columns: dict[int, tuple[int, str, int, int]],
) -> list[str]:
    mirrored_groups = get_mirrored_rate_groups(config)
    if not mirrored_groups:
        return []

    columns = config["columns"]
    data_start_row = int(config["data_start_row"])
    group_col = int(columns["group"])
    zone_col = int(columns["zone"])
    pickup_col = int(columns["pickup_start_date"])
    rate_cols = sorted({value[0] for value in duration_columns.values()})
    protected_periods = get_protected_rate_periods(config)
    rows_by_group_key: dict[tuple[str, str, date], list[int]] = defaultdict(list)

    for row in range(data_start_row, ws.max_row + 1):
        group = normalize_code(ws.cell(row, group_col).value)
        zone = normalize_code(ws.cell(row, zone_col).value)
        pickup_date = parse_date_value(ws.cell(row, pickup_col).value)
        if group and zone and pickup_date is not None:
            rows_by_group_key[(group, zone, pickup_date)].append(row)

    issues: list[str] = []
    for group, settings in mirrored_groups.items():
        template_group = settings["template_group"]
        template_rows = {
            (zone, pickup_date): rows[0]
            for (row_group, zone, pickup_date), rows in rows_by_group_key.items()
            if row_group == template_group
        }
        for zone, pickup_date in sorted(template_rows):
            mirror_rows = rows_by_group_key.get((group, zone, pickup_date), [])
            if not mirror_rows:
                issues.append(f"brak {group}/{zone}/{pickup_date.isoformat()}")
                continue
            if date_is_rate_protected(pickup_date, protected_periods):
                continue
            template_row = template_rows[(zone, pickup_date)]
            for mirror_row in mirror_rows:
                for col in rate_cols:
                    template_rate = parse_number(ws.cell(template_row, col).value)
                    mirror_rate = parse_number(ws.cell(mirror_row, col).value)
                    if template_rate is None or mirror_rate is None or abs(template_rate - mirror_rate) >= 0.001:
                        issues.append(
                            f"{group}/{zone}/{pickup_date.isoformat()} {get_column_letter(col)}: "
                            f"{format_rate_for_comment(mirror_rate)} != {template_group} "
                            f"{format_rate_for_comment(template_rate)}"
                        )
    return issues


def expand_pickup_date_rows(ws: Any, config: dict[str, Any]) -> dict[str, Any]:
    settings = config.get("pickup_date_expansion") or {}
    if not settings.get("enabled"):
        return {"enabled": False, "source_row_count": 0, "expanded_row_count": 0}

    time_zone = str(settings.get("time_zone") or "Europe/Warsaw")
    start_date = resolve_config_date(settings.get("start_date", "today"), time_zone)
    if settings.get("months_ahead") is not None:
        end_date = add_calendar_months(start_date, int(settings["months_ahead"]))
    else:
        end_date = resolve_config_date(settings.get("end_date", "2027-01-31"), time_zone)
    if start_date > end_date:
        raise ValueError(
            f"Pickup date expansion start {start_date.isoformat()} is after end {end_date.isoformat()}; Sheet1 was not modified."
        )
    columns = config["columns"]
    data_start_row = int(config["data_start_row"])
    group_col = int(columns["group"])
    zone_col = int(columns["zone"])
    pickup_start_col = int(columns["pickup_start_date"])
    pickup_end_col = int(columns["pickup_end_date"])
    max_col = ws.max_column
    source_rows: list[tuple[dict[str, Any], date, date, str, str]] = []
    row_segments: list[tuple[str, Any]] = []
    template_rows_by_group_zone: dict[tuple[str, str], dict[str, Any]] = {}
    template_dates_by_group_zone: dict[tuple[str, str], date] = {}
    included_group_zones: set[tuple[str, str]] = set()
    preserved_out_of_range_row_count = 0
    dropped_before_start_row_count = 0
    dropped_after_end_row_count = 0
    drop_rows_before_start_date = bool(settings.get("drop_rows_before_start_date"))
    drop_rows_after_end_date = bool(settings.get("drop_rows_after_end_date"))

    for row in range(data_start_row, ws.max_row + 1):
        group = ws.cell(row, group_col).value
        zone = ws.cell(row, zone_col).value
        normalized_group = normalize_code(group)
        normalized_zone = normalize_code(zone)
        pickup_start = parse_date_value(ws.cell(row, pickup_start_col).value)
        pickup_end = parse_date_value(ws.cell(row, pickup_end_col).value) or pickup_start
        if not group and not zone and pickup_start is None:
            continue
        if not group or not zone or pickup_start is None:
            continue
        if pickup_end is None or pickup_end < pickup_start:
            pickup_end = pickup_start

        row_snapshot = snapshot_row(ws, row, max_col)
        group_zone_key = (normalized_group, normalized_zone)
        if pickup_start >= template_dates_by_group_zone.get(group_zone_key, date.min):
            template_rows_by_group_zone[group_zone_key] = row_snapshot
            template_dates_by_group_zone[group_zone_key] = pickup_start

        if pickup_start < start_date and drop_rows_before_start_date:
            dropped_before_start_row_count += 1
            continue
        if pickup_start > end_date and drop_rows_after_end_date:
            dropped_after_end_row_count += 1
            continue
        if pickup_end < start_date or pickup_start > end_date:
            row_segments.append(("preserve", row_snapshot))
            preserved_out_of_range_row_count += 1
            continue

        clipped_start = max(pickup_start, start_date)
        clipped_end = min(pickup_end, end_date)
        if clipped_start > clipped_end:
            continue
        source_row = (row_snapshot, clipped_start, clipped_end, normalized_group, normalized_zone)
        source_rows.append(source_row)
        row_segments.append(("expand", source_row))
        included_group_zones.add(group_zone_key)

    restored_group_zones: list[str] = []
    for (group, zone), row_snapshot in sorted(template_rows_by_group_zone.items()):
        if (group, zone) in included_group_zones:
            continue
        source_row = (row_snapshot, start_date, end_date, group, zone)
        source_rows.append(source_row)
        row_segments.append(("expand", source_row))
        included_group_zones.add((group, zone))
        restored_group_zones.append(f"{group}/{zone}")

    expanded_rows: list[tuple[dict[str, Any], date, str, str]] = []
    output_rows: list[tuple[dict[str, Any], date | None]] = []
    expanded_groups: set[str] = set()
    expanded_group_zones: set[tuple[str, str]] = set()
    for segment_type, segment in row_segments:
        if segment_type == "preserve":
            output_rows.append((segment, None))
            continue

        row_snapshot, clipped_start, clipped_end, group, zone = segment
        expanded_groups.add(group)
        expanded_group_zones.add((group, zone))
        for pickup_date in iter_dates_inclusive(clipped_start, clipped_end):
            expanded_rows.append((row_snapshot, pickup_date, group, zone))
            output_rows.append((row_snapshot, pickup_date))

    appended_horizon_rows: list[tuple[date, str, str, dict[str, Any]]] = []
    latest_source_date = max(template_dates_by_group_zone.values())
    if start_date <= latest_source_date < end_date:
        append_start_date = latest_source_date + timedelta(days=1)
        for group_zone_key in sorted(template_rows_by_group_zone):
            if template_dates_by_group_zone[group_zone_key] != latest_source_date:
                continue
            group, zone = group_zone_key
            row_snapshot = template_rows_by_group_zone[group_zone_key]
            for pickup_date in iter_dates_inclusive(append_start_date, end_date):
                appended_horizon_rows.append((pickup_date, group, zone, row_snapshot))

    for pickup_date, group, zone, row_snapshot in appended_horizon_rows:
        expanded_rows.append((row_snapshot, pickup_date, group, zone))
        output_rows.append((row_snapshot, pickup_date))

    if not template_rows_by_group_zone:
        raise ValueError("Pickup date expansion found no valid Group + Zone source rows; Sheet1 was not modified.")
    if not expanded_rows:
        raise ValueError("Pickup date expansion produced no rows; Sheet1 was not modified.")

    if ws.max_row >= data_start_row:
        ws.delete_rows(data_start_row, ws.max_row - data_start_row + 1)

    for row_index, (row_snapshot, pickup_date) in enumerate(output_rows, start=data_start_row):
        write_row_snapshot(ws, row_index, row_snapshot)
        if pickup_date is None:
            continue
        start_template = row_snapshot["cells"][pickup_start_col - 1]["value"]
        end_template = row_snapshot["cells"][pickup_end_col - 1]["value"]
        ws.cell(row_index, pickup_start_col).value = format_pickup_date_like_template(pickup_date, start_template)
        ws.cell(row_index, pickup_end_col).value = format_pickup_date_like_template(pickup_date, end_template)

    return {
        "enabled": True,
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "source_row_count": len(source_rows),
        "expanded_row_count": len(expanded_rows),
        "preserved_out_of_range_row_count": preserved_out_of_range_row_count,
        "dropped_before_start_row_count": dropped_before_start_row_count,
        "dropped_after_end_row_count": dropped_after_end_row_count,
        "output_row_count": len(output_rows),
        "source_groups": sorted({group for group, _zone in template_rows_by_group_zone}),
        "expanded_groups": sorted(expanded_groups),
        "missing_source_groups_after_expansion": sorted(
            {group for group, _zone in template_rows_by_group_zone} - expanded_groups
        ),
        "source_group_zones": sorted(f"{group}/{zone}" for group, zone in template_rows_by_group_zone),
        "missing_source_group_zones_after_expansion": sorted(
            f"{group}/{zone}"
            for group, zone in set(template_rows_by_group_zone) - expanded_group_zones
        ),
        "restored_group_zone_count": len(restored_group_zones),
        "restored_group_zones": restored_group_zones[:50],
        "appended_horizon_row_count": len(appended_horizon_rows),
    }


def get_pickup_row_duration_days(pickup_start: date | None, pickup_end: date | None) -> int | None:
    if pickup_start is None or pickup_end is None:
        return None
    duration = (pickup_end - pickup_start).days
    return duration if duration > 0 else None


def find_targets_for_row(
    targets_by_date: dict[date, list[dict[str, Any]]],
    pickup_start: date | None,
    pickup_end: date | None = None,
    match_pickup_end_duration: bool = False,
) -> list[dict[str, Any]]:
    if pickup_start is None:
        return []
    row_targets = targets_by_date.get(pickup_start, [])
    if not match_pickup_end_duration:
        return row_targets

    row_duration = get_pickup_row_duration_days(pickup_start, pickup_end)
    if row_duration is None:
        return []
    return [
        target
        for target in row_targets
        if int(parse_number(target.get("rental_days")) or 0) == row_duration
    ]



def maybe_normalize_pickup_end_date(ws: Any, row: int, columns: dict[str, Any], dry_run: bool) -> bool:
    start_cell = ws.cell(row, int(columns["pickup_start_date"]))
    end_cell = ws.cell(row, int(columns["pickup_end_date"]))
    if start_cell.value in (None, ""):
        return False
    if end_cell.value == start_cell.value:
        return False

    start_date = parse_date_value(start_cell.value)
    end_date = parse_date_value(end_cell.value)
    if start_date is not None and end_date == start_date and str(end_cell.value) == str(start_cell.value):
        return False

    if not dry_run:
        end_cell.value = start_cell.value
        end_cell.number_format = start_cell.number_format
    return True


def maybe_sync_booking_end_to_pickup_end(ws: Any, row: int, columns: dict[str, Any], dry_run: bool) -> bool:
    booking_end_col = columns.get("booking_end_date")
    pickup_end_col = columns.get("pickup_end_date")
    if not booking_end_col or not pickup_end_col:
        return False

    booking_end_cell = ws.cell(row, int(booking_end_col))
    pickup_end_cell = ws.cell(row, int(pickup_end_col))
    if pickup_end_cell.value in (None, ""):
        return False
    if booking_end_cell.value == pickup_end_cell.value:
        return False

    if not dry_run:
        booking_end_cell.value = pickup_end_cell.value
        booking_end_cell.number_format = pickup_end_cell.number_format
    return True


def get_import_row_limit(config: dict[str, Any]) -> int:
    configured_max = int(config.get("max_import_rows") or 0)
    if configured_max <= 0:
        raise ValueError("max_import_rows must be a positive integer.")
    return min(configured_max, BROKER_IMPORT_ROW_LIMIT)


def validate_import_row_limit(ws: Any, config: dict[str, Any]) -> int:
    max_rows = get_import_row_limit(config)
    if ws.max_row > max_rows:
        raise ValueError(
            f"Import workbook row limit exceeded: Sheet1 has {ws.max_row} rows, "
            f"but the configured maximum is {max_rows}. No Excel files were saved."
        )
    return max_rows


def save_import_ready_workbook(workbook: Any, sheet_name: str, output_path: Path) -> None:
    for sheet in list(workbook.worksheets):
        if sheet.title != sheet_name:
            workbook.remove(sheet)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def enforce_group_price_parity(
    ws: Any,
    config: dict[str, Any],
    duration_columns: dict[int, tuple[int, str, int, int]],
    dry_run: bool,
    scope: set[tuple[str, date, int]] | None = None,
) -> list[dict[str, Any]]:
    parity = get_group_price_parity(config)
    if parity is None:
        return []

    base_groups, premium_adjustments = parity
    tracked_groups = set(base_groups) | set(premium_adjustments)
    columns = config["columns"]
    data_start_row = int(config["data_start_row"])
    group_col = int(columns["group"])
    zone_col = int(columns["zone"])
    pickup_col = int(columns["pickup_start_date"])
    rate_cols = sorted({value[0] for value in duration_columns.values()})
    duration_band_by_col = {
        value[0]: (value[1], value[2], value[3])
        for value in duration_columns.values()
    }
    protected_periods = get_protected_rate_periods(config)
    min_change = float(config.get("min_excel_change_pln_day", 0.01))
    rows_by_key: dict[tuple[str, date], dict[str, int]] = defaultdict(dict)

    for row in range(data_start_row, ws.max_row + 1):
        group = normalize_code(ws.cell(row, group_col).value)
        if group not in tracked_groups:
            continue
        zone = normalize_code(ws.cell(row, zone_col).value)
        pickup_start = parse_date_value(ws.cell(row, pickup_col).value)
        if not zone or pickup_start is None:
            continue
        rows_by_key[(zone, pickup_start)][group] = row

    changes: list[dict[str, Any]] = []
    for (zone, pickup_start), groups in rows_by_key.items():
        if date_is_rate_protected(pickup_start, protected_periods):
            continue
        for col in rate_cols:
            if scope is not None and (zone, pickup_start, col) not in scope:
                continue
            base_rate = None
            for group in base_groups:
                row = groups.get(group)
                if row is None:
                    continue
                base_rate = parse_number(ws.cell(row, col).value)
                if base_rate is not None:
                    break
            if base_rate is None:
                continue

            target_rates = {group: base_rate for group in base_groups}
            target_rates.update({
                group: base_rate + adjustment
                for group, adjustment in premium_adjustments.items()
            })

            for group, target_rate in target_rates.items():
                row = groups.get(group)
                if row is None:
                    continue
                cell = ws.cell(row, col)
                old_rate = parse_number(cell.value)
                if old_rate is not None and abs(target_rate - old_rate) < min_change:
                    continue

                duration_band, duration_min_days, duration_max_days = duration_band_by_col[col]
                change = {
                    "action": classify_actual_action(old_rate, target_rate, "increase"),
                    "recommendation_action": "parity",
                    "recommendation_type": "group_parity",
                    "target_rank": "",
                    "reason": "Ujednolicenie stawek grup bazowych oraz korekta skonfigurowanych grup premium.",
                    "location": "",
                    "zone": zone,
                    "group": group,
                    "pickup_date": pickup_start.isoformat(),
                    "duration_days": None,
                    "duration_band": duration_band,
                    "duration_min_days": duration_min_days,
                    "duration_max_days": duration_max_days,
                    "cell": cell.coordinate,
                    "old_rate": old_rate,
                    "new_rate": target_rate,
                    "delta": None if old_rate is None else round(target_rate - old_rate, 2),
                    "minimum_rate_pln_day": 0,
                    "minimum_reason": "",
                    "group_adjustment_pln_day": premium_adjustments.get(group, 0),
                    "target_achievable": True,
                    "source_decision_count": 0,
                }
                changes.append(change)
                if dry_run:
                    continue

                cell.value = int(target_rate) if float(target_rate).is_integer() else round(target_rate, 2)
                cell.fill = get_delta_fill(
                    change,
                    config,
                )
                cell.comment = build_rate_comment(change)

    return changes


def highlight_excluded_group_rates(
    ws: Any,
    row: int,
    config: dict[str, Any],
    duration_columns: dict[int, tuple[int, str, int, int]],
    dry_run: bool,
    scoped_rate_cols: set[int] | None = None,
) -> int:
    group = ws.cell(row, int(config["columns"]["group"])).value
    threshold = get_excluded_group_highlight_threshold(group, config)
    if threshold is None:
        return 0

    fill = PatternFill(fill_type="solid", fgColor=str((config.get("colors") or {}).get("limited", "FCE4D6")))
    highlighted = 0
    for col, _, _, _ in duration_columns.values():
        if scoped_rate_cols is not None and col not in scoped_rate_cols:
            continue
        rate = parse_number(ws.cell(row, col).value)
        if rate is None or rate >= threshold:
            continue
        highlighted += 1
        if not dry_run:
            ws.cell(row, col).fill = copy(fill)
    return highlighted


def apply_updates(
    workbook_path: Path,
    recommendations_path: Path,
    output_path: Path | None,
    config: dict[str, Any],
    cli_groups: str | None,
    dry_run: bool,
    accepted_only: bool = False,
    acceptance_workbook_path: Path | None = None,
    import_output_path: Path | None = None,
) -> dict[str, Any]:
    input_workbook_sha256 = sha256_file(workbook_path)
    baseline_confirmation = load_baseline_confirmation(config, input_workbook_sha256)
    allowed_groups = resolve_apply_groups(config, cli_groups)
    recommendations = load_recommendation_items(recommendations_path)

    workbook = openpyxl.load_workbook(workbook_path)
    sheet_name = config.get("worksheet") or workbook.sheetnames[0]
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Worksheet '{sheet_name}' not found. Available sheets: {', '.join(workbook.sheetnames)}")
    ws = workbook[sheet_name]

    expansion_summary = expand_pickup_date_rows(ws, config)
    recommendations, recommendation_out_of_pickup_range_count = filter_recommendations_to_pickup_date_range(
        recommendations,
        config,
        expansion_summary,
    )
    recommendation_scope_dates = get_recommendation_scope_dates(recommendations)
    match_pickup_end_duration = False
    duration_columns = get_duration_columns(ws, config)
    fixed_rate_group_summary = ensure_fixed_rate_group_rows(ws, config, duration_columns, dry_run)
    mirrored_rate_group_summary = ensure_mirrored_rate_group_rows(ws, config, duration_columns, dry_run)
    targets, skipped_targets = build_targets(recommendations, duration_columns, config)
    accepted_target_count = 0
    filtered_unaccepted_target_count = 0
    if accepted_only:
        accepted_keys = (
            load_acceptance_keys(
                acceptance_workbook_path,
                str(config.get("recommendations_review_sheet") or "Recommendations Review"),
            )
            if acceptance_workbook_path
            else set()
        )
        targets, accepted_target_count, filtered_unaccepted_target_count = filter_targets_by_acceptance(targets, accepted_keys)
    city_top1_airport_caps, missing_airport_references = build_city_top1_airport_rate_caps(ws, config, targets)
    if missing_airport_references:
        details = "; ".join(
            f"{item['city_zone']}/{item['pickup_date']} {item['duration_band']} -> "
            + ",".join(item["airport_zones"])
            for item in missing_airport_references[:8]
        )
        raise ValueError(
            "Missing airport rate required by city Top1 cap; no workbook changes were saved. "
            + details
        )
    group_price_parity_scope = build_group_price_parity_scope(targets)
    columns = config["columns"]
    data_start_row = int(config["data_start_row"])
    min_change = float(config.get("min_excel_change_pln_day", 0.01))
    changes: list[dict[str, Any]] = []
    normalized_pickup_end_count = 0
    synced_booking_end_count = 0
    group_price_parity_changes: list[dict[str, Any]] = []
    excluded_group_highlight_count = 0
    city_top1_airport_cap_applied_count = 0
    protected_periods = get_protected_rate_periods(config)

    for row in range(data_start_row, ws.max_row + 1):
        if config.get("normalize_pickup_end_to_start", True) and not match_pickup_end_duration:
            if maybe_normalize_pickup_end_date(ws, row, columns, dry_run):
                normalized_pickup_end_count += 1
        if maybe_sync_booking_end_to_pickup_end(ws, row, columns, dry_run):
            synced_booking_end_count += 1

        zone = normalize_code(ws.cell(row, int(columns["zone"])).value)
        if zone not in targets:
            continue

        group = ws.cell(row, int(columns["group"])).value
        pickup_start = parse_date_value(ws.cell(row, int(columns["pickup_start_date"])).value)
        pickup_end = parse_date_value(ws.cell(row, int(columns["pickup_end_date"])).value)
        if date_is_rate_protected(pickup_start, protected_periods):
            continue
        row_targets = find_targets_for_row(
            targets[zone],
            pickup_start,
            pickup_end,
            match_pickup_end_duration,
        )
        if not row_targets:
            continue

        if group_is_excluded(group, config):
            excluded_group_highlight_count += highlight_excluded_group_rates(
                ws,
                row,
                config,
                duration_columns,
                dry_run,
                scoped_rate_cols={int(target["rate_col"]) for target in row_targets},
            )
            continue

        if not group_is_allowed(group, allowed_groups):
            continue

        for target in row_targets:
            cell = ws.cell(row, int(target["rate_col"]))
            old_rate = parse_number(cell.value)
            group_adjustment = get_group_rate_adjustment(group, config)
            current_base_equivalent = None if old_rate is None else old_rate - group_adjustment
            base_rate, suggested_rate, minimum_rate, minimum_reason = calculate_target_base_rate(
                target,
                current_base_equivalent,
                config,
            )
            uncapped_base_rate = base_rate
            cap_info = city_top1_airport_caps.get((zone, target["target_date"], int(target["rate_col"])))
            city_airport_cap_active = cap_info is not None
            city_airport_cap_applied = False
            city_airport_cap_conflict = False
            city_airport_max_rate = None
            airport_reference_rate = None
            if cap_info is not None:
                base_rate_cap = float(cap_info["base_rate_cap_pln_day"])
                city_airport_max_rate = round(base_rate_cap + group_adjustment, 2)
                airport_reference_rate = (cap_info.get("airport_rates_by_group") or {}).get(normalize_code(group))
                if minimum_rate > base_rate_cap + 0.001:
                    city_airport_cap_conflict = True
                    raise ValueError(
                        "City Top1 airport cap conflicts with the configured price floor; no workbook changes were saved. "
                        f"{normalize_code(group)}/{zone}/{target['target_date'].isoformat()} "
                        f"{target['duration_band']}: floor {format_rate_for_comment(minimum_rate)} > "
                        f"cap {format_rate_for_comment(base_rate_cap)}."
                    )
                if base_rate > base_rate_cap:
                    base_rate = base_rate_cap
                    city_airport_cap_applied = True
                    city_top1_airport_cap_applied_count += 1
            new_rate = base_rate + group_adjustment
            if old_rate is not None and abs(new_rate - old_rate) < min_change:
                continue

            actual_action = classify_actual_action(old_rate, new_rate, str(target["action"]))
            constraint_evaluation = evaluate_target_constraints(target, new_rate)
            broker_markup_multiplier = parse_number(target.get("broker_markup_multiplier")) or 1
            predicted_site_rate = round(new_rate * broker_markup_multiplier, 2)
            change = {
                "action": actual_action,
                "recommendation_action": target["action"],
                "recommendation_type": target.get("recommendation_type", ""),
                "target_rank": target.get("target_rank", ""),
                "reason": target.get("reason", ""),
                "location": target.get("location", ""),
                "zone": zone,
                "group": normalize_code(group),
                "pickup_date": target["target_date"].isoformat(),
                "duration_days": target.get("rental_days"),
                "duration_band": target["duration_band"],
                "duration_min_days": target.get("duration_min_days"),
                "duration_max_days": target.get("duration_max_days"),
                "cell": cell.coordinate,
                "old_rate": old_rate,
                "new_rate": new_rate,
                "delta": None if old_rate is None else round(new_rate - old_rate, 2),
                "suggested_rate_before_minimum": suggested_rate,
                "suggested_rate_before_airport_cap": uncapped_base_rate + group_adjustment,
                "site_target_rate": target.get("site_cap_rate_pln_day") or target.get("site_target_rate_pln_day"),
                "predicted_site_rate": predicted_site_rate,
                "broker_markup_multiplier": broker_markup_multiplier,
                "broker_markup_percent": target.get("broker_markup_percent"),
                "broker_markup_source": target.get("broker_markup_source", ""),
                "minimum_rate_pln_day": minimum_rate,
                "minimum_reason": minimum_reason if base_rate > suggested_rate else "",
                "group_adjustment_pln_day": group_adjustment,
                "city_top1_airport_cap_active": city_airport_cap_active,
                "city_top1_airport_cap_applied": city_airport_cap_applied,
                "city_top1_airport_cap_conflict": city_airport_cap_conflict,
                "city_top1_airport_max_multiplier": cap_info.get("max_multiplier") if cap_info else None,
                "city_top1_airport_max_rate": city_airport_max_rate,
                "airport_reference_zone": cap_info.get("airport_zone", "") if cap_info else "",
                "airport_reference_location": cap_info.get("airport_location", "") if cap_info else "",
                "airport_reference_rate": airport_reference_rate,
                "currency": target.get("currency", ""),
                "mm_rank": target.get("mm_rank", ""),
                "mm_provider": target.get("mm_provider", ""),
                "mm_rate": target.get("mm_rate_pln_day"),
                "top1_provider": target.get("top1_provider", ""),
                "top1_rate": target.get("top1_rate_pln_day"),
                "top2_provider": target.get("top2_provider", ""),
                "top2_rate": target.get("top2_rate_pln_day"),
                "top3_provider": target.get("top3_provider", ""),
                "top3_rate": target.get("top3_rate_pln_day"),
                "benchmark_provider": target.get("benchmark_provider", ""),
                "benchmark_rate": target.get("benchmark_rate_pln_day"),
                "dropoff_date": target.get("dropoff_date", ""),
                "source_generated_at": target.get("source_generated_at", ""),
                "scenario_id": target.get("scenario_id", ""),
                "source_decision_count": target.get("source_decision_count", 1),
                "source_active_count": target.get("source_active_count", 1),
                "source_actions": target.get("source_actions", []),
                "source_recommendation_types": target.get("source_recommendation_types", []),
                "source_locations": target.get("source_locations", []),
                "controlling_duration_days": target.get("controlling_duration_days"),
                "aggregation_conflict": bool(target.get("aggregation_conflict")),
                "covered_duration_days": target.get("covered_duration_days", []),
                "missing_duration_days": target.get("missing_duration_days", []),
                "duration_band_coverage_complete": target.get("duration_band_coverage_complete", True),
                **constraint_evaluation,
            }

            if not dry_run:
                cell.value = int(new_rate) if float(new_rate).is_integer() else round(new_rate, 2)
                cell.fill = get_delta_fill(change, config)
                cell.comment = build_rate_comment(change)

            changes.append(change)

    group_price_parity_changes = enforce_group_price_parity(
        ws,
        config,
        duration_columns,
        dry_run,
        scope=group_price_parity_scope,
    )
    changes.extend(group_price_parity_changes)
    changes_outside_recommendation_scope = sorted({
        str(change.get("pickup_date") or "")
        for change in changes
        if parse_date_value(change.get("pickup_date")) not in recommendation_scope_dates
    })
    if changes_outside_recommendation_scope:
        raise ValueError(
            "Excel changes fall outside the recommendation date scope; no Excel files were saved. "
            + "; ".join(changes_outside_recommendation_scope[:8])
        )
    validation_rows = build_validation_rows(
        ws,
        config,
        duration_columns,
        changes,
        skipped_targets,
        expansion_summary,
        assume_fixed_rate_changes_applied=dry_run,
    )

    city_top1_airport_cap_violations = (
        [] if dry_run else find_city_top1_airport_cap_violations(ws, config, city_top1_airport_caps)
    )
    if city_top1_airport_cap_violations:
        raise ValueError(
            "Final rates violate the city Top1 airport cap after group parity; no workbook changes were saved. "
            + "; ".join(city_top1_airport_cap_violations[:8])
        )

    if not dry_run:
        if output_path is None:
            raise ValueError("Output path is required unless --dry-run is used.")
        validate_import_row_limit(ws, config)
        write_changed_positions_sheet(workbook, ws, config, changes)
        write_recommendations_review_sheet(workbook, ws, config, changes)
        write_competitor_evidence_sheet(workbook, ws, config, changes)
        write_validation_sheet(
            workbook,
            ws,
            config,
            duration_columns,
            changes,
            skipped_targets,
            expansion_summary,
            validation_rows,
        )
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        if import_output_path is not None:
            save_import_ready_workbook(workbook, sheet_name, import_output_path)

    return {
        "workbook": str(workbook_path),
        "input_workbook_sha256": input_workbook_sha256,
        "baseline_confirmation": baseline_confirmation,
        "output": str(output_path) if output_path else None,
        "import_output": str(import_output_path) if import_output_path else None,
        "max_import_rows": get_import_row_limit(config),
        "import_row_count": ws.max_row,
        "recommendation_date_scope": {
            "start_date": min(recommendation_scope_dates).isoformat() if recommendation_scope_dates else None,
            "end_date": max(recommendation_scope_dates).isoformat() if recommendation_scope_dates else None,
            "date_count": len(recommendation_scope_dates),
        },
        "recommendation_out_of_pickup_range_count": recommendation_out_of_pickup_range_count,
        "dry_run": dry_run,
        "change_count": len(changes),
        "change_statistics": build_change_statistics(changes),
        "group_price_parity_change_count": len(group_price_parity_changes),
        "group_price_parity_scope_count": len(group_price_parity_scope),
        "city_top1_airport_cap_scope_count": len(city_top1_airport_caps),
        "city_top1_airport_cap_applied_count": city_top1_airport_cap_applied_count,
        "city_top1_airport_cap_violation_count": len(city_top1_airport_cap_violations),
        "excluded_group_highlight_count": excluded_group_highlight_count,
        "normalized_pickup_end_count": normalized_pickup_end_count,
        "synced_booking_end_count": synced_booking_end_count,
        "pickup_date_expansion": expansion_summary,
        "fixed_rate_groups": fixed_rate_group_summary,
        "mirrored_rate_groups": mirrored_rate_group_summary,
        "skipped_target_count": len(skipped_targets),
        "accepted_only": accepted_only,
        "accepted_target_count": accepted_target_count,
        "filtered_unaccepted_target_count": filtered_unaccepted_target_count,
        "validation": [
            {
                "check": row[0],
                "status": row[1],
                "issue_count": row[2],
                "details": row[3],
            }
            for row in validation_rows
        ],
        "broker_markup_observations": build_broker_markup_observations(
            changes,
            config,
            baseline_confirmation,
        ),
        "changes": changes[:100],
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Update rental-rate Excel workbook from pricing recommendations.")
    parser.add_argument("--workbook", required=True, help="Input .xlsx workbook path.")
    parser.add_argument("--recommendations", required=True, help="pricing-recommendations.json path.")
    parser.add_argument("--config", required=True, help="Excel rate update config JSON path.")
    parser.add_argument("--output", help="Output .xlsx path.")
    parser.add_argument("--import-output", help="Optional clean import .xlsx path containing only Sheet1.")
    parser.add_argument("--groups", help="Comma-separated car groups to update, or 'all'. Overrides config apply_groups.")
    parser.add_argument("--accepted-only", action="store_true", help="Apply only recommendations marked as accepted.")
    parser.add_argument("--acceptance-workbook", help="Workbook containing a Recommendations Review sheet with Akceptacja?/Accept? decisions.")
    parser.add_argument("--dry-run", action="store_true", help="Calculate matching changes without saving an .xlsx file.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    config = load_config(Path(args.config))
    summary = apply_updates(
        workbook_path=Path(args.workbook),
        recommendations_path=Path(args.recommendations),
        output_path=Path(args.output) if args.output else None,
        config=config,
        cli_groups=args.groups,
        dry_run=args.dry_run,
        accepted_only=args.accepted_only,
        acceptance_workbook_path=Path(args.acceptance_workbook) if args.acceptance_workbook else None,
        import_output_path=Path(args.import_output) if args.import_output else None,
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
